from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.modules.auctions.catalog.brand_normalizer import canonical_brand
from app.modules.auctions.price_loaders.base import BasePriceLoader
from app.modules.auctions.price_loaders.models import OurCategory, PriceRow

logger = logging.getLogger(__name__)


_PRINTER_CATEGORIES: set[str] = {
    "Печатная техника",
}


_IGNORE_FIRST: set[str] = {
    "тумба", "комплект", "лоток", "модуль", "автоподатчик",
    "дополнительный", "сканер", "стенд", "крышка", "податчик",
}

_ADJECTIVE_FIRST: set[str] = {
    "цветное", "цветной", "лазерное", "лазерный",
    "монохромное", "монохромный", "струйное", "струйный",
}


def _classify_resursmedia(name: str) -> OurCategory:
    parts = name.split()
    first = parts[0].lower() if parts else ""
    second = parts[1].lower() if len(parts) > 1 else ""

    if first == "мфу":
        return "mfu"
    if first == "принтер":
        return "printer"
    if first == "плоттер":
        return "printer"
    if first == "фабрика":
        return "printer"

    if first in _IGNORE_FIRST:
        return "ignore"

    if first in _ADJECTIVE_FIRST:
        if second == "мфу":
            return "mfu"
        if second in ("принтер", "плоттер"):
            return "printer"
        logger.info(
            "ResursMedia %s: classified as ignore (first=%r, second=%r)",
            name, first, second,
        )
        return "ignore"

    if second == "мфу":
        return "mfu"
    if second in ("принтер", "плоттер"):
        return "printer"
    if second in _ADJECTIVE_FIRST:
        third = parts[2].lower() if len(parts) > 2 else ""
        if third == "мфу":
            return "mfu"
        if third in ("принтер", "плоттер"):
            return "printer"

    logger.info(
        "ResursMedia %s: classified as ignore (first=%r, second=%r)",
        name, first, second,
    )
    return "ignore"


_COL_NUM = 0
_COL_CATEGORY = 1
_COL_BRAND = 2
_COL_ARTICLE = 3
_COL_MPN = 4
_COL_NAME = 5
_COL_VOLUME = 6
_COL_WEIGHT = 7
_COL_PRICE_USD = 8
_COL_PRICE_RUB = 9
_COL_STOCK_MSK = 10
_COL_STOCK_FACT = 11
_COL_TRANSIT = 12

HEADER_ROW = 2
DATA_START_ROW = 3


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _normalize(s) -> str:
    return str(s).strip() if s is not None else ""


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


_RESURS_QUAL_STOCK: dict[str, int] = {
    "нет": 0,
    "мало": 5,
    "много": 50,
    "ожидается": 0,
}


def _parse_stock(value) -> int:
    if value is None:
        return 0
    s = str(value).strip().lower()
    if not s:
        return 0
    if s in _RESURS_QUAL_STOCK:
        return _RESURS_QUAL_STOCK[s]
    return _parse_int(value)


class ResursMediaPriceLoader(BasePriceLoader):
    supplier_code = "resursmedia"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "resurs" in name or "ресурс" in name


    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Price" if "Price" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            category = _normalize(_cell(row, _COL_CATEGORY))
            if category not in _PRINTER_CATEGORIES:
                continue

            brand = canonical_brand(_normalize(_cell(row, _COL_BRAND))) or None
            supplier_sku = _normalize(_cell(row, _COL_ARTICLE))
            mpn = _normalize(_cell(row, _COL_MPN)) or None
            name = _normalize(_cell(row, _COL_NAME))
            price_rub = _parse_price(_cell(row, _COL_PRICE_RUB))
            price_usd = _parse_price(_cell(row, _COL_PRICE_USD))
            stock_msk = _parse_stock(_cell(row, _COL_STOCK_MSK))
            stock_fact = _parse_stock(_cell(row, _COL_STOCK_FACT))
            transit = _parse_stock(_cell(row, _COL_TRANSIT))

            if not supplier_sku and not name:
                continue
            if not supplier_sku:
                logger.warning(
                    "Resurs-Media строка %d: пустой Артикул — пропущена.", row_idx
                )
                continue

            if price_rub is not None:
                price = price_rub
                currency = "RUB"
            elif price_usd is not None:
                price = price_usd
                currency = "USD"
            else:
                continue

            stock = max(stock_msk, stock_fact)

            yield PriceRow(
                supplier_sku=supplier_sku,
                mpn=mpn,
                gtin=None,
                brand=brand,
                raw_category=category,
                name=name,
                price=price,
                currency=currency,
                stock=stock,
                transit=transit,
                our_category=_classify_resursmedia(name),
                row_number=row_idx,
            )
