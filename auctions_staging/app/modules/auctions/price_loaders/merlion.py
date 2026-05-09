from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.modules.auctions.catalog.brand_normalizer import canonical_brand
from app.modules.auctions.price_loaders.base import BasePriceLoader
from app.modules.auctions.price_loaders.models import OurCategory, PriceRow

logger = logging.getLogger(__name__)


_PRINTER_GROUPS: set[tuple[str, str]] = {
    ("Периферия и аксессуары", "Принтеры"),
}


_G3_CATEGORY_MAP: dict[str, OurCategory] = {
    "МФУ лазерные": "mfu",
    "Лазерные": "printer",
    "МФУ струйные": "mfu",
    "Струйные": "printer",
    "Термопринтеры": "ignore",
    "Мини-Фото-принтеры": "ignore",
    "Матричные": "ignore",
    "": "ignore",
}


def _classify_merlion(g3: str) -> OurCategory:
    if g3 in _G3_CATEGORY_MAP:
        return _G3_CATEGORY_MAP[g3]
    logger.info("Merlion G3=%r: classified as ignore (unknown subcategory)", g3)
    return "ignore"


_COL_GROUP_1 = 0
_COL_GROUP_2 = 1
_COL_GROUP_3 = 2
_COL_BRAND = 3
_COL_NUMBER = 4
_COL_MPN = 6
_COL_NAME = 7
_COL_PRICE_USD = 9
_COL_PRICE_RUB = 10
_COL_STOCK = 11
_COL_TRANSIT_1 = 12
_COL_TRANSIT_2 = 13

HEADER_ROW = 11
DATA_START_ROW = 12


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


_MERLION_QUAL_STOCK: dict[str, int] = {
    "+": 5,
    "++": 15,
    "+++": 50,
    "++++": 100,
    "call": 0,
}


def _parse_stock(value) -> int:
    if value is None:
        return 0
    s = str(value).strip().lower()
    if not s:
        return 0
    if s in _MERLION_QUAL_STOCK:
        return _MERLION_QUAL_STOCK[s]
    return _parse_int(value)


def _normalize(s) -> str:
    return str(s).strip() if s is not None else ""


def _build_raw_path(g1: str, g2: str, g3: str) -> str:
    return " | ".join(x for x in (g1, g2, g3) if x)


def _is_printer_group(g1: str, g2: str) -> bool:
    return (g1, g2) in _PRINTER_GROUPS


class MerlionPriceLoader(BasePriceLoader):
    supplier_code = "merlion"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "merlion" in name or "мерлион" in name

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Price List"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист «{sheet_name}» не найден в файле {filepath}. "
                f"Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            g1 = _normalize(_cell(row, _COL_GROUP_1))
            g2 = _normalize(_cell(row, _COL_GROUP_2))
            g3 = _normalize(_cell(row, _COL_GROUP_3))

            if not _is_printer_group(g1, g2):
                continue

            brand = canonical_brand(_normalize(_cell(row, _COL_BRAND))) or None
            supplier_sku = _normalize(_cell(row, _COL_NUMBER))
            mpn = _normalize(_cell(row, _COL_MPN)) or None
            name = _normalize(_cell(row, _COL_NAME))
            price_usd = _parse_price(_cell(row, _COL_PRICE_USD))
            price_rub = _parse_price(_cell(row, _COL_PRICE_RUB))
            stock = _parse_stock(_cell(row, _COL_STOCK))
            transit_1 = _parse_stock(_cell(row, _COL_TRANSIT_1))
            transit_2 = _parse_stock(_cell(row, _COL_TRANSIT_2))

            if not supplier_sku and not name:
                continue
            if not supplier_sku:
                logger.warning(
                    "Merlion строка %d: пустой «Номер» — пропущена.", row_idx
                )
                continue

            if price_rub is not None:
                price = price_rub
                currency = "RUB"
            elif price_usd is not None:
                price = price_usd
                currency = "USD"
            else:
                continue

            yield PriceRow(
                supplier_sku=supplier_sku,
                mpn=mpn,
                gtin=None,
                brand=brand,
                raw_category=_build_raw_path(g1, g2, g3),
                name=name,
                price=price,
                currency=currency,
                stock=stock,
                transit=transit_1 + transit_2,
                our_category=_classify_merlion(g3),
                row_number=row_idx,
            )
