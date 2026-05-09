from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.modules.auctions.catalog.brand_normalizer import canonical_brand
from app.modules.auctions.price_loaders.base import BasePriceLoader
from app.modules.auctions.price_loaders.models import OurCategory, PriceRow

logger = logging.getLogger(__name__)


_PRINTER_CATEGORIES_B: set[str] = {
    "Принтеры",
    "МФУ",
}


_BC_CATEGORY_MAP: dict[tuple[str, str], OurCategory] = {
    ("Принтеры", "Принтеры лазерные"): "printer",
    ("Принтеры", "Принтеры струйные"): "printer",
    ("Принтеры", "Принтеры матричные"): "ignore",
    ("МФУ", "МФУ лазерные"): "mfu",
    ("МФУ", "МФУ струйные"): "mfu",
    ("МФУ", "МФУ матричные"): "ignore",
}


def _classify_ocs(cat_b: str, kind_c: str) -> OurCategory:
    key = (cat_b, kind_c)
    if key in _BC_CATEGORY_MAP:
        return _BC_CATEGORY_MAP[key]
    logger.info(
        "OCS (B=%r, C=%r): classified as ignore (unknown combination)", cat_b, kind_c
    )
    return "ignore"


_COL_CAT_B = 1
_COL_KIND_C = 2
_COL_MAKER = 3
_COL_SUPPLIER_SKU = 4
_COL_SKU = 6
_COL_NAME = 7
_COL_PRICE = 8
_COL_CURRENCY = 9
_COL_STOCK = 11
_COL_TRANSIT = 17


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


def _normalize_gtin(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if "e" in s.lower():
        try:
            s = str(int(Decimal(s)))
        except InvalidOperation:
            return None
    digits = re.sub(r"\D", "", s)
    return digits or None


def _find_ean_column(header_row: tuple) -> int | None:
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().upper().replace(" ", "")
        if name in {"EAN128", "EAN", "GTIN"}:
            return idx
    return None


class OcsPriceLoader(BasePriceLoader):
    supplier_code = "ocs"

    @classmethod
    def detect(cls, filename: str) -> bool:
        return "ocs" in filename.lower()

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Наличие и цены"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист «{sheet_name}» не найден в файле {filepath}. "
                f"Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ean_idx = _find_ean_column(header)
        if ean_idx is None:
            logger.info("OCS: колонка EAN128 не найдена — GTIN не будет заполняться.")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            cat_b = str(_cell(row, _COL_CAT_B) or "").strip()
            kind_c = str(_cell(row, _COL_KIND_C) or "").strip()

            if cat_b not in _PRINTER_CATEGORIES_B:
                continue

            manufacturer = canonical_brand(str(_cell(row, _COL_MAKER) or "").strip()) or None
            supplier_sku = str(_cell(row, _COL_SUPPLIER_SKU) or "").strip()
            mpn = str(_cell(row, _COL_SKU) or "").strip()
            name = str(_cell(row, _COL_NAME) or "").strip()
            price_raw = _cell(row, _COL_PRICE)
            currency_raw = _cell(row, _COL_CURRENCY)
            stock_raw = _cell(row, _COL_STOCK)
            transit_raw = _cell(row, _COL_TRANSIT)
            ean_raw = _cell(row, ean_idx) if ean_idx is not None else None

            price = _parse_price(price_raw)
            if price is None:
                continue

            currency = (str(currency_raw).strip().upper() if currency_raw else "RUB")[:3] or "RUB"

            if not mpn:
                continue

            yield PriceRow(
                supplier_sku=supplier_sku or "",
                mpn=mpn,
                gtin=_normalize_gtin(ean_raw),
                brand=manufacturer,
                raw_category=" | ".join(x for x in (cat_b, kind_c) if x),
                name=name,
                price=price,
                currency=currency,
                stock=_parse_int(stock_raw),
                transit=_parse_int(transit_raw),
                our_category=_classify_ocs(cat_b, kind_c),
                row_number=row_idx,
            )
