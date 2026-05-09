from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.modules.auctions.catalog.brand_normalizer import canonical_brand
from app.modules.auctions.price_loaders.base import BasePriceLoader
from app.modules.auctions.price_loaders.models import OurCategory, PriceRow

logger = logging.getLogger(__name__)


_PRINTER_PATH_PREFIXES: tuple[str, ...] = (
    "Принтеры, сканеры, МФУ->Принтеры->",
    "Принтеры, сканеры, МФУ->МФУ->",
    "Принтеры, сканеры, МФУ->Широкоформатные Принтеры",
    "Принтеры, сканеры, МФУ->Широкоформатные МФУ",
)


_TREOLAN_ROOT = "Принтеры, сканеры, МФУ"
_WIDE_PRINTER_PREFIXES: tuple[str, ...] = (
    f"{_TREOLAN_ROOT}->Широкоформатные Принтеры",
    f"{_TREOLAN_ROOT}->Широкоформатные Принтеры/Плоттеры",
)
_WIDE_MFU_PREFIX = f"{_TREOLAN_ROOT}->Широкоформатные МФУ"


def _classify_treolan(path: str) -> OurCategory:
    if any(path.startswith(p) for p in _WIDE_PRINTER_PREFIXES):
        return "printer"
    if path.startswith(_WIDE_MFU_PREFIX):
        return "mfu"
    parts = path.split("->")
    if len(parts) >= 2 and parts[0] == _TREOLAN_ROOT:
        middle = parts[1]
        if middle == "Принтеры":
            return "printer"
        if middle == "МФУ":
            return "mfu"
    logger.info("Treolan path=%r: classified as ignore", path)
    return "ignore"


_COL_ARTICLE = 0
_COL_NAME = 1
_COL_BRAND = 2
_COL_STOCK = 3
_COL_TRANSIT_1 = 4
_COL_TRANSIT_2 = 5
_COL_PRICE_USD = 6
_COL_PRICE_RUB = 7
_COL_GTIN = 9

HEADER_ROW = 3
DATA_START_ROW = 4


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


_TREOLAN_QUAL_STOCK: dict[str, int] = {
    "<10": 5,
    "много": 50,
    ">10": 20,
    ">100": 100,
}


def _parse_stock(value) -> int:
    if value is None:
        return 0
    s = str(value).strip().lower().replace(" ", "")
    if not s:
        return 0
    if s in _TREOLAN_QUAL_STOCK:
        return _TREOLAN_QUAL_STOCK[s]
    return _parse_int(value)


def _normalize(s) -> str:
    return str(s).strip() if s is not None else ""


def _normalize_gtin(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if "e" in s.lower():
        try:
            s = str(int(Decimal(s)))
        except InvalidOperation:
            return None
    digits = re.sub(r"\D", "", s)
    return digits or None


def _is_category_separator(row: tuple) -> str | None:
    a_val = _cell(row, _COL_ARTICLE)
    if a_val is None:
        return None
    a_str = str(a_val).strip()
    if not a_str or "->" not in a_str:
        return None
    for idx in (_COL_NAME, _COL_BRAND, _COL_PRICE_USD, _COL_PRICE_RUB):
        if _normalize(_cell(row, idx)):
            return None
    return a_str


def _path_is_printer(path: str) -> bool:
    return any(path.startswith(p) for p in _PRINTER_PATH_PREFIXES)


class TreolanPriceLoader(BasePriceLoader):
    supplier_code = "treolan"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "treolan" in name or "catalog" in name

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Каталог"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист «{sheet_name}» не найден в файле {filepath}. "
                f"Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        current_raw_category: str = ""

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            sep = _is_category_separator(row)
            if sep is not None:
                current_raw_category = sep
                continue

            if not _path_is_printer(current_raw_category):
                continue

            article = _normalize(_cell(row, _COL_ARTICLE))
            name = _normalize(_cell(row, _COL_NAME))
            brand = canonical_brand(_normalize(_cell(row, _COL_BRAND))) or None
            stock = _parse_stock(_cell(row, _COL_STOCK))
            transit = _parse_stock(_cell(row, _COL_TRANSIT_1)) + _parse_stock(
                _cell(row, _COL_TRANSIT_2)
            )
            price_usd = _parse_price(_cell(row, _COL_PRICE_USD))
            price_rub = _parse_price(_cell(row, _COL_PRICE_RUB))
            gtin = _normalize_gtin(_cell(row, _COL_GTIN))

            if not article and not name:
                continue
            if not article:
                logger.warning("Treolan строка %d: пустой артикул — пропущена.", row_idx)
                continue

            if price_rub is not None:
                price = price_rub
                currency = "RUB"
            elif price_usd is not None:
                price = price_usd
                currency = "USD"
            else:
                continue

            yield PriceRow(
                supplier_sku=article,
                mpn=article,
                gtin=gtin,
                brand=brand,
                raw_category=current_raw_category,
                name=name,
                price=price,
                currency=currency,
                stock=stock,
                transit=transit,
                our_category=_classify_treolan(current_raw_category),
                row_number=row_idx,
            )
