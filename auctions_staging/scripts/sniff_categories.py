"""Diagnostic: extract unique category values from each price file (after printer/mfu pre-filter)."""
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

SEED_DIR = Path(__file__).resolve().parent.parent / ".business" / "seed" / "supplier_prices"


def _norm(v):
    return str(v).strip() if v is not None else ""


def sniff_merlion(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Price List"]
    g3_counter = Counter()
    g2_counter = Counter()
    for row in ws.iter_rows(min_row=12, values_only=True):
        if not row:
            continue
        g1, g2, g3 = _norm(row[0]), _norm(row[1]), _norm(row[2])
        if (g1, g2) != ("Периферия и аксессуары", "Принтеры"):
            continue
        g2_counter[(g1, g2)] += 1
        g3_counter[g3] += 1
    return {"g2_after_prefilter": g2_counter, "g3_after_prefilter": g3_counter}


def sniff_ocs(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Наличие и цены"]
    cat_b_counter = Counter()
    kind_c_counter = Counter()
    pair_counter = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        cat_b = _norm(row[1])
        kind_c = _norm(row[2])
        if cat_b not in {"Принтеры", "МФУ"}:
            continue
        cat_b_counter[cat_b] += 1
        kind_c_counter[kind_c] += 1
        pair_counter[(cat_b, kind_c)] += 1
    return {
        "cat_b_after_prefilter": cat_b_counter,
        "kind_c_after_prefilter": kind_c_counter,
        "pairs": pair_counter,
    }


def sniff_treolan(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Каталог"]
    path_counter = Counter()
    current_path = ""
    PRINTER_PREFIXES = (
        "Принтеры, сканеры, МФУ->Принтеры->",
        "Принтеры, сканеры, МФУ->МФУ->",
        "Принтеры, сканеры, МФУ->Широкоформатные Принтеры",
        "Принтеры, сканеры, МФУ->Широкоформатные МФУ",
    )
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        a = _norm(row[0])
        if a and "->" in a and not any(_norm(row[i]) for i in (1, 2, 6, 7)):
            current_path = a
            continue
        if not any(current_path.startswith(p) for p in PRINTER_PREFIXES):
            continue
        path_counter[current_path] += 1
    return {"path_after_prefilter": path_counter}


def sniff_resursmedia(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Price"] if "Price" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())

    cat_counter = Counter()
    subcat_counter = Counter()  # колонка 0 (А)
    name_first_word = Counter()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        category = _norm(row[1]) if len(row) > 1 else ""
        if category != "Печатная техника":
            continue
        cat_counter[category] += 1
        subcat_counter[_norm(row[0]) if len(row) > 0 else ""] += 1
        name = _norm(row[5]) if len(row) > 5 else ""
        if name:
            name_first_word[name.split()[0] if name.split() else ""] += 1
    return {
        "category_after_prefilter": cat_counter,
        "subcat_col_a": subcat_counter,
        "name_first_word": name_first_word,
        "headers": headers,
    }


def main():
    print("=== MERLION ===")
    r = sniff_merlion(SEED_DIR / "merlion_price.xlsm")
    print(f"После фильтра (g1='Периферия и аксессуары', g2='Принтеры'):")
    print(f"  Всего: {sum(r['g3_after_prefilter'].values())} строк")
    print(f"  Уникальные значения G3 (Группа 3):")
    for v, c in r["g3_after_prefilter"].most_common():
        print(f"    {c:4}  «{v}»")

    print("\n=== OCS ===")
    r = sniff_ocs(SEED_DIR / "ocs_price.xlsx")
    print(f"После фильтра (cat_b in {{Принтеры, МФУ}}):")
    print(f"  Всего: {sum(r['cat_b_after_prefilter'].values())} строк")
    print(f"  Распределение по B (Категория):")
    for v, c in r["cat_b_after_prefilter"].most_common():
        print(f"    {c:4}  «{v}»")
    print(f"  Распределение по C (Тип) внутри них:")
    for v, c in r["kind_c_after_prefilter"].most_common():
        print(f"    {c:4}  «{v}»")
    print(f"  Пары (B, C):")
    for (b, c), n in r["pairs"].most_common():
        print(f"    {n:4}  B=«{b}» C=«{c}»")

    print("\n=== TREOLAN ===")
    r = sniff_treolan(SEED_DIR / "treolan_catalog.xlsx")
    print(f"После фильтра (путь начинается с одного из printer-prefixes):")
    print(f"  Всего: {sum(r['path_after_prefilter'].values())} строк")
    print(f"  Уникальные пути:")
    for v, c in r["path_after_prefilter"].most_common():
        print(f"    {c:4}  «{v}»")

    print("\n=== RESURS-MEDIA ===")
    r = sniff_resursmedia(SEED_DIR / "priceresurs.xlsx")
    print(f"После фильтра (B='Печатная техника'):")
    print(f"  Всего: {sum(r['category_after_prefilter'].values())} строк")
    print(f"  Headers row 2 (для контекста колонок):")
    for i, h in enumerate(r["headers"][:13]):
        print(f"    [{i}] {h}")
    print(f"  Колонка A (col 0) — может быть подкатегорией:")
    for v, n in r["subcat_col_a"].most_common():
        print(f"    {n:4}  «{v}»")
    print(f"  Первое слово имени модели (топ-15):")
    for w, n in r["name_first_word"].most_common(15):
        print(f"    {n:4}  «{w}»")


if __name__ == "__main__":
    main()
