# Excel-выгрузка списка аукционных лотов (Backlog #12, Фаза 2 плана
# 2026-05-13-auctions-excel-export.md).
#
# Read-only экспорт: менеджер открывает /auctions, настраивает фильтры,
# нажимает «Скачать Excel» — получает файл с теми лотами, которые сейчас
# видит в инбоксе. Главный сценарий — сортировка по «Маржа, %» в Excel,
# чтобы найти жирные позиции (порог 15% решено оставить, фильтрация
# делается менеджером в файле).
#
# Гранулярность B (решение собственника 2026-05-16):
# 1 строка Excel = 1 tender_item + primary match (если есть). Tender-meta
# (reg_number, заказчик, регион, статус) дублируется в каждой строке лота.
# Это позволяет сортировать по марже % и видеть, какому лоту принадлежит
# позиция, без перекрёстных ссылок.
#
# Структура листа (паттерн повторяет catalog/excel_export.py):
#   Строка 1: A1 = «Курс ЦБ (USD→RUB)», B1 = число (LATEST из
#     exchange_rates, fallback 90.0). Курс декоративный — все cost/price
#     в БД уже в RUB; формулы маржи опираются на ячейки cost/price в той
#     же строке, не на $B$1. Курс оставлен для консистентности с
#     каталог-экспортом и будущего расширения, если cheapest_supplier
#     USD-цены попадут в выгрузку.
#   Строка 2: пусто, разделитель.
#   Строка 3: заголовки, autofilter.
#   Строка 4..N: данные.
#
# Маржа RUB и % — формулы Excel. Это даёт менеджеру возможность вручную
# поправить cost/price в файле и сразу увидеть новый margin, плюс
# Excel-сортировка по марже % работает на формулах.
#
# Lim cap: 10 000 tender_items за один экспорт. На текущем prod-объёме
# ≈800 позиций cap не достигается; защита от случайного «скачать всё»
# при гипотетическом росте. При cap_reached=True флаг попадает в audit.

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from portal.services.auctions_service import InboxFilters, STATUS_LABELS
from shared.db import SessionLocal


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------
# Стили и константы листа
# ---------------------------------------------------------------------

_FILL_RATE = PatternFill("solid", fgColor="DDEBF7")        # светло-голубой
_FILL_FORMULA = PatternFill("solid", fgColor="FFF4CE")     # жёлтый — маржа-формула
_FILL_HEADER = PatternFill("solid", fgColor="E7E6E6")      # серый — шапка
_FONT_HEADER = Font(bold=True)
_FONT_LINK = Font(color="0000EE", underline="single")
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Fallback курс при пустой exchange_rates — совпадает с #11
# (portal/services/catalog/excel_export.py:84).
_FALLBACK_RATE = Decimal("90.0000")

_RATE_LABEL_CELL = "A1"
_RATE_VALUE_CELL = "B1"
_RATE_ABS_REF = "$B$1"

_HEADER_ROW = 3
_DATA_START_ROW = 4

# Cap на количество строк (tender_items) за один экспорт. SQL запрашивает
# CAP+1 строк, чтобы детектить превышение и выставить флаг cap_reached.
_ROW_CAP = 10_000

# Имя листа и шаблон имени файла.
_SHEET_TITLE = "Аукционы"


# ---------------------------------------------------------------------
# Описание колонок (порядок — фиксирован)
# ---------------------------------------------------------------------

@dataclass(frozen=True)
class _Col:
    title: str
    width: int                       # ширина колонки в Excel-единицах
    hidden: bool = False
    is_formula: bool = False         # подсветить жёлтым (формула)


# Порядок и набор колонок Excel-листа. Изменение порядка — допустимо;
# индексы вычисляются по титулам через _col_index(), формулы маржи
# тоже опираются на _col_index().
_COLUMNS: tuple[_Col, ...] = (
    _Col("id",                       8, hidden=True),       # 1: tender_items.id (стабильный ключ, для аудита)
    _Col("№ извещения",              22),                   # 2
    _Col("Карточка zakupki",         18),                   # 3 — hyperlink
    _Col("Заказчик",                 42),                   # 4
    _Col("Регион",                   22),                   # 5
    _Col("Статус",                   14),                   # 6
    _Col("Дата публикации",          18),                   # 7
    _Col("Дедлайн подачи",           18),                   # 8
    _Col("Дата поставки",            18),                   # 9
    _Col("НМЦК лота, ₽",             14),                   # 10
    _Col("KTRU лота",                24),                   # 11
    _Col("№ позиции",                 8),                   # 12
    _Col("Название позиции",         42),                   # 13
    _Col("KTRU позиции",             22),                   # 14
    _Col("Количество",               10),                   # 15
    _Col("Цена за единицу, ₽",       14),                   # 16
    _Col("Бренд SKU",                14),                   # 17
    _Col("Артикул SKU",              18),                   # 18
    _Col("Название SKU",             42),                   # 19
    _Col("Cost base, ₽",             14),                   # 20
    _Col("Поставщик (cheapest)",     22),                   # 21
    _Col("Маржа, ₽",                 14, is_formula=True),  # 22 — формула
    _Col("Маржа, %",                 10, is_formula=True),  # 23 — формула
    _Col("Флаги",                    24),                   # 24
    _Col("Дата ингеста",             18),                   # 25
    _Col("Изменён",                  18),                   # 26
)


def _col_index(title: str) -> int:
    """1-based Excel column index по заголовку колонки."""
    for i, c in enumerate(_COLUMNS, start=1):
        if c.title == title:
            return i
    raise KeyError(f"Колонка не найдена: {title!r}")


def _col_letter(title: str) -> str:
    """Excel-буква колонки (A, B, ..., AA, ...) по заголовку."""
    return get_column_letter(_col_index(title))


# ---------------------------------------------------------------------
# Отчёт о выгрузке
# ---------------------------------------------------------------------

@dataclass
class ExportReport:
    file_path: Path
    rows_count: int = 0
    rate_used: Decimal = _FALLBACK_RATE
    rate_date: date | None = None
    rate_is_fallback: bool = True
    cap_reached: bool = False
    filter_summary: dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------
# Утилиты
# ---------------------------------------------------------------------

def _read_latest_rate(db: Session) -> tuple[Decimal, date | None, bool]:
    """LATEST курс из exchange_rates. Если таблица пустая — fallback 90.0
    (как в catalog/excel_export.py:_read_latest_rate)."""
    row = db.execute(
        text(
            "SELECT rate_usd_rub, rate_date "
            "FROM exchange_rates "
            "ORDER BY rate_date DESC, fetched_at DESC LIMIT 1"
        )
    ).first()
    if row is None:
        return _FALLBACK_RATE, None, True
    return Decimal(str(row.rate_usd_rub)), row.rate_date, False


def _to_msk(dt: datetime | None) -> datetime | None:
    """UTC/naive → Europe/Moscow. Naive интерпретируется как UTC."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    try:
        from zoneinfo import ZoneInfo
        return dt.astimezone(ZoneInfo("Europe/Moscow"))
    except Exception:
        return dt.astimezone(timezone(timedelta(hours=3)))


def _array_to_str(arr: Any) -> str:
    """TEXT[] → строка через запятую без пробелов; None/[] → ''."""
    if not arr:
        return ""
    if isinstance(arr, (list, tuple)):
        return ",".join(str(x).strip() for x in arr if str(x) and str(x).strip())
    return str(arr)


def _flags_to_str(flags: Any) -> str:
    """flags_jsonb → ключи через запятую, где значение truthy.
    Bool-значения берём как есть, остальные через bool(). Пустой dict → ''."""
    if not flags or not isinstance(flags, dict):
        return ""
    return ",".join(str(k) for k, v in flags.items() if v)


def _filter_summary(filters: InboxFilters) -> dict[str, Any]:
    """Сериализация фильтров для audit_log payload."""
    return {
        "statuses":                 list(filters.statuses) if filters.statuses else [],
        "nmck_min":                 str(filters.nmck_min) if filters.nmck_min is not None else None,
        "nmck_max":                 str(filters.nmck_max) if filters.nmck_max is not None else None,
        "search":                   filters.search or "",
        "urgent_only":              bool(filters.urgent_only),
        "print_only":               bool(filters.print_only),
        "include_excluded_regions": bool(filters.include_excluded_regions),
    }


# ---------------------------------------------------------------------
# SQL — основной запрос
# ---------------------------------------------------------------------

# Стратегия:
#   * Один запрос tenders ⋈ tender_items ⋈ tender_status ⋈ matches(primary)
#     ⋈ printers_mfu, плюс CTE items_breakdown для фильтра print_only
#     (повторяет логику auctions_service._INBOX_SQL).
#   * cheapest_supplier — коррелированный subquery по supplier_prices с
#     stock_qty > 0 и активным поставщиком. Возвращает имя поставщика с
#     min ценой для (printer|mfu).
#   * Фильтры — те же 7 полей, что в InboxFilters. urgent_only/нужен
#     deadline_alert_hours; передаётся параметром (читается из settings
#     роутом, не сервисом — у сервиса задача стабильная по фильтрам, а
#     порог дедлайна — настраиваемый).
#   * LIMIT — :limit (передаём CAP+1, чтобы понять, не урезали ли).
#   * ORDER BY publish_date DESC NULLS LAST, reg_number, position_num —
#     свежие лоты сверху, позиции одного лота — подряд.
_EXPORT_SQL = """
WITH items_breakdown AS (
    SELECT ti.tender_id,
           COUNT(*) AS total_cnt,
           COUNT(*) FILTER (
               WHERE ti.ktru_code LIKE '26.20.18.000-%%'
                  OR ti.ktru_code LIKE '26.20.16.120-%%'
           ) AS printer_cnt
      FROM tender_items ti
     GROUP BY ti.tender_id
)
SELECT
    ti.id                                                AS tender_item_id,
    t.reg_number                                         AS reg_number,
    t.url                                                AS url,
    t.customer                                           AS customer,
    t.customer_region                                    AS customer_region,
    COALESCE(ts.status, 'new')                           AS status,
    t.publish_date                                       AS publish_date,
    t.submit_deadline                                    AS submit_deadline,
    t.delivery_deadline                                  AS delivery_deadline,
    t.nmck_total                                         AS nmck_total,
    t.ktru_codes_array                                   AS ktru_codes_array,
    ti.position_num                                      AS position_num,
    ti.name                                              AS item_name,
    ti.ktru_code                                         AS ktru_code,
    ti.qty                                               AS qty,
    ti.nmck_per_unit                                     AS price_per_unit,
    pmu.brand                                            AS brand,
    pmu.sku                                              AS sku,
    pmu.name                                             AS sku_name,
    pmu.cost_base_rub                                    AS cost_base,
    (
        SELECT s.name
          FROM supplier_prices sp
          JOIN suppliers s ON s.id = sp.supplier_id
         WHERE sp.component_id = pmu.id
           AND sp.category IN ('printer', 'mfu')
           AND sp.stock_qty > 0
           AND s.is_active = TRUE
         ORDER BY sp.price ASC
         LIMIT 1
    )                                                    AS cheapest_supplier,
    t.flags_jsonb                                        AS flags_jsonb,
    t.ingested_at                                        AS ingested_at,
    t.last_modified_at                                   AS last_modified_at
  FROM tenders t
  JOIN tender_items ti           ON ti.tender_id = t.reg_number
  LEFT JOIN tender_status ts     ON ts.tender_id = t.reg_number
  LEFT JOIN matches m            ON m.tender_item_id = ti.id AND m.match_type = 'primary'
  LEFT JOIN printers_mfu pmu     ON pmu.id = m.nomenclature_id
  LEFT JOIN items_breakdown ib   ON ib.tender_id = t.reg_number
 WHERE (:has_status_filter = 0 OR COALESCE(ts.status, 'new') = ANY(CAST(:statuses AS text[])))
   AND (:nmck_min IS NULL OR t.nmck_total >= :nmck_min)
   AND (:nmck_max IS NULL OR t.nmck_total <= :nmck_max)
   AND (
        :search IS NULL
        OR t.reg_number      ILIKE :search_like
        OR t.customer        ILIKE :search_like
        OR t.customer_region ILIKE :search_like
       )
   AND (
        :print_only = 0
        OR (COALESCE(ib.total_cnt, 0) > 0
            AND COALESCE(ib.total_cnt, 0) = COALESCE(ib.printer_cnt, 0))
       )
   AND (
        :include_excluded_regions = 1
        OR NOT COALESCE((t.flags_jsonb->>'excluded_by_region')::boolean, false)
       )
   AND (
        :urgent_only = 0
        OR (
            t.submit_deadline IS NOT NULL
            AND t.submit_deadline > NOW()
            AND t.submit_deadline < NOW() + make_interval(hours => :deadline_alert_hours)
        )
       )
 ORDER BY t.publish_date DESC NULLS LAST, t.reg_number, ti.position_num
 LIMIT :limit
"""


def _execute_query(
    db: Session, filters: InboxFilters, *, deadline_alert_hours: int, limit: int,
) -> list[Any]:
    params: dict[str, Any] = {
        "has_status_filter":         1 if filters.statuses else 0,
        "statuses":                  list(filters.statuses) if filters.statuses else [],
        "nmck_min":                  filters.nmck_min,
        "nmck_max":                  filters.nmck_max,
        "search":                    filters.search,
        "search_like":               f"%{filters.search}%" if filters.search else None,
        "print_only":                1 if filters.print_only else 0,
        "include_excluded_regions":  1 if filters.include_excluded_regions else 0,
        "urgent_only":               1 if filters.urgent_only else 0,
        "deadline_alert_hours":      int(deadline_alert_hours),
        "limit":                     int(limit),
    }
    return list(db.execute(text(_EXPORT_SQL), params).all())


# ---------------------------------------------------------------------
# Запись листа
# ---------------------------------------------------------------------

def _write_header(ws: Worksheet, rate: Decimal) -> None:
    # Строка 1: курс
    ws[_RATE_LABEL_CELL] = "Курс ЦБ (USD→RUB)"
    ws[_RATE_LABEL_CELL].font = _FONT_HEADER
    ws[_RATE_LABEL_CELL].fill = _FILL_RATE
    ws[_RATE_VALUE_CELL] = float(rate)
    ws[_RATE_VALUE_CELL].number_format = "0.0000"
    ws[_RATE_VALUE_CELL].font = _FONT_HEADER
    ws[_RATE_VALUE_CELL].fill = _FILL_RATE

    # Строка 3: шапка таблицы
    for idx, col in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=idx, value=col.title)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_HEADER
        cell.fill = _FILL_HEADER
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = col.width
        if col.hidden:
            ws.column_dimensions[letter].hidden = True

    # Autofilter на шапку
    last_letter = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A{_HEADER_ROW}:{last_letter}{_HEADER_ROW}"

    # Freeze: при прокрутке шапка остаётся видимой.
    ws.freeze_panes = f"A{_DATA_START_ROW}"


def _write_data_row(ws: Worksheet, excel_row: int, row: Any) -> None:
    """Заполняет одну строку (1 tender_item) по результату SQL.

    `row` — SQLAlchemy Row из `_execute_query` (доступ через attrs).
    """
    # Простые data-ячейки. Set value через ws.cell(row,col).value = ...,
    # потому что для каждой ячейки мы сразу применяем number_format.

    def setv(title: str, value: Any) -> Any:
        c = ws.cell(row=excel_row, column=_col_index(title))
        if value is not None:
            c.value = value
        return c

    # Системные и meta
    setv("id", int(row.tender_item_id))
    setv("№ извещения", row.reg_number)

    # Hyperlink на карточку zakupki: текст «Открыть», url — в hyperlink-поле.
    url_cell = ws.cell(row=excel_row, column=_col_index("Карточка zakupki"))
    if row.url:
        url_cell.value = "Открыть"
        url_cell.hyperlink = row.url
        url_cell.font = _FONT_LINK

    setv("Заказчик", row.customer or "")
    setv("Регион", row.customer_region or "")
    setv("Статус", STATUS_LABELS.get(row.status, row.status))

    # Даты — МСК в формате dd.mm.yyyy HH:MM
    for title, dt in (
        ("Дата публикации",  row.publish_date),
        ("Дедлайн подачи",   row.submit_deadline),
        ("Дата поставки",    row.delivery_deadline),
        ("Дата ингеста",     row.ingested_at),
        ("Изменён",          row.last_modified_at),
    ):
        msk = _to_msk(dt)
        if msk is not None:
            c = setv(title, msk.replace(tzinfo=None))  # openpyxl не любит tz-aware
            c.number_format = "dd.mm.yyyy hh:mm"

    # Числовые поля
    nmck_total = setv("НМЦК лота, ₽", float(row.nmck_total) if row.nmck_total is not None else None)
    if row.nmck_total is not None:
        nmck_total.number_format = "0.00"

    setv("KTRU лота", _array_to_str(row.ktru_codes_array))

    setv("№ позиции", int(row.position_num) if row.position_num is not None else None)
    setv("Название позиции", row.item_name or "")
    setv("KTRU позиции", row.ktru_code or "")

    qty_cell = setv("Количество", float(row.qty) if row.qty is not None else None)
    if row.qty is not None:
        qty_cell.number_format = "0.###"

    price_cell = setv(
        "Цена за единицу, ₽",
        float(row.price_per_unit) if row.price_per_unit is not None else None,
    )
    if row.price_per_unit is not None:
        price_cell.number_format = "0.00"

    setv("Бренд SKU", row.brand or "")
    setv("Артикул SKU", row.sku or "")
    setv("Название SKU", row.sku_name or "")

    cost_cell = setv(
        "Cost base, ₽",
        float(row.cost_base) if row.cost_base is not None else None,
    )
    if row.cost_base is not None:
        cost_cell.number_format = "0.00"

    setv("Поставщик (cheapest)", row.cheapest_supplier or "")

    # Маржа RUB и % — формулы, опираются на ячейки той же строки.
    # Пишем формулу только если есть price и cost. Если чего-то не хватает —
    # ячейка остаётся пустой (менеджер видит «—»).
    has_cost = row.cost_base is not None and float(row.cost_base) > 0
    has_price = row.price_per_unit is not None
    has_qty = row.qty is not None and float(row.qty) > 0

    margin_rub_cell = ws.cell(row=excel_row, column=_col_index("Маржа, ₽"))
    margin_pct_cell = ws.cell(row=excel_row, column=_col_index("Маржа, %"))
    margin_rub_cell.fill = _FILL_FORMULA
    margin_pct_cell.fill = _FILL_FORMULA

    if has_cost and has_price and has_qty:
        price_ref = f"{_col_letter('Цена за единицу, ₽')}{excel_row}"
        cost_ref = f"{_col_letter('Cost base, ₽')}{excel_row}"
        qty_ref = f"{_col_letter('Количество')}{excel_row}"
        margin_rub_cell.value = f"=({price_ref}-{cost_ref})*{qty_ref}"
        margin_rub_cell.number_format = "0.00"
        margin_pct_cell.value = f"=({price_ref}-{cost_ref})/{cost_ref}"
        margin_pct_cell.number_format = "0.00%"

    setv("Флаги", _flags_to_str(row.flags_jsonb))


# ---------------------------------------------------------------------
# Публичная функция экспорта
# ---------------------------------------------------------------------

def export_auctions(
    output_path: Path,
    filters: InboxFilters,
    *,
    deadline_alert_hours: int = 24,
    db: Session | None = None,
) -> ExportReport:
    """Выгружает один Excel-файл с аукционными лотами под фильтры.

    Args:
        output_path: путь, куда сохранить .xlsx.
        filters: фильтры инбокса (тот же dataclass, что использует route
            `/auctions`).
        deadline_alert_hours: окно «срочно» в часах. Передаётся роутом
            из settings; для прямых вызовов используется дефолт 24.
        db: опциональная SQLAlchemy-сессия. Если None — открываем свою.

    Returns:
        ExportReport — путь к файлу, кол-во строк, использованный курс,
        флаги rate_is_fallback / cap_reached, сериализация фильтров.
    """
    if db is None:
        db = SessionLocal()
        try:
            return _build_workbook(
                db, output_path, filters, deadline_alert_hours=deadline_alert_hours,
            )
        finally:
            db.close()
    return _build_workbook(
        db, output_path, filters, deadline_alert_hours=deadline_alert_hours,
    )


def _build_workbook(
    db: Session,
    output_path: Path,
    filters: InboxFilters,
    *,
    deadline_alert_hours: int,
) -> ExportReport:
    rate, rate_date, is_fallback = _read_latest_rate(db)
    if is_fallback:
        logger.warning(
            "auctions excel-export: exchange_rates пуст, использую fallback "
            "%.4f RUB/USD.",
            float(rate),
        )

    # Запрашиваем CAP+1 — детектим cap_reached.
    rows = _execute_query(
        db, filters, deadline_alert_hours=deadline_alert_hours, limit=_ROW_CAP + 1,
    )
    cap_reached = len(rows) > _ROW_CAP
    if cap_reached:
        rows = rows[:_ROW_CAP]
        logger.warning(
            "auctions excel-export: достигнут cap %d строк, остальные обрезаны.",
            _ROW_CAP,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_TITLE

    _write_header(ws, rate)

    for offset, row in enumerate(rows):
        _write_data_row(ws, _DATA_START_ROW + offset, row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return ExportReport(
        file_path=output_path,
        rows_count=len(rows),
        rate_used=rate,
        rate_date=rate_date,
        rate_is_fallback=is_fallback,
        cap_reached=cap_reached,
        filter_summary=_filter_summary(filters),
    )


# ---------------------------------------------------------------------
# Имя файла по умолчанию
# ---------------------------------------------------------------------

def default_filename(today: date | None = None) -> str:
    """Имя файла «Аукционы_YYYY-MM-DD.xlsx»."""
    if today is None:
        today = date.today()
    return f"Аукционы_{today.strftime('%Y-%m-%d')}.xlsx"
