# Excel-выгрузка каталога товаров (Фаза 2 плана 2026-05-13).
#
# Два независимых файла:
#   * «Комплектующие_ПК.xlsx» — 8 листов (CPU/Motherboard/RAM/GPU/Storage/
#     Case/PSU/Cooler), по одному на таблицу комплектующих.
#   * «Печатная_техника.xlsx» — 2 листа (Принтеры/МФУ), оба читаются из
#     таблицы printers_mfu с фильтром category in ('printer','mfu').
#
# Структура каждого листа:
#   - Строка 1: служебная — A1 «Курс ЦБ (USD→RUB)», B1 — численное
#     значение курса (берётся LATEST из exchange_rates; если таблица
#     пуста — fallback 90.0 без обращения к ЦБ). B1 редактируемый: при
#     ручной правке RUB-формулы пересчитываются автоматически.
#   - Строка 2: пустая, разделитель.
#   - Строка 3: заголовки колонок. По шапке вешается autofilter.
#   - Строка 4+: данные.
#
# Категории колонок:
#   - hidden: внутренний id (column_dimensions[...].hidden = True). Это
#     ключ строки для importer'а Фазы 3.
#   - edit: редактируемая ячейка, белый фон.
#   - ro: read-only, жёлтая заливка (FFF4CE). Цены/поставщики/даты —
#     importer Фазы 3 их игнорирует.
#
# Цены:
#   - Цена min, USD — статика, минимум по supplier_prices(currency='USD')
#     среди активных поставщиков и активных позиций (stock_qty>0 OR
#     transit_qty>0). Если USD-предложений нет — ячейка пустая.
#   - Цена min, RUB — если есть USD-цена: формула «=<USD_cell>*$B$1»
#     (абсолютная ссылка на ячейку курса). Если есть только RUB —
#     статика. Если ни того ни другого — пустая.
#   - Поставщик (min) — имя того поставщика, чьё предложение дало min.
#     Приоритет USD-min над RUB-min (т.к. курс редактируем — единая
#     валюта сравнения).
#   - Цена обновлена — supplier_prices.updated_at у выбранного
#     предложения.
#
# Сериализация массивов (TEXT[], list в attrs_jsonb): в одной ячейке
# через запятую без пробелов, например «ATX,mATX,ITX». Пустые элементы
# отбрасываются.
#
# Пустые / NA значения:
#   - NULL в БД → пустая ячейка.
#   - Строка 'n/a' в attrs_jsonb → ячейка 'n/a' (маркер «искали —
#     не нашли»). Сохраняется как есть.

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from portal.services.auctions.catalog.enrichment.schema import (
    PRINTER_MFU_ATTRS,
    PRINTER_MFU_DIMENSION_ATTRS,
)
from shared.db import SessionLocal


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------

_FILL_RO = PatternFill("solid", fgColor="FFF4CE")
_FONT_HEADER = Font(bold=True)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Fallback курс на случай пустой таблицы exchange_rates (новый контейнер,
# тестовая БД без ЦБ-фикстуры и т.п.). Совпадает с фактическим fallback'ом
# в portal/services/databases/component_service.py:314 — 90.0 RUB/USD.
_FALLBACK_RATE = Decimal("90.0000")

# Позиция ячейки курса. Используется при генерации формул RUB-цен.
_RATE_LABEL_CELL = "A1"
_RATE_VALUE_CELL = "B1"
# Абсолютная ссылка на ячейку курса для формул в Excel.
_RATE_ABS_REF = "$B$1"

# Строка, на которой стоит шапка таблицы (autofilter крепится к ней).
_HEADER_ROW = 3

# Цвета фона.
_FILL_RATE = PatternFill("solid", fgColor="DDEBF7")  # светло-голубой
# Зрительно отделяет служебную строку 1 от данных.


# ---------------------------------------------------------------------
# Описание колонок листов
# ---------------------------------------------------------------------

# kind: 'hidden' | 'edit' | 'ro'
# source: 'col:<db_column>'      — значение колонки БД из основной таблицы.
#         'array:<db_column>'    — TEXT[]-колонка БД → через запятую.
#         'attr:<key>'           — ключ attrs_jsonb (printers_mfu).
#         'price:usd'            — min-цена в USD (ro).
#         'price:rub'            — min-цена в RUB: формула или статика (ro).
#         'price:supplier'       — имя поставщика min-предложения (ro).
#         'price:updated_at'     — updated_at min-предложения (ro).
#         'stock:on_hand'        — сумма stock_qty по активным предложениям (ro).
#         'stock:in_transit'     — сумма transit_qty по активным предложениям (ro).
#         'stock:suppliers'      — count(DISTINCT supplier) с положительным
#                                  наличием/транзитом (ro).
@dataclass(frozen=True)
class _Col:
    title: str
    kind: str
    source: str


# Общие колонки идентификации компонента ПК (id + базовые edit-поля).
_PC_COMMON_PREFIX = (
    _Col("id",           "hidden", "col:id"),
    _Col("model",        "edit",   "col:model"),
    _Col("manufacturer", "edit",   "col:manufacturer"),
    _Col("sku",          "edit",   "col:sku"),
    _Col("gtin",         "edit",   "col:gtin"),
    _Col("is_hidden",    "edit",   "col:is_hidden"),
)

# Общий «хвост» цен и наличия у всех листов ПК. Сначала price-блок
# (USD/RUB/Поставщик/Обновлена) — слитная группа; затем блок наличия
# (Склад/Транзит/Поставщиков) — отдельная группа. Подмножество поставщиков
# для наличия совпадает с подмножеством для min-цены (активные поставщики
# + активные позиции stock>0 OR transit>0), цифры читаются вместе с
# «Поставщик (min)» и не противоречат друг другу.
_PC_COMMON_PRICE_SUFFIX = (
    _Col("Цена min, USD",   "ro", "price:usd"),
    _Col("Цена min, RUB",   "ro", "price:rub"),
    _Col("Поставщик (min)", "ro", "price:supplier"),
    _Col("Цена обновлена",  "ro", "price:updated_at"),
    _Col("Склад, шт",       "ro", "stock:on_hand"),
    _Col("Транзит, шт",     "ro", "stock:in_transit"),
    _Col("Поставщиков, шт", "ro", "stock:suppliers"),
)


def _cpu_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("socket",                  "edit", "col:socket"),
        _Col("cores",                   "edit", "col:cores"),
        _Col("threads",                 "edit", "col:threads"),
        _Col("base_clock_ghz",          "edit", "col:base_clock_ghz"),
        _Col("turbo_clock_ghz",         "edit", "col:turbo_clock_ghz"),
        _Col("tdp_watts",               "edit", "col:tdp_watts"),
        _Col("has_integrated_graphics", "edit", "col:has_integrated_graphics"),
        _Col("memory_type",             "edit", "col:memory_type"),
        _Col("package_type",            "edit", "col:package_type"),
        _Col("process_nm",              "edit", "col:process_nm"),
        _Col("l3_cache_mb",             "edit", "col:l3_cache_mb"),
        _Col("max_memory_freq",         "edit", "col:max_memory_freq"),
        _Col("release_year",            "edit", "col:release_year"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _motherboard_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("socket",          "edit", "col:socket"),
        _Col("chipset",         "edit", "col:chipset"),
        _Col("form_factor",     "edit", "col:form_factor"),
        _Col("memory_type",     "edit", "col:memory_type"),
        _Col("has_m2_slot",     "edit", "col:has_m2_slot"),
        _Col("memory_slots",    "edit", "col:memory_slots"),
        _Col("max_memory_gb",   "edit", "col:max_memory_gb"),
        _Col("max_memory_freq", "edit", "col:max_memory_freq"),
        _Col("sata_ports",      "edit", "col:sata_ports"),
        _Col("m2_slots",        "edit", "col:m2_slots"),
        _Col("has_wifi",        "edit", "col:has_wifi"),
        _Col("has_bluetooth",   "edit", "col:has_bluetooth"),
        _Col("pcie_version",    "edit", "col:pcie_version"),
        _Col("pcie_x16_slots",  "edit", "col:pcie_x16_slots"),
        _Col("usb_ports",       "edit", "col:usb_ports"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _ram_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("memory_type",    "edit", "col:memory_type"),
        _Col("form_factor",    "edit", "col:form_factor"),
        _Col("module_size_gb", "edit", "col:module_size_gb"),
        _Col("modules_count",  "edit", "col:modules_count"),
        _Col("frequency_mhz",  "edit", "col:frequency_mhz"),
        _Col("cl_timing",      "edit", "col:cl_timing"),
        _Col("voltage",        "edit", "col:voltage"),
        _Col("has_heatsink",   "edit", "col:has_heatsink"),
        _Col("has_rgb",        "edit", "col:has_rgb"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _gpu_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("vram_gb",               "edit", "col:vram_gb"),
        _Col("vram_type",             "edit", "col:vram_type"),
        _Col("tdp_watts",             "edit", "col:tdp_watts"),
        _Col("needs_extra_power",     "edit", "col:needs_extra_power"),
        _Col("video_outputs",         "edit", "col:video_outputs"),
        _Col("core_clock_mhz",        "edit", "col:core_clock_mhz"),
        _Col("memory_clock_mhz",      "edit", "col:memory_clock_mhz"),
        _Col("gpu_chip",              "edit", "col:gpu_chip"),
        _Col("recommended_psu_watts", "edit", "col:recommended_psu_watts"),
        _Col("length_mm",             "edit", "col:length_mm"),
        _Col("height_mm",             "edit", "col:height_mm"),
        _Col("power_connectors",      "edit", "col:power_connectors"),
        _Col("fans_count",            "edit", "col:fans_count"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _storage_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("storage_type",   "edit", "col:storage_type"),
        _Col("form_factor",    "edit", "col:form_factor"),
        _Col("interface",      "edit", "col:interface"),
        _Col("capacity_gb",    "edit", "col:capacity_gb"),
        _Col("read_speed_mb",  "edit", "col:read_speed_mb"),
        _Col("write_speed_mb", "edit", "col:write_speed_mb"),
        _Col("tbw",            "edit", "col:tbw"),
        _Col("rpm",            "edit", "col:rpm"),
        _Col("cache_mb",       "edit", "col:cache_mb"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _case_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("supported_form_factors", "edit", "array:supported_form_factors"),
        _Col("has_psu_included",       "edit", "col:has_psu_included"),
        _Col("included_psu_watts",     "edit", "col:included_psu_watts"),
        _Col("max_gpu_length_mm",      "edit", "col:max_gpu_length_mm"),
        _Col("max_cooler_height_mm",   "edit", "col:max_cooler_height_mm"),
        _Col("psu_form_factor",        "edit", "col:psu_form_factor"),
        _Col("color",                  "edit", "col:color"),
        _Col("material",               "edit", "col:material"),
        _Col("drive_bays",             "edit", "col:drive_bays"),
        _Col("fans_included",          "edit", "col:fans_included"),
        _Col("has_glass_panel",        "edit", "col:has_glass_panel"),
        _Col("has_rgb",                "edit", "col:has_rgb"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _psu_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("power_watts",          "edit", "col:power_watts"),
        _Col("form_factor",          "edit", "col:form_factor"),
        _Col("efficiency_rating",    "edit", "col:efficiency_rating"),
        _Col("modularity",           "edit", "col:modularity"),
        _Col("has_12vhpwr",          "edit", "col:has_12vhpwr"),
        _Col("sata_connectors",      "edit", "col:sata_connectors"),
        _Col("main_cable_length_mm", "edit", "col:main_cable_length_mm"),
        _Col("warranty_years",       "edit", "col:warranty_years"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _cooler_columns() -> tuple[_Col, ...]:
    return (
        *_PC_COMMON_PREFIX,
        _Col("supported_sockets", "edit", "array:supported_sockets"),
        _Col("max_tdp_watts",     "edit", "col:max_tdp_watts"),
        _Col("cooler_type",       "edit", "col:cooler_type"),
        _Col("height_mm",         "edit", "col:height_mm"),
        _Col("radiator_size_mm",  "edit", "col:radiator_size_mm"),
        _Col("fans_count",        "edit", "col:fans_count"),
        _Col("noise_db",          "edit", "col:noise_db"),
        _Col("has_rgb",           "edit", "col:has_rgb"),
        *_PC_COMMON_PRICE_SUFFIX,
    )


def _printer_mfu_columns() -> tuple[_Col, ...]:
    """Колонки для листов «Принтеры» / «МФУ» (одинаковая структура).

    Источник полей attrs_jsonb — словарь PRINTER_MFU_ATTRS_ALL из
    enrichment/schema.py (общий со словарём ПЭК-логистики). Все 9 ключей
    PRINTER_MFU_ATTRS обязательные, 4 ключа PRINTER_MFU_DIMENSION_ATTRS —
    опциональные.
    """
    base: list[_Col] = [
        _Col("id",                "hidden", "col:id"),
        _Col("sku",               "edit",   "col:sku"),
        _Col("mpn",               "edit",   "col:mpn"),
        _Col("gtin",              "edit",   "col:gtin"),
        _Col("brand",             "edit",   "col:brand"),
        _Col("name",              "edit",   "col:name"),
        _Col("category",          "edit",   "col:category"),
        _Col("ktru_codes_array",  "edit",   "array:ktru_codes_array"),
        _Col("is_hidden",         "edit",   "col:is_hidden"),
        _Col("cost_base_rub",     "edit",   "col:cost_base_rub"),
        _Col("margin_pct_target", "edit",   "col:margin_pct_target"),
    ]
    for key in PRINTER_MFU_ATTRS:
        base.append(_Col(key, "edit", f"attr:{key}"))
    for key in PRINTER_MFU_DIMENSION_ATTRS:
        base.append(_Col(key, "edit", f"attr:{key}"))
    base.extend([
        _Col("attrs_source",    "ro", "col:attrs_source"),
        _Col("Цена min, USD",   "ro", "price:usd"),
        _Col("Цена min, RUB",   "ro", "price:rub"),
        _Col("Поставщик (min)", "ro", "price:supplier"),
        _Col("Цена обновлена",  "ro", "price:updated_at"),
        _Col("Склад, шт",       "ro", "stock:on_hand"),
        _Col("Транзит, шт",     "ro", "stock:in_transit"),
        _Col("Поставщиков, шт", "ro", "stock:suppliers"),
    ])
    return tuple(base)


@dataclass(frozen=True)
class _SheetSpec:
    """Описание листа: имя, источник данных и набор колонок."""
    title: str
    table: str               # таблица БД (cpus / motherboards / ... / printers_mfu)
    price_category: str      # значение supplier_prices.category для джойна
    columns: tuple[_Col, ...]
    extra_where: str = ""    # доп. фильтр в WHERE (для printers_mfu разделить
                             # 'printer' / 'mfu' по колонке category).


def _components_sheet_specs() -> tuple[_SheetSpec, ...]:
    return (
        _SheetSpec("CPU",         "cpus",         "cpu",         _cpu_columns()),
        _SheetSpec("Motherboard", "motherboards", "motherboard", _motherboard_columns()),
        _SheetSpec("RAM",         "rams",         "ram",         _ram_columns()),
        _SheetSpec("GPU",         "gpus",         "gpu",         _gpu_columns()),
        _SheetSpec("Storage",     "storages",     "storage",     _storage_columns()),
        _SheetSpec("Case",        "cases",        "case",        _case_columns()),
        _SheetSpec("PSU",         "psus",         "psu",         _psu_columns()),
        _SheetSpec("Cooler",      "coolers",      "cooler",      _cooler_columns()),
    )


def _printers_mfu_sheet_specs() -> tuple[_SheetSpec, ...]:
    cols = _printer_mfu_columns()
    return (
        _SheetSpec(
            "Принтеры", "printers_mfu", "printer", cols,
            extra_where="category = 'printer'",
        ),
        _SheetSpec(
            "МФУ", "printers_mfu", "mfu", cols,
            extra_where="category = 'mfu'",
        ),
    )


# ---------------------------------------------------------------------
# Отчёт о выгрузке
# ---------------------------------------------------------------------

@dataclass
class ExportReport:
    """Что было выгружено. Используется CLI/тестами для проверки."""
    file_path: Path
    sheet_counts: dict[str, int] = field(default_factory=dict)
    total_rows: int = 0
    rate_used: Decimal = _FALLBACK_RATE
    rate_date: date | None = None
    rate_is_fallback: bool = True


# ---------------------------------------------------------------------
# Курс ЦБ — читаем самый свежий из БД, без обращения в интернет
# ---------------------------------------------------------------------

def _read_latest_rate(db: Session) -> tuple[Decimal, date | None, bool]:
    """LATEST курс из exchange_rates. Возвращает (rate, rate_date, is_fallback).

    Если таблица пустая — fallback 90.0 (без обращения к ЦБ: экспорт
    должен работать офлайн в любой момент, в том числе на свежей БД).
    """
    row = db.execute(
        text(
            "SELECT rate_usd_rub, rate_date "
            "FROM exchange_rates "
            "ORDER BY rate_date DESC, fetched_at DESC LIMIT 1"
        )
    ).first()
    if row is None:
        return _FALLBACK_RATE, None, True
    return Decimal(str(row.rate_usd_rub)), row.rate_date, False


# ---------------------------------------------------------------------
# Чтение строк таблицы + min-цена
# ---------------------------------------------------------------------

def _fetch_main_rows(
    db: Session, spec: _SheetSpec, db_columns: list[str],
) -> list[dict[str, Any]]:
    """Читает все строки указанной таблицы (`spec.table`), отдаёт список
    словарей в формате column_name -> value. `db_columns` — какие колонки
    тащить из БД (минимально необходимое подмножество)."""
    cols_sql = ", ".join(db_columns)
    where = f" WHERE {spec.extra_where}" if spec.extra_where else ""
    sql = f"SELECT {cols_sql} FROM {spec.table}{where} ORDER BY id"
    rows = db.execute(text(sql)).mappings().all()
    return [dict(r) for r in rows]


def _fetch_min_prices(
    db: Session, category: str, component_ids: Sequence[int],
) -> dict[int, dict[str, Any]]:
    """Для каждого component_id из списка возвращает min-цену.

    Алгоритм:
      - Среди активных поставщиков и активных позиций (stock_qty > 0 OR
        transit_qty > 0) находим min по price для каждой валюты отдельно.
      - Имя поставщика — у того, чьё предложение дало min (приоритет USD).
      - Дата обновления — updated_at той же строки.

    Result-схема:
      {
        <component_id>: {
          'min_usd':       Decimal | None,
          'min_rub':       Decimal | None,
          'supplier_name': str | None,
          'updated_at':    datetime | None,
        }
      }
    Если компонент не имеет активных предложений — его в результате нет.
    """
    if not component_ids:
        return {}
    sql = text(
        """
        WITH active_prices AS (
            SELECT sp.component_id,
                   sp.currency,
                   sp.price,
                   sp.updated_at,
                   s.name AS supplier_name,
                   ROW_NUMBER() OVER (
                       PARTITION BY sp.component_id, sp.currency
                       ORDER BY sp.price ASC, sp.updated_at DESC
                   ) AS rn
              FROM supplier_prices sp
              JOIN suppliers s ON s.id = sp.supplier_id
             WHERE sp.category = :cat
               AND s.is_active = TRUE
               AND (sp.stock_qty > 0 OR sp.transit_qty > 0)
               AND sp.component_id = ANY(:ids)
        )
        SELECT component_id, currency, price, updated_at, supplier_name
          FROM active_prices
         WHERE rn = 1
        """
    )
    rows = db.execute(
        sql, {"cat": category, "ids": list(component_ids)},
    ).mappings().all()

    out: dict[int, dict[str, Any]] = {}
    for r in rows:
        cid = int(r["component_id"])
        slot = out.setdefault(cid, {
            "min_usd": None, "min_rub": None,
            "supplier_name": None, "updated_at": None,
        })
        currency = (r["currency"] or "").upper()
        price = Decimal(str(r["price"])) if r["price"] is not None else None
        if currency == "USD":
            slot["min_usd"] = price
            # USD-предложение имеет приоритет над RUB при выборе
            # «представительного» поставщика — единая шкала.
            slot["supplier_name"] = r["supplier_name"]
            slot["updated_at"] = r["updated_at"]
        else:
            slot["min_rub"] = price
            if slot["supplier_name"] is None:
                slot["supplier_name"] = r["supplier_name"]
                slot["updated_at"] = r["updated_at"]
    return out


def _fetch_availability(
    db: Session, category: str, component_ids: Sequence[int],
) -> dict[int, dict[str, int]]:
    """Агрегаты наличия по поставщикам для каждой component_id.

    Подмножество строк — то же, что для min-цены: только активные
    поставщики и активные позиции (`stock_qty > 0 OR transit_qty > 0`).
    Тем самым «Склад, шт» / «Транзит, шт» / «Поставщиков, шт» когерентны
    с «Поставщик (min)»: видишь у поставщика min-цену → его остаток
    включён в суммы.

    Result-схема:
      {
        <component_id>: {
          'on_hand':       int,  # SUM(stock_qty)
          'in_transit':    int,  # SUM(transit_qty)
          'suppliers':     int,  # COUNT(DISTINCT supplier_id)
        }
      }
    Компоненты без активных предложений в результате отсутствуют (Excel
    показывает пустые ячейки — означает «нет наличия ни у кого»).
    """
    if not component_ids:
        return {}
    sql = text(
        """
        SELECT sp.component_id,
               COALESCE(SUM(sp.stock_qty),   0) AS on_hand,
               COALESCE(SUM(sp.transit_qty), 0) AS in_transit,
               COUNT(DISTINCT sp.supplier_id)   AS suppliers
          FROM supplier_prices sp
          JOIN suppliers s ON s.id = sp.supplier_id
         WHERE sp.category = :cat
           AND s.is_active = TRUE
           AND (sp.stock_qty > 0 OR sp.transit_qty > 0)
           AND sp.component_id = ANY(:ids)
         GROUP BY sp.component_id
        """
    )
    rows = db.execute(
        sql, {"cat": category, "ids": list(component_ids)},
    ).mappings().all()
    return {
        int(r["component_id"]): {
            "on_hand":    int(r["on_hand"] or 0),
            "in_transit": int(r["in_transit"] or 0),
            "suppliers":  int(r["suppliers"] or 0),
        }
        for r in rows
    }


# ---------------------------------------------------------------------
# Преобразование значений ячеек
# ---------------------------------------------------------------------

def _serialize_array(value: Any) -> str | None:
    """TEXT[] / list → 'X,Y,Z'. Пустой массив или None → None (пустая
    ячейка в Excel)."""
    if value is None:
        return None
    if isinstance(value, str):
        # Уже строка (например, attrs_jsonb приехало строкой).
        s = value.strip()
        return s or None
    if isinstance(value, (list, tuple)):
        cleaned = [str(x).strip() for x in value if x not in (None, "")]
        cleaned = [x for x in cleaned if x]
        if not cleaned:
            return None
        return ",".join(cleaned)
    # Fallback: всё остальное приводим к str.
    return str(value)


def _attrs_value(attrs: dict | None, key: str) -> Any:
    """Достаёт ключ из attrs_jsonb с учётом сериализации list-ов."""
    if not attrs:
        return None
    if key not in attrs:
        return None
    value = attrs[key]
    if isinstance(value, list):
        return _serialize_array(value)
    return value


def _cell_value_for_column(
    col: _Col, row: dict[str, Any], attrs: dict | None,
) -> Any:
    """Возвращает значение для записи в Excel-ячейку (без price:* —
    они обрабатываются отдельно, так как требуют ссылок на другие ячейки)."""
    src = col.source
    if src.startswith("col:"):
        return row.get(src[4:])
    if src.startswith("array:"):
        return _serialize_array(row.get(src[6:]))
    if src.startswith("attr:"):
        return _attrs_value(attrs, src[5:])
    # price:* колонки обрабатываются вне этой функции.
    return None


# ---------------------------------------------------------------------
# Запись листа
# ---------------------------------------------------------------------

def _required_db_columns(columns: Sequence[_Col]) -> list[str]:
    """Список колонок БД, которые нужно SELECT'нуть, чтобы заполнить лист.

    Включает все col:/array:-источники, плюс id (если он не среди них —
    нужен для джойна цен) и attrs_jsonb (если есть attr:-колонки).
    """
    need: list[str] = []
    has_attrs = False
    for col in columns:
        if col.source.startswith("col:"):
            need.append(col.source[4:])
        elif col.source.startswith("array:"):
            need.append(col.source[6:])
        elif col.source.startswith("attr:"):
            has_attrs = True
    if "id" not in need:
        need.insert(0, "id")
    if has_attrs and "attrs_jsonb" not in need:
        need.append("attrs_jsonb")
    # Удаляем дубликаты, сохраняя порядок.
    seen: set[str] = set()
    ordered: list[str] = []
    for c in need:
        if c not in seen:
            seen.add(c)
            ordered.append(c)
    return ordered


def _write_sheet(
    ws: Worksheet,
    spec: _SheetSpec,
    rows: list[dict[str, Any]],
    prices: dict[int, dict[str, Any]],
    availability: dict[int, dict[str, int]],
    rate: Decimal,
) -> None:
    """Заполняет один лист по spec'у. Возвращает число записанных строк
    через len(rows) — но фактически использует side-effect на ws."""
    # 1. Служебная строка 1: курс ЦБ.
    ws[_RATE_LABEL_CELL] = "Курс ЦБ (USD→RUB)"
    ws[_RATE_LABEL_CELL].font = _FONT_HEADER
    ws[_RATE_LABEL_CELL].fill = _FILL_RATE
    ws[_RATE_VALUE_CELL] = float(rate)
    ws[_RATE_VALUE_CELL].number_format = "0.0000"
    ws[_RATE_VALUE_CELL].fill = _FILL_RATE
    ws[_RATE_VALUE_CELL].font = _FONT_HEADER

    # 2. Шапка таблицы на строке _HEADER_ROW.
    for col_idx, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=col.title)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_HEADER
        if col.kind == "ro":
            cell.fill = _FILL_RO

    # 3. Скрытые колонки.
    for col_idx, col in enumerate(spec.columns, start=1):
        if col.kind == "hidden":
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # 4. Pre-compute Excel-координат USD-колонки (для формулы RUB).
    usd_col_idx: int | None = None
    for col_idx, col in enumerate(spec.columns, start=1):
        if col.source == "price:usd":
            usd_col_idx = col_idx
            break

    # 5. Запись данных, начиная со строки _HEADER_ROW + 1 (4).
    data_start_row = _HEADER_ROW + 1
    for r_offset, row in enumerate(rows):
        excel_row = data_start_row + r_offset
        # attrs_jsonb (для printers_mfu) приходит в row под этим ключом —
        # драйвер psycopg2 уже распаковал его в dict.
        attrs = row.get("attrs_jsonb")
        component_id = row.get("id")
        price_info = prices.get(int(component_id), {}) if component_id else {}
        min_usd = price_info.get("min_usd")
        min_rub = price_info.get("min_rub")
        supplier_name = price_info.get("supplier_name")
        price_updated_at = price_info.get("updated_at")
        avail_info = availability.get(int(component_id), {}) if component_id else {}
        on_hand = avail_info.get("on_hand")
        in_transit = avail_info.get("in_transit")
        suppliers_with_stock = avail_info.get("suppliers")

        for col_idx, col in enumerate(spec.columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            if col.kind == "ro":
                cell.fill = _FILL_RO

            src = col.source
            if src == "price:usd":
                if min_usd is not None:
                    cell.value = float(min_usd)
                    cell.number_format = "0.00"
            elif src == "price:rub":
                if min_usd is not None and usd_col_idx is not None:
                    usd_letter = get_column_letter(usd_col_idx)
                    cell.value = f"={usd_letter}{excel_row}*{_RATE_ABS_REF}"
                    cell.number_format = "0.00"
                elif min_rub is not None:
                    cell.value = float(min_rub)
                    cell.number_format = "0.00"
            elif src == "price:supplier":
                if supplier_name:
                    cell.value = supplier_name
            elif src == "price:updated_at":
                if price_updated_at is not None:
                    cell.value = price_updated_at
                    cell.number_format = "yyyy-mm-dd hh:mm"
            elif src == "stock:on_hand":
                # 0 — это значащее «нет на складе ни у кого», пишем число;
                # отсутствие записи (нет активных предложений вообще) —
                # пустая ячейка.
                if on_hand is not None:
                    cell.value = on_hand
                    cell.number_format = "0"
            elif src == "stock:in_transit":
                if in_transit is not None:
                    cell.value = in_transit
                    cell.number_format = "0"
            elif src == "stock:suppliers":
                if suppliers_with_stock is not None:
                    cell.value = suppliers_with_stock
                    cell.number_format = "0"
            else:
                value = _cell_value_for_column(col, row, attrs)
                if value is not None:
                    cell.value = value

    # 6. Autofilter на шапку.
    if spec.columns:
        last_letter = get_column_letter(len(spec.columns))
        ws.auto_filter.ref = f"A{_HEADER_ROW}:{last_letter}{_HEADER_ROW}"


# ---------------------------------------------------------------------
# Публичные функции экспорта
# ---------------------------------------------------------------------

def _build_workbook(
    db: Session,
    specs: Sequence[_SheetSpec],
    output_path: Path,
) -> ExportReport:
    """Общий код двух экспортов: создаёт Workbook, прогоняет specs,
    сохраняет файл, возвращает ExportReport."""
    rate, rate_date, is_fallback = _read_latest_rate(db)
    if is_fallback:
        logger.warning(
            "catalog excel-export: exchange_rates пуст, использую fallback "
            "%.4f RUB/USD. RUB-формулы будут считаться с этим значением.",
            float(rate),
        )

    wb = Workbook()
    # У свежесозданного Workbook'а уже есть один лист — переименовываем
    # его в первый spec, остальные создаём create_sheet'ом.
    first = True
    sheet_counts: dict[str, int] = {}
    for spec in specs:
        if first:
            ws = wb.active
            ws.title = spec.title
            first = False
        else:
            ws = wb.create_sheet(title=spec.title)
        db_cols = _required_db_columns(spec.columns)
        rows = _fetch_main_rows(db, spec, db_cols)
        component_ids = [int(r["id"]) for r in rows if r.get("id") is not None]
        prices = _fetch_min_prices(db, spec.price_category, component_ids)
        availability = _fetch_availability(db, spec.price_category, component_ids)
        _write_sheet(ws, spec, rows, prices, availability, rate)
        sheet_counts[spec.title] = len(rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return ExportReport(
        file_path=output_path,
        sheet_counts=sheet_counts,
        total_rows=sum(sheet_counts.values()),
        rate_used=rate,
        rate_date=rate_date,
        rate_is_fallback=is_fallback,
    )


def export_components_pc(
    output_path: Path, *, db: Session | None = None,
) -> ExportReport:
    """Выгружает «Комплектующие_ПК.xlsx» (8 листов).

    Args:
        output_path: куда сохранить файл (.xlsx).
        db: опциональная SQLAlchemy-сессия. Если None — открываем свою
            на время вызова (для CLI и UI-роута).
    """
    if db is None:
        db = SessionLocal()
        try:
            return _build_workbook(db, _components_sheet_specs(), output_path)
        finally:
            db.close()
    return _build_workbook(db, _components_sheet_specs(), output_path)


def export_printers_mfu(
    output_path: Path, *, db: Session | None = None,
) -> ExportReport:
    """Выгружает «Печатная_техника.xlsx» (2 листа: Принтеры/МФУ)."""
    if db is None:
        db = SessionLocal()
        try:
            return _build_workbook(db, _printers_mfu_sheet_specs(), output_path)
        finally:
            db.close()
    return _build_workbook(db, _printers_mfu_sheet_specs(), output_path)


# ---------------------------------------------------------------------
# Имена файлов для скачивания (используются UI-роутом и CLI)
# ---------------------------------------------------------------------

def default_filename(target: str, today: date | None = None) -> str:
    """Имя файла по умолчанию: «Комплектующие_ПК_YYYY-MM-DD.xlsx» /
    «Печатная_техника_YYYY-MM-DD.xlsx»."""
    if today is None:
        today = date.today()
    stamp = today.strftime("%Y-%m-%d")
    if target == "pc":
        return f"Комплектующие_ПК_{stamp}.xlsx"
    if target == "printers":
        return f"Печатная_техника_{stamp}.xlsx"
    raise ValueError(f"неизвестный target='{target}'")
