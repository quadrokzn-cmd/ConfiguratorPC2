# Импорт каталога товаров из Excel (Фаза 3 плана 2026-05-13).
#
# Принимает файл, сгенерированный Фазой 2 экспорта, и применяет правки
# к таблицам каталога:
#   - комплектующие ПК: cpus / motherboards / rams / gpus / storages /
#     cases / psus / coolers (8 листов);
#   - печатная техника: printers_mfu (2 листа, фильтр category in
#     ('printer','mfu'); attrs_jsonb обновляется per-key merge'ом).
#
# Поведение по строке:
#   - id есть, запись найдена → UPDATE редактируемых полей.
#   - id пустой, name заполнен → INSERT нового товара.
#   - id есть, в БД не найден → skip + запись в report.
#   - Полностью пустая строка (только id или ничего) → skip без ошибки.
#
# Цены, поставщик и даты обновления — read-only: ячейки этих колонок
# игнорируются. Если в шапке листа такие колонки присутствовали, в
# report добавляется одна строка-предупреждение «read-only columns
# ignored».
#
# Валидация — на уровне ячейки (тип, обязательные поля, enum). Ошибки
# собираются в ImportReport.errors и НЕ прерывают всю загрузку: валидные
# строки применяются. Один общий transaction на весь файл — при SQL-ошибке
# rollback всего файла (валидационные ошибки rollback не вызывают).
#
# Batched-операции: UPDATE и INSERT идут через `executemany` — psycopg2
# собирает их в одну сетевую отправку (insertmanyvalues с SQLAlchemy 2.0).
# Это критично для Railway (см. memory feedback_remote_db_n1_pattern).

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from portal.services.auctions.catalog.enrichment.schema import (
    DUPLEX_VALUES,
    COLORNESS_VALUES,
    MAX_FORMAT_VALUES,
    NA,
    NETWORK_INTERFACE_VALUES,
    PRINT_TECH_VALUES,
    PRINTER_MFU_ATTRS,
    PRINTER_MFU_DIMENSION_ATTRS,
    USB_VALUES,
)
from shared.db import SessionLocal


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Структуры отчёта
# ---------------------------------------------------------------------------


@dataclass
class ImportRowError:
    sheet: str
    row: int
    message: str


@dataclass
class ImportReport:
    """Сводный отчёт по импорту одного xlsx-файла."""

    updated: int = 0
    inserted: int = 0
    skipped: int = 0
    errors: list[ImportRowError] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    saved_path: str | None = None

    @property
    def error_count(self) -> int:
        return len(self.errors)

    def to_dict(self) -> dict:
        return {
            "updated":     self.updated,
            "inserted":    self.inserted,
            "skipped":     self.skipped,
            "errors":      [
                {"sheet": e.sheet, "row": e.row, "message": e.message}
                for e in self.errors
            ],
            "warnings":    list(self.warnings),
            "saved_path":  self.saved_path,
        }


# ---------------------------------------------------------------------------
# Описание колонок листов
# ---------------------------------------------------------------------------
#
# Каждая колонка: ('header', 'db_field' | None, 'kind', 'type').
# kind:
#   'id'        — скрытая первая колонка (PK строки в БД).
#   'edit'      — редактируемое поле; пишется в UPDATE/INSERT.
#   'readonly'  — игнорируется importer'ом (цены, поставщик, даты).
#   'attrs'     — ключ attrs_jsonb (per-key merge с валидацией).
# type:
#   'int' / 'float' / 'str' / 'bool' / 'array' /
#   'enum:<...>'  (для attrs-полей с фикс-набором).
#
# Заголовки совпадают с теми, что пишет Фаза 2 экспорта (см. план).
# Read-only-колонки перечислены для документации и для предупреждения
# в report'е, если пользователь их случайно поправил.


# Базовые edit-колонки, общие для всех 8 категорий комплектующих ПК.
_COMMON_EDIT_PC: list[tuple[str, str, str, str]] = [
    ("model",        "model",        "edit", "str"),
    ("manufacturer", "manufacturer", "edit", "str"),
    ("sku",          "sku",          "edit", "str"),
    ("gtin",         "gtin",         "edit", "str"),
    ("is_hidden",    "is_hidden",    "edit", "bool"),
]

_RO_PC: list[tuple[str, str, str, str]] = [
    ("Цена min, USD",   None, "readonly", "float"),
    ("Цена min, RUB",   None, "readonly", "float"),
    ("Поставщик (min)", None, "readonly", "str"),
    ("Цена обновлена",  None, "readonly", "datetime"),
    # Колонки наличия — агрегаты по supplier_prices. Импортер их
    # игнорирует, пишет общий read-only warning. Реальные значения
    # обновляются только автозагрузкой прайсов.
    ("Склад, шт",       None, "readonly", "int"),
    ("Транзит, шт",     None, "readonly", "int"),
    ("Поставщиков, шт", None, "readonly", "int"),
]


# INSERT-дефолты для PC-таблиц: только is_hidden имеет NOT NULL DEFAULT
# (миграция 013_components_is_hidden), остальные NOT NULL поля без
# DEFAULT ловятся через required_for_insert.
_PC_INSERT_DEFAULTS: dict[str, str] = {
    "is_hidden": "FALSE",
}


def _pc_sheet(
    sheet_name: str,
    table: str,
    category: str,
    spec_columns: list[tuple[str, str, str, str]],
    required_for_insert: list[str],
) -> dict:
    return {
        "sheet_name": sheet_name,
        "table":      table,
        "category":   category,
        "kind":       "pc_component",
        "columns": (
            [("id", "id", "id", "int")]
            + _COMMON_EDIT_PC
            + spec_columns
            + _RO_PC
        ),
        "required_for_insert": required_for_insert,
        "insert_defaults":     _PC_INSERT_DEFAULTS,
    }


# Минимальный набор NOT NULL-полей для INSERT (см. миграцию 001 +
# 013_components_is_hidden). model и manufacturer NOT NULL у всех; is_hidden
# имеет DEFAULT FALSE — не обязателен.
PC_SHEETS: dict[str, dict] = {
    "cpu": _pc_sheet(
        "CPU", "cpus", "cpu",
        spec_columns=[
            ("socket",                  "socket",                  "edit", "str"),
            ("cores",                   "cores",                   "edit", "int"),
            ("threads",                 "threads",                 "edit", "int"),
            ("base_clock_ghz",          "base_clock_ghz",          "edit", "float"),
            ("turbo_clock_ghz",         "turbo_clock_ghz",         "edit", "float"),
            ("tdp_watts",               "tdp_watts",               "edit", "int"),
            ("has_integrated_graphics", "has_integrated_graphics", "edit", "bool"),
            ("memory_type",             "memory_type",             "edit", "str"),
            ("package_type",            "package_type",            "edit", "str"),
            ("process_nm",              "process_nm",              "edit", "int"),
            ("l3_cache_mb",             "l3_cache_mb",             "edit", "int"),
            ("max_memory_freq",         "max_memory_freq",         "edit", "int"),
            ("release_year",            "release_year",            "edit", "int"),
        ],
        required_for_insert=[
            "model", "manufacturer", "socket", "cores", "threads",
            "base_clock_ghz", "turbo_clock_ghz", "tdp_watts",
            "has_integrated_graphics", "memory_type", "package_type",
        ],
    ),
    "motherboard": _pc_sheet(
        "Motherboard", "motherboards", "motherboard",
        spec_columns=[
            ("socket",          "socket",          "edit", "str"),
            ("chipset",         "chipset",         "edit", "str"),
            ("form_factor",     "form_factor",     "edit", "str"),
            ("memory_type",     "memory_type",     "edit", "str"),
            ("has_m2_slot",     "has_m2_slot",     "edit", "bool"),
            ("memory_slots",    "memory_slots",    "edit", "int"),
            ("max_memory_gb",   "max_memory_gb",   "edit", "int"),
            ("max_memory_freq", "max_memory_freq", "edit", "int"),
            ("sata_ports",      "sata_ports",      "edit", "int"),
            ("m2_slots",        "m2_slots",        "edit", "int"),
            ("has_wifi",        "has_wifi",        "edit", "bool"),
            ("has_bluetooth",   "has_bluetooth",   "edit", "bool"),
            ("pcie_version",    "pcie_version",    "edit", "str"),
            ("pcie_x16_slots",  "pcie_x16_slots",  "edit", "int"),
            ("usb_ports",       "usb_ports",       "edit", "int"),
        ],
        required_for_insert=[
            "model", "manufacturer", "socket", "chipset", "form_factor",
            "memory_type", "has_m2_slot",
        ],
    ),
    "ram": _pc_sheet(
        "RAM", "rams", "ram",
        spec_columns=[
            ("memory_type",    "memory_type",    "edit", "str"),
            ("form_factor",    "form_factor",    "edit", "str"),
            ("module_size_gb", "module_size_gb", "edit", "int"),
            ("modules_count",  "modules_count",  "edit", "int"),
            ("frequency_mhz",  "frequency_mhz",  "edit", "int"),
            ("cl_timing",      "cl_timing",      "edit", "int"),
            ("voltage",        "voltage",        "edit", "float"),
            ("has_heatsink",   "has_heatsink",   "edit", "bool"),
            ("has_rgb",        "has_rgb",        "edit", "bool"),
        ],
        required_for_insert=[
            "model", "manufacturer", "memory_type", "form_factor",
            "module_size_gb", "modules_count", "frequency_mhz",
        ],
    ),
    "gpu": _pc_sheet(
        "GPU", "gpus", "gpu",
        spec_columns=[
            ("vram_gb",               "vram_gb",               "edit", "int"),
            ("vram_type",             "vram_type",             "edit", "str"),
            ("tdp_watts",             "tdp_watts",             "edit", "int"),
            ("needs_extra_power",     "needs_extra_power",     "edit", "bool"),
            ("video_outputs",         "video_outputs",         "edit", "str"),
            ("core_clock_mhz",        "core_clock_mhz",        "edit", "int"),
            ("memory_clock_mhz",      "memory_clock_mhz",      "edit", "int"),
            ("gpu_chip",              "gpu_chip",              "edit", "str"),
            ("recommended_psu_watts", "recommended_psu_watts", "edit", "int"),
            ("length_mm",             "length_mm",             "edit", "int"),
            ("height_mm",             "height_mm",             "edit", "int"),
            ("power_connectors",      "power_connectors",      "edit", "str"),
            ("fans_count",            "fans_count",            "edit", "int"),
        ],
        required_for_insert=[
            "model", "manufacturer", "vram_gb", "vram_type", "tdp_watts",
            "needs_extra_power", "video_outputs", "core_clock_mhz",
            "memory_clock_mhz",
        ],
    ),
    "storage": _pc_sheet(
        "Storage", "storages", "storage",
        spec_columns=[
            ("storage_type",   "storage_type",   "edit", "str"),
            ("form_factor",    "form_factor",    "edit", "str"),
            ("interface",      "interface",      "edit", "str"),
            ("capacity_gb",    "capacity_gb",    "edit", "int"),
            ("read_speed_mb",  "read_speed_mb",  "edit", "int"),
            ("write_speed_mb", "write_speed_mb", "edit", "int"),
            ("tbw",            "tbw",            "edit", "int"),
            ("rpm",            "rpm",            "edit", "int"),
            ("cache_mb",       "cache_mb",       "edit", "int"),
        ],
        required_for_insert=[
            "model", "manufacturer", "storage_type", "form_factor",
            "interface", "capacity_gb",
        ],
    ),
    "case": _pc_sheet(
        "Case", "cases", "case",
        spec_columns=[
            ("supported_form_factors", "supported_form_factors", "edit", "array"),
            ("has_psu_included",       "has_psu_included",       "edit", "bool"),
            ("included_psu_watts",     "included_psu_watts",     "edit", "int"),
            ("max_gpu_length_mm",      "max_gpu_length_mm",      "edit", "int"),
            ("max_cooler_height_mm",   "max_cooler_height_mm",   "edit", "int"),
            ("psu_form_factor",        "psu_form_factor",        "edit", "str"),
            ("color",                  "color",                  "edit", "str"),
            ("material",               "material",               "edit", "str"),
            ("drive_bays",             "drive_bays",             "edit", "int"),
            ("fans_included",          "fans_included",          "edit", "int"),
            ("has_glass_panel",        "has_glass_panel",        "edit", "bool"),
            ("has_rgb",                "has_rgb",                "edit", "bool"),
        ],
        required_for_insert=[
            "model", "manufacturer", "supported_form_factors", "has_psu_included",
        ],
    ),
    "psu": _pc_sheet(
        "PSU", "psus", "psu",
        spec_columns=[
            ("power_watts",          "power_watts",          "edit", "int"),
            ("form_factor",          "form_factor",          "edit", "str"),
            ("efficiency_rating",    "efficiency_rating",    "edit", "str"),
            ("modularity",           "modularity",           "edit", "str"),
            ("has_12vhpwr",          "has_12vhpwr",          "edit", "bool"),
            ("sata_connectors",      "sata_connectors",      "edit", "int"),
            ("main_cable_length_mm", "main_cable_length_mm", "edit", "int"),
            ("warranty_years",       "warranty_years",       "edit", "int"),
        ],
        required_for_insert=["model", "manufacturer", "power_watts"],
    ),
    "cooler": _pc_sheet(
        "Cooler", "coolers", "cooler",
        spec_columns=[
            ("supported_sockets", "supported_sockets", "edit", "array"),
            ("max_tdp_watts",     "max_tdp_watts",     "edit", "int"),
            ("cooler_type",       "cooler_type",       "edit", "str"),
            ("height_mm",         "height_mm",         "edit", "int"),
            ("radiator_size_mm",  "radiator_size_mm",  "edit", "int"),
            ("fans_count",        "fans_count",        "edit", "int"),
            ("noise_db",          "noise_db",          "edit", "float"),
            ("has_rgb",           "has_rgb",           "edit", "bool"),
        ],
        required_for_insert=[
            "model", "manufacturer", "supported_sockets", "max_tdp_watts",
        ],
    ),
}


# Колонки печатной техники (Принтеры/МФУ — один список, отличаются только
# фильтром category).
_PRINTER_COMMON_EDIT: list[tuple[str, str, str, str]] = [
    ("sku",                "sku",                "edit", "str"),
    ("mpn",                "mpn",                "edit", "str"),
    ("gtin",               "gtin",               "edit", "str"),
    ("brand",              "brand",              "edit", "str"),
    ("name",               "name",               "edit", "str"),
    ("category",           "category",           "edit", "str"),
    ("ktru_codes_array",   "ktru_codes_array",   "edit", "array"),
    ("is_hidden",          "is_hidden",          "edit", "bool"),
    ("cost_base_rub",      "cost_base_rub",      "edit", "float"),
    ("margin_pct_target",  "margin_pct_target",  "edit", "float"),
]

# attrs_jsonb-ключи: 9 обязательных + 4 опциональных. type — для подсказки
# валидатору. Реальная валидация делается через validate_attrs() из
# auctions/catalog/enrichment/schema.py.
_PRINTER_ATTRS_COLS: list[tuple[str, str, str, str]] = [
    ("print_speed_ppm",         "print_speed_ppm",         "attrs", "int"),
    ("colorness",               "colorness",               "attrs", "enum:colorness"),
    ("max_format",              "max_format",              "attrs", "enum:max_format"),
    ("duplex",                  "duplex",                  "attrs", "enum:duplex"),
    ("resolution_dpi",          "resolution_dpi",          "attrs", "int"),
    ("network_interface",       "network_interface",       "attrs", "enum_array:network_interface"),
    ("usb",                     "usb",                     "attrs", "enum:usb"),
    ("starter_cartridge_pages", "starter_cartridge_pages", "attrs", "int"),
    ("print_technology",        "print_technology",        "attrs", "enum:print_technology"),
    ("weight_kg",               "weight_kg",               "attrs", "float"),
    ("box_width_cm",            "box_width_cm",            "attrs", "float"),
    ("box_height_cm",           "box_height_cm",           "attrs", "float"),
    ("box_depth_cm",            "box_depth_cm",            "attrs", "float"),
]

_PRINTER_RO_COLS: list[tuple[str, str, str, str]] = [
    ("attrs_source",    None, "readonly", "str"),
    ("Цена min, USD",   None, "readonly", "float"),
    ("Цена min, RUB",   None, "readonly", "float"),
    ("Поставщик (min)", None, "readonly", "str"),
    ("Цена обновлена",  None, "readonly", "datetime"),
    ("Склад, шт",       None, "readonly", "int"),
    ("Транзит, шт",     None, "readonly", "int"),
    ("Поставщиков, шт", None, "readonly", "int"),
]


# INSERT-дефолты для полей с NOT NULL DEFAULT в БД. Если соответствующая
# ячейка Excel пустая → используем DB-default через COALESCE. Сводный
# источник правды; добавляется в sheet_cfg.
_PRINTER_INSERT_DEFAULTS: dict[str, str] = {
    "is_hidden":        "FALSE",
    "ktru_codes_array": "ARRAY[]::TEXT[]",
}


PRINTER_SHEETS: dict[str, dict] = {
    "printer": {
        "sheet_name": "Принтеры",
        "table":      "printers_mfu",
        "category":   "printer",
        "kind":       "printer_mfu",
        "columns": (
            [("id", "id", "id", "int")]
            + _PRINTER_COMMON_EDIT
            + _PRINTER_ATTRS_COLS
            + _PRINTER_RO_COLS
        ),
        "required_for_insert": ["sku", "brand", "name", "category"],
        "insert_defaults":     _PRINTER_INSERT_DEFAULTS,
    },
    "mfu": {
        "sheet_name": "МФУ",
        "table":      "printers_mfu",
        "category":   "mfu",
        "kind":       "printer_mfu",
        "columns": (
            [("id", "id", "id", "int")]
            + _PRINTER_COMMON_EDIT
            + _PRINTER_ATTRS_COLS
            + _PRINTER_RO_COLS
        ),
        "required_for_insert": ["sku", "brand", "name", "category"],
        "insert_defaults":     _PRINTER_INSERT_DEFAULTS,
    },
}


# ---------------------------------------------------------------------------
# Парсинг значений ячейки → Python-тип
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _parse_bool(value: Any) -> bool | None:
    """TRUE/FALSE из Excel-ячейки. None если ячейка пуста."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes", "да", "y"):
        return True
    if s in ("false", "0", "no", "нет", "n"):
        return False
    return None


def _parse_int(value: Any) -> int | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        # bool — подвид int в Python; не хотим, чтобы True/False прокрался
        # в числовое поле.
        raise ValueError(f"ожидалось целое число, пришло bool ({value!r})")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != int(value):
            raise ValueError(f"ожидалось целое число, пришло {value!r}")
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    return int(float(s.replace(",", ".")))


def _parse_float(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        raise ValueError(f"ожидалось число, пришло bool ({value!r})")
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    s = str(value).strip().replace(",", ".")
    try:
        return float(s)
    except (ValueError, InvalidOperation) as exc:
        raise ValueError(f"ожидалось число, пришло {value!r}") from exc


def _parse_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _parse_array(value: Any) -> list[str] | None:
    """Сериализация массива через запятую: 'ATX,mATX,ITX' → ['ATX','mATX','ITX'].
    Пустая ячейка → None (т.е. ключ не обновляется при UPDATE)."""
    if _is_blank(value):
        return None
    s = str(value)
    items = [t.strip() for t in s.split(",")]
    return [t for t in items if t]


def _parse_attrs_value(
    *,
    field_name: str,
    raw: Any,
    type_hint: str,
) -> tuple[Any, str | None]:
    """Парсит значение для ключа attrs_jsonb.

    Возвращает (parsed, error). Если ячейка пустая — (None, None) (sentinel:
    «ключ не обновляется», см. план: «пустая ячейка → ключ не обновляется»).
    Если ячейка содержит 'n/a' — возвращает (NA, None).
    """
    if _is_blank(raw):
        return None, None

    s_raw = str(raw).strip()
    if s_raw.lower() == NA:
        return NA, None

    try:
        if type_hint == "int":
            return _parse_int(raw), None
        if type_hint == "float":
            v = _parse_float(raw)
            return v, None
        if type_hint.startswith("enum:"):
            allowed_key = type_hint[len("enum:"):]
            allowed = _ATTRS_ENUMS.get(allowed_key)
            if allowed is None:
                return None, f"{field_name}: внутренняя ошибка, неизвестный enum {allowed_key!r}"
            if s_raw not in allowed:
                return None, (
                    f"{field_name}: значение {s_raw!r} не в {sorted(allowed)}"
                )
            return s_raw, None
        if type_hint.startswith("enum_array:"):
            allowed_key = type_hint[len("enum_array:"):]
            allowed = _ATTRS_ENUMS.get(allowed_key)
            if allowed is None:
                return None, f"{field_name}: внутренняя ошибка, неизвестный enum-array {allowed_key!r}"
            items = [t.strip() for t in s_raw.split(",")]
            items = [t for t in items if t]
            if not items:
                return None, None
            for it in items:
                if it not in allowed:
                    return None, (
                        f"{field_name}: элемент {it!r} не в {sorted(allowed)}"
                    )
            return items, None
        return s_raw, None
    except ValueError as exc:
        return None, f"{field_name}: {exc}"


_ATTRS_ENUMS: dict[str, frozenset[str]] = {
    "colorness":         COLORNESS_VALUES,
    "max_format":        MAX_FORMAT_VALUES,
    "duplex":            DUPLEX_VALUES,
    "usb":               USB_VALUES,
    "print_technology":  PRINT_TECH_VALUES,
    "network_interface": NETWORK_INTERFACE_VALUES,
}


# ---------------------------------------------------------------------------
# Чтение листа + сбор изменений
# ---------------------------------------------------------------------------


# Структура шапки: row 1 — служебная (курс), row 2 — пусто, row 3 —
# заголовки колонок, row 4+ — данные. Это зафиксировано Фазой 1 плана.
_HEADER_ROW = 3
_DATA_START_ROW = 4


def _build_header_index(ws, expected_columns: list[tuple[str, str, str, str]]) -> dict[str, int]:
    """Возвращает {header_text: column_index_1based} по заголовкам в строке 3.

    Колонки, которых нет в файле, просто игнорируются — это допускается
    для будущей обратной совместимости (export может добавить новые
    колонки, а старый импорт читать их не обязан)."""
    header_index: dict[str, int] = {}
    # Сканируем все заполненные ячейки шапки до первой пустой подряд.
    col_idx = 1
    max_scan = 200  # защитник от бесконечного цикла
    empty_streak = 0
    while col_idx <= max_scan and empty_streak < 5:
        cell_value = ws.cell(row=_HEADER_ROW, column=col_idx).value
        if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
            empty_streak += 1
        else:
            header_index[str(cell_value).strip()] = col_idx
            empty_streak = 0
        col_idx += 1
    return header_index


def _row_is_empty(ws, row: int, max_col: int) -> bool:
    """Все ячейки строки (кроме, возможно, колонки id) пустые."""
    for c in range(2, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if not _is_blank(v):
            return False
    return True


def _last_data_row(ws) -> int:
    """Возвращает индекс последней заполненной строки данных (>= 4 или 3 если данных нет)."""
    return ws.max_row or _HEADER_ROW


def _read_pc_row(
    *,
    ws,
    row: int,
    header_index: dict[str, int],
    sheet_cfg: dict,
    errors: list[ImportRowError],
) -> tuple[dict[str, Any] | None, int | None]:
    """Читает одну строку листа PC-компонента. Возвращает (fields, id).

    fields — словарь {db_field: parsed_value} только для edit-колонок.
    Если значение клетки пустое — ключ ВКЛЮЧАЕТСЯ в fields со значением None
    (что приведёт к записи NULL при UPDATE/INSERT); это согласуется с
    логикой Excel-flow «что видишь, то и попадёт в БД».

    Возвращает (None, None) для полностью пустой строки.
    """
    sheet_name = sheet_cfg["sheet_name"]

    id_col = header_index.get("id")
    id_value: int | None = None
    if id_col:
        raw_id = ws.cell(row=row, column=id_col).value
        if not _is_blank(raw_id):
            try:
                id_value = _parse_int(raw_id)
            except ValueError as exc:
                errors.append(ImportRowError(
                    sheet=sheet_name, row=row,
                    message=f"id: {exc}",
                ))
                return None, None

    # Проверяем, не пустая ли строка целиком (для пропуска без ошибки).
    max_col = max(header_index.values()) if header_index else 1
    if _row_is_empty(ws, row, max_col):
        return None, id_value

    fields: dict[str, Any] = {}
    row_has_error = False
    for header, db_field, kind, type_hint in sheet_cfg["columns"]:
        if kind not in ("edit",):
            continue
        col = header_index.get(header)
        if not col:
            # Колонка отсутствует в файле — пропускаем, в UPDATE/INSERT
            # её просто не будет.
            continue
        raw = ws.cell(row=row, column=col).value
        try:
            if type_hint == "int":
                parsed = _parse_int(raw)
            elif type_hint == "float":
                parsed = _parse_float(raw)
            elif type_hint == "bool":
                parsed = _parse_bool(raw)
            elif type_hint == "array":
                parsed = _parse_array(raw)
            else:
                parsed = _parse_str(raw)
        except ValueError as exc:
            errors.append(ImportRowError(
                sheet=sheet_name, row=row,
                message=f"{db_field}: {exc}",
            ))
            row_has_error = True
            continue
        fields[db_field] = parsed

    if row_has_error:
        return None, id_value

    return fields, id_value


def _read_printer_row(
    *,
    ws,
    row: int,
    header_index: dict[str, int],
    sheet_cfg: dict,
    errors: list[ImportRowError],
) -> tuple[dict[str, Any] | None, dict[str, Any] | None, int | None]:
    """Читает строку листа Принтеры/МФУ. Возвращает (fields, attrs_partial, id).

    attrs_partial — dict {key: value | NA} с уже валидированными значениями.
    Пустые ячейки в эту таблицу НЕ попадают (per-key merge: «пустая ячейка
    → ключ не обновляется»). Это отличает поведение от PC-листов.
    """
    sheet_name = sheet_cfg["sheet_name"]

    id_col = header_index.get("id")
    id_value: int | None = None
    if id_col:
        raw_id = ws.cell(row=row, column=id_col).value
        if not _is_blank(raw_id):
            try:
                id_value = _parse_int(raw_id)
            except ValueError as exc:
                errors.append(ImportRowError(
                    sheet=sheet_name, row=row,
                    message=f"id: {exc}",
                ))
                return None, None, None

    max_col = max(header_index.values()) if header_index else 1
    if _row_is_empty(ws, row, max_col):
        return None, None, id_value

    fields: dict[str, Any] = {}
    attrs: dict[str, Any] = {}
    row_has_error = False

    for header, db_field, kind, type_hint in sheet_cfg["columns"]:
        col = header_index.get(header)
        if not col:
            continue
        raw = ws.cell(row=row, column=col).value
        if kind == "edit":
            try:
                if type_hint == "int":
                    parsed = _parse_int(raw)
                elif type_hint == "float":
                    parsed = _parse_float(raw)
                elif type_hint == "bool":
                    parsed = _parse_bool(raw)
                elif type_hint == "array":
                    parsed = _parse_array(raw)
                else:
                    parsed = _parse_str(raw)
            except ValueError as exc:
                errors.append(ImportRowError(
                    sheet=sheet_name, row=row,
                    message=f"{db_field}: {exc}",
                ))
                row_has_error = True
                continue
            fields[db_field] = parsed
        elif kind == "attrs":
            parsed, err = _parse_attrs_value(
                field_name=db_field, raw=raw, type_hint=type_hint,
            )
            if err:
                errors.append(ImportRowError(
                    sheet=sheet_name, row=row, message=err,
                ))
                row_has_error = True
                continue
            if parsed is not None:
                attrs[db_field] = parsed
        # readonly — игнорируется.

    if row_has_error:
        return None, None, id_value

    # category на листе должна совпадать с типом листа (printer/mfu).
    if "category" in fields and fields["category"] not in (None, sheet_cfg["category"]):
        errors.append(ImportRowError(
            sheet=sheet_name, row=row,
            message=(
                f"category: значение {fields['category']!r} не совпадает с "
                f"листом ({sheet_cfg['category']!r}); правка category между "
                f"листами не поддерживается"
            ),
        ))
        return None, None, id_value
    # Подставляем category, если ячейка была пустой.
    fields["category"] = sheet_cfg["category"]

    return fields, attrs, id_value


# ---------------------------------------------------------------------------
# Применение изменений к БД
# ---------------------------------------------------------------------------


def _pg_type_for(ftype: str) -> str | None:
    """Возвращает CAST-тип PostgreSQL для bind-параметра, либо None если
    каст не нужен. Нужно для COALESCE(:field, table.field) — psycopg2 шлёт
    None без типа, и PG может «не знать», к чему приводить."""
    if ftype == "bool":
        return "BOOLEAN"
    if ftype == "array":
        return "TEXT[]"
    if ftype == "int":
        return "INTEGER"
    if ftype == "float":
        return "NUMERIC"
    return None


def _coalesce_param(name: str, ftype: str, fallback_sql: str) -> str:
    """Возвращает выражение `COALESCE(CAST(:name AS T), fallback_sql)`,
    с опущенным CAST для типов, не требующих явного приведения."""
    pg_t = _pg_type_for(ftype)
    if pg_t:
        return f"COALESCE(CAST(:{name} AS {pg_t}), {fallback_sql})"
    return f"COALESCE(:{name}, {fallback_sql})"


def _apply_pc_sheet(
    *,
    session: Session,
    sheet_cfg: dict,
    rows_update: list[tuple[int, dict[str, Any]]],
    rows_insert: list[dict[str, Any]],
    report: ImportReport,
    errors: list[ImportRowError],
    insert_row_numbers: list[int],
) -> None:
    """Применяет UPDATE (per-row, с COALESCE) + INSERT (batched) для PC-листа.

    UPDATE-семантика: пустая ячейка в Excel → значение в БД остаётся
    прежним. Это согласуется с per-key merge в attrs_jsonb для печатной
    техники и подходит для usecase «правлю пару колонок, остальные не
    трогаю». Минус: пользователь не может «обнулить» значение через
    Excel — для этого нужен UI или CSV-manual-edit-flow. Документировано
    в плане 2026-05-13.

    INSERT — один executemany на весь лист: SQLAlchemy 2.0 +
    psycopg2 insertmanyvalues автоматически батчит запросы (см.
    feedback_remote_db_n1_pattern: на Railway это критично).
    """
    sheet_name = sheet_cfg["sheet_name"]
    table = sheet_cfg["table"]

    edit_specs: list[tuple[str, str]] = [
        (db_field, ftype)
        for _h, db_field, kind, ftype in sheet_cfg["columns"]
        if kind == "edit"
    ]
    edit_fields = [f for f, _t in edit_specs]

    # --- UPDATE: единый SQL-шаблон, COALESCE на каждое поле. Одинаков
    # для всех строк → можно гонять через session.execute(stmt, params)
    # в цикле; psycopg2 кэширует prepared statement.
    if rows_update:
        set_parts = [
            f"{f} = " + _coalesce_param(f, ftype, f"{table}.{f}")
            for f, ftype in edit_specs
        ]
        update_sql = text(
            f"UPDATE {table} SET {', '.join(set_parts)} "
            f"WHERE id = :__id RETURNING id"
        )

        for cid, fields in rows_update:
            params: dict[str, Any] = {"__id": cid}
            for f in edit_fields:
                params[f] = fields.get(f)
            try:
                res = session.execute(update_sql, params)
                row = res.first()
                if row is None:
                    report.skipped += 1
                    report.warnings.append(
                        f"{sheet_name}: id={cid} не найден в БД, строка пропущена"
                    )
                else:
                    report.updated += 1
            except Exception as exc:
                errors.append(ImportRowError(
                    sheet=sheet_name, row=0,
                    message=f"UPDATE id={cid}: {type(exc).__name__}: {exc}",
                ))
                raise

    # --- INSERT: один SQL-шаблон, batched executemany. Для полей с
    # NOT NULL DEFAULT (см. sheet_cfg.insert_defaults) — COALESCE с
    # SQL-defaultом, чтобы пустая ячейка не валила INSERT. Прочие NOT NULL
    # поля без DEFAULT обязаны быть в required_for_insert.
    if rows_insert:
        defaults = sheet_cfg.get("insert_defaults") or {}
        cols_sql = ", ".join(edit_fields)
        value_parts: list[str] = []
        for f, ftype in edit_specs:
            if f in defaults:
                value_parts.append(_coalesce_param(f, ftype, defaults[f]))
            else:
                pg_t = _pg_type_for(ftype)
                if pg_t:
                    value_parts.append(f"CAST(:{f} AS {pg_t})")
                else:
                    value_parts.append(f":{f}")
        insert_sql = text(
            f"INSERT INTO {table} ({cols_sql}) "
            f"VALUES ({', '.join(value_parts)})"
        )

        params_list = [
            {f: fields.get(f) for f in edit_fields}
            for fields in rows_insert
        ]
        try:
            session.execute(insert_sql, params_list)
            report.inserted += len(params_list)
        except Exception as exc:
            errors.append(ImportRowError(
                sheet=sheet_name, row=0,
                message=f"INSERT batch ({len(params_list)} rows): {type(exc).__name__}: {exc}",
            ))
            raise


def _apply_printer_sheet(
    *,
    session: Session,
    sheet_cfg: dict,
    rows_update: list[tuple[int, dict[str, Any], dict[str, Any]]],
    rows_insert: list[tuple[dict[str, Any], dict[str, Any]]],
    report: ImportReport,
    errors: list[ImportRowError],
) -> None:
    """UPDATE (per-row, COALESCE-семантика для edit-полей, jsonb-merge
    для attrs) + batched INSERT для printers_mfu.

    Семантика: пустая ячейка edit-поля → значение в БД не меняется
    (COALESCE). Пустая ячейка attrs_jsonb-ключа → ключ не обновляется
    (per-key merge через jsonb-оператор `||`). 'n/a' → пишется строкой.
    """
    sheet_name = sheet_cfg["sheet_name"]
    table = sheet_cfg["table"]

    edit_specs: list[tuple[str, str]] = [
        (db_field, ftype)
        for _h, db_field, kind, ftype in sheet_cfg["columns"]
        if kind == "edit"
    ]
    edit_fields = [f for f, _t in edit_specs]

    # UPDATE: единый SQL с COALESCE на каждое edit-поле + jsonb-merge.
    if rows_update:
        set_parts = [
            f"{f} = " + _coalesce_param(f, ftype, f"{table}.{f}")
            for f, ftype in edit_specs
        ]
        set_parts.append(
            "attrs_jsonb = attrs_jsonb || CAST(:__attrs_partial AS JSONB)"
        )
        upd = text(
            f"UPDATE {table} SET {', '.join(set_parts)} "
            f"WHERE id = :__id RETURNING id"
        )
        for cid, fields, attrs_partial in rows_update:
            params: dict[str, Any] = {"__id": cid}
            for f in edit_fields:
                params[f] = fields.get(f)
            params["__attrs_partial"] = json.dumps(
                attrs_partial or {}, ensure_ascii=False,
            )
            try:
                res = session.execute(upd, params)
                row = res.first()
                if row is None:
                    report.skipped += 1
                    report.warnings.append(
                        f"{sheet_name}: id={cid} не найден в БД, строка пропущена"
                    )
                else:
                    report.updated += 1
            except Exception as exc:
                errors.append(ImportRowError(
                    sheet=sheet_name, row=0,
                    message=f"UPDATE id={cid}: {type(exc).__name__}: {exc}",
                ))
                raise

    # INSERT: batched executemany. NOT NULL DEFAULT поля → COALESCE с
    # SQL-defaultом (см. sheet_cfg.insert_defaults). attrs_jsonb —
    # собранный partial для новой строки.
    if rows_insert:
        defaults = sheet_cfg.get("insert_defaults") or {}
        cols = edit_fields + ["attrs_jsonb"]
        cols_sql = ", ".join(cols)
        value_parts: list[str] = []
        for f, ftype in edit_specs:
            if f in defaults:
                value_parts.append(_coalesce_param(f, ftype, defaults[f]))
            else:
                pg_t = _pg_type_for(ftype)
                if pg_t:
                    value_parts.append(f"CAST(:{f} AS {pg_t})")
                else:
                    value_parts.append(f":{f}")
        value_parts.append("CAST(:__attrs AS JSONB)")
        ins_sql = text(
            f"INSERT INTO {table} ({cols_sql}) "
            f"VALUES ({', '.join(value_parts)})"
        )

        params_list = []
        for fields, attrs_partial in rows_insert:
            d: dict[str, Any] = {f: fields.get(f) for f in edit_fields}
            d["__attrs"] = json.dumps(attrs_partial or {}, ensure_ascii=False)
            params_list.append(d)
        try:
            session.execute(ins_sql, params_list)
            report.inserted += len(params_list)
        except Exception as exc:
            errors.append(ImportRowError(
                sheet=sheet_name, row=0,
                message=f"INSERT batch ({len(params_list)} rows): {type(exc).__name__}: {exc}",
            ))
            raise


# ---------------------------------------------------------------------------
# Чтение листов и сбор изменений (общий слой)
# ---------------------------------------------------------------------------


def _process_pc_sheet(
    *,
    ws,
    sheet_cfg: dict,
    report: ImportReport,
    session: Session,
) -> None:
    sheet_name = sheet_cfg["sheet_name"]
    header_index = _build_header_index(ws, sheet_cfg["columns"])

    if not header_index:
        report.warnings.append(f"{sheet_name}: шапка не найдена, лист пропущен")
        return

    # Предупреждение о наличии read-only колонок (единоразовое, не на строку).
    ro_present = [
        h for h, db_field, kind, _t in sheet_cfg["columns"]
        if kind == "readonly" and h in header_index
    ]
    if ro_present:
        report.warnings.append(
            f"{sheet_name}: read-only columns ignored: {', '.join(ro_present)}"
        )

    rows_update: list[tuple[int, dict[str, Any]]] = []
    rows_insert: list[dict[str, Any]] = []
    insert_row_numbers: list[int] = []

    last_row = _last_data_row(ws)
    if last_row < _DATA_START_ROW:
        return  # пустой лист — нормально, идём дальше.

    errors_for_apply: list[ImportRowError] = []

    for row in range(_DATA_START_ROW, last_row + 1):
        fields, id_value = _read_pc_row(
            ws=ws, row=row, header_index=header_index,
            sheet_cfg=sheet_cfg, errors=report.errors,
        )
        if fields is None:
            # либо ошибка (уже в report.errors), либо пустая строка.
            continue

        if id_value is not None:
            rows_update.append((id_value, fields))
        else:
            # INSERT-путь: проверяем NOT NULL поля.
            missing = [
                f for f in sheet_cfg["required_for_insert"]
                if fields.get(f) is None
            ]
            if missing:
                report.errors.append(ImportRowError(
                    sheet=sheet_name, row=row,
                    message=f"INSERT: отсутствуют обязательные поля: {missing}",
                ))
                continue
            rows_insert.append(fields)
            insert_row_numbers.append(row)

    if not rows_update and not rows_insert:
        return

    _apply_pc_sheet(
        session=session, sheet_cfg=sheet_cfg,
        rows_update=rows_update, rows_insert=rows_insert,
        report=report, errors=errors_for_apply,
        insert_row_numbers=insert_row_numbers,
    )
    # ошибки SQL уже пробросились наружу через raise; сюда они не дойдут.


def _process_printer_sheet(
    *,
    ws,
    sheet_cfg: dict,
    report: ImportReport,
    session: Session,
) -> None:
    sheet_name = sheet_cfg["sheet_name"]
    header_index = _build_header_index(ws, sheet_cfg["columns"])

    if not header_index:
        report.warnings.append(f"{sheet_name}: шапка не найдена, лист пропущен")
        return

    ro_present = [
        h for h, db_field, kind, _t in sheet_cfg["columns"]
        if kind == "readonly" and h in header_index
    ]
    if ro_present:
        report.warnings.append(
            f"{sheet_name}: read-only columns ignored: {', '.join(ro_present)}"
        )

    rows_update: list[tuple[int, dict[str, Any], dict[str, Any]]] = []
    rows_insert: list[tuple[dict[str, Any], dict[str, Any]]] = []

    last_row = _last_data_row(ws)
    if last_row < _DATA_START_ROW:
        return

    for row in range(_DATA_START_ROW, last_row + 1):
        fields, attrs_partial, id_value = _read_printer_row(
            ws=ws, row=row, header_index=header_index,
            sheet_cfg=sheet_cfg, errors=report.errors,
        )
        if fields is None:
            continue

        if id_value is not None:
            rows_update.append((id_value, fields, attrs_partial or {}))
        else:
            missing = [
                f for f in sheet_cfg["required_for_insert"]
                if fields.get(f) is None
            ]
            if missing:
                report.errors.append(ImportRowError(
                    sheet=sheet_name, row=row,
                    message=f"INSERT: отсутствуют обязательные поля: {missing}",
                ))
                continue
            rows_insert.append((fields, attrs_partial or {}))

    if not rows_update and not rows_insert:
        return

    _apply_printer_sheet(
        session=session, sheet_cfg=sheet_cfg,
        rows_update=rows_update, rows_insert=rows_insert,
        report=report, errors=report.errors,
    )


# ---------------------------------------------------------------------------
# Точки входа
# ---------------------------------------------------------------------------


def import_components_pc(
    file_path: str | Path,
    user_id: int,
    *,
    session: Session | None = None,
) -> ImportReport:
    """Импорт файла «Комплектующие_ПК.xlsx».

    Читает 8 листов (CPU/Motherboard/RAM/GPU/Storage/Case/PSU/Cooler) и
    применяет UPDATE/INSERT в соответствующих таблицах. Read-only поля
    (цены, поставщик, даты) игнорируются.

    user_id — для логирования и audit_log (саму запись audit_log пишет
    роутер; здесь user_id не используется напрямую, но включён в API ради
    симметрии и будущих per-user проверок).
    """
    report = ImportReport()
    own_session = session is None
    if own_session:
        session = SessionLocal()
    try:
        wb = load_workbook(filename=str(file_path), read_only=False, data_only=True)
        try:
            for cat, cfg in PC_SHEETS.items():
                sheet_name = cfg["sheet_name"]
                if sheet_name not in wb.sheetnames:
                    report.warnings.append(
                        f"{sheet_name}: лист отсутствует в файле, пропущен"
                    )
                    continue
                ws = wb[sheet_name]
                _process_pc_sheet(
                    ws=ws, sheet_cfg=cfg, report=report, session=session,
                )
            session.commit()
        except Exception:
            session.rollback()
            logger.exception("Excel-import: SQL-ошибка, откатываем транзакцию")
            raise
        finally:
            wb.close()
    finally:
        if own_session:
            session.close()
    return report


def import_printers_mfu(
    file_path: str | Path,
    user_id: int,
    *,
    session: Session | None = None,
) -> ImportReport:
    """Импорт файла «Печатная_техника.xlsx» (листы Принтеры / МФУ).

    attrs_jsonb обновляется per-key merge'ом (`attrs_jsonb || partial`):
    пустая ячейка → ключ не трогается, 'n/a' → пишется 'n/a',
    непустое значение → валидируется и пишется.
    """
    report = ImportReport()
    own_session = session is None
    if own_session:
        session = SessionLocal()
    try:
        wb = load_workbook(filename=str(file_path), read_only=False, data_only=True)
        try:
            for cat, cfg in PRINTER_SHEETS.items():
                sheet_name = cfg["sheet_name"]
                if sheet_name not in wb.sheetnames:
                    report.warnings.append(
                        f"{sheet_name}: лист отсутствует в файле, пропущен"
                    )
                    continue
                ws = wb[sheet_name]
                _process_printer_sheet(
                    ws=ws, sheet_cfg=cfg, report=report, session=session,
                )
            session.commit()
        except Exception:
            session.rollback()
            logger.exception("Excel-import: SQL-ошибка, откатываем транзакцию")
            raise
        finally:
            wb.close()
    finally:
        if own_session:
            session.close()
    return report
