# Адаптер OCS: чтение прайса с листа «Наличие и цены».
#
# Отличия от монолитного price_loader.py (до этапа 7):
#   - здесь ТОЛЬКО парсинг, без записи в БД (этим занимается orchestrator);
#   - дополнительно читается колонка EAN128 → PriceRow.gtin
#     (раньше GTIN не использовался вообще);
#   - категории определяются по тем же парам (колонка B, колонка C),
#     что и раньше — маппинг не менялся, чтобы не ломать существующую БД.

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from portal.services.configurator.price_loaders.base import BasePriceLoader
from portal.services.configurator.price_loaders.models import PriceRow

logger = logging.getLogger(__name__)


# (колонка B, колонка C) → (our_category, исторически имя таблицы).
# Формат совпадает с прежним _CATEGORY_MAP, чтобы ничего не сломать.
_CATEGORY_MAP: dict[tuple[str, str | None], str] = {
    ("Процессоры",                       None): "cpu",
    ("Материнские платы",                None): "motherboard",
    ("Оперативная память",               None): "ram",
    ("Видеокарты",                       None): "gpu",
    ("Накопители информации", "Жёсткие диски"):           "storage",
    ("Накопители информации", "Твердотельные накопители"): "storage",
    ("Корпуса",                          None): "case",
    ("Блоки питания",                    None): "psu",
    ("Системы охлаждения для ПК", "Воздушное охлаждение для процессоров"): "cooler",
    ("Системы охлаждения для ПК",
     "Системы жидкостного охлаждения «всё-в-одном» для процессоров"):     "cooler",
}


# Печатная техника OCS (Этап 4 слияния, 2026-05-08).
# Решение по (B, C) — точные значения из реальных прайсов OCS:
# колонка B = тип группы («Принтеры»/«МФУ»), C = подтип
# («Принтеры лазерные»/«МФУ струйные» и т.п.). Запись в БД пока
# не подключена — orchestrator скипнет такие позиции с инкрементом
# pending_printers_mfu (Этап 6 даст таблицу `printers_mfu`).
_BC_PRINTER_MFU_MAP: dict[tuple[str, str], str] = {
    ("Принтеры", "Принтеры лазерные"):  "printer",
    ("Принтеры", "Принтеры струйные"):  "printer",
    ("Принтеры", "Принтеры матричные"): "ignore",
    ("МФУ",      "МФУ лазерные"):       "mfu",
    ("МФУ",      "МФУ струйные"):       "mfu",
    ("МФУ",      "МФУ матричные"):      "ignore",
}


def _classify_ocs(cat_b: str, kind_c: str) -> str:
    """Определяет печатную категорию OCS по (B, C). 'printer' / 'mfu'
    идут в pending_printers_mfu в orchestrator; 'ignore' для
    matрричных и неизвестных пар (для отладки прайса)."""
    if (cat_b, kind_c) in _BC_PRINTER_MFU_MAP:
        return _BC_PRINTER_MFU_MAP[(cat_b, kind_c)]
    logger.info(
        "OCS (B=%r, C=%r): классифицирован как ignore (неизвестная пара)",
        cat_b, kind_c,
    )
    return "ignore"


# Индексы колонок (0-based) в листе «Наличие и цены».
# A=cat_a, B=cat_b, C=kind_c, D=maker, E=supplier_sku, G=mpn, H=name,
# I=price, J=currency, L=stock, R=transit (итого минимум 18 колонок).
# EAN128 в листе OCS идёт после основной группы и определяется по
# заголовку (см. _find_ean_column ниже) — индекс у разных версий прайса
# может отличаться, поэтому хардкодить небезопасно.
_COL_CAT_B        = 1
_COL_KIND_C       = 2
_COL_MAKER        = 3
_COL_SUPPLIER_SKU = 4    # E
_COL_SKU          = 6    # G — MPN
_COL_NAME         = 7
_COL_PRICE        = 8
_COL_CURRENCY     = 9
_COL_STOCK        = 11
_COL_TRANSIT      = 17


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


def _resolve_category(cat_b: str, kind_c: str) -> str | None:
    """Сначала ПК-карта (точная (B, C), затем (B, None)); если ПК-карта
    промахнулась — пробуем печатные пары (Этап 4 слияния).

    Возвращает 'cpu'/.../'cooler' (ПК) или 'printer'/'mfu' (печать;
    orchestrator пока пропустит с pending_printers_mfu) или None.
    'ignore'-результат _classify_ocs (матричные / неизвестные пары)
    приводим к None — семантика «не пишем в БД» в C-PC2 одна.
    """
    b = (cat_b or "").strip()
    c = (kind_c or "").strip()
    hit = _CATEGORY_MAP.get((b, c if c else None))
    if hit:
        return hit
    hit = _CATEGORY_MAP.get((b, None))
    if hit:
        return hit
    if c:
        cat = _classify_ocs(b, c)
        if cat in ("printer", "mfu"):
            return cat
    return None


def _normalize_gtin(value) -> str | None:
    """EAN128 из OCS может прийти как число (12 цифр без ведущих нулей),
    как строка, или как пустое значение. Приводим к строке и оставляем
    только цифры — чтобы одинаково легло в VARCHAR(20)."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Float-представление 1.23456789012e12 — защита, если Excel решил
    # отдать число в экспоненциальной записи.
    if "e" in s.lower():
        try:
            s = str(int(Decimal(s)))
        except InvalidOperation:
            return None
    digits = re.sub(r"\D", "", s)
    return digits or None


def _find_ean_column(header_row: tuple) -> int | None:
    """Возвращает индекс колонки EAN128 по заголовку или None.

    EAN128 — глобальный штрихкод товара, у OCS он лежит в одной из
    крайних колонок (R/S/T в разных версиях прайса). Вместо хардкода
    индекса ищем по имени — это стабильнее.
    """
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().upper().replace(" ", "")
        if name in {"EAN128", "EAN", "GTIN"}:
            return idx
    return None


class OcsLoader(BasePriceLoader):
    supplier_name = "OCS"

    @classmethod
    def detect(cls, filename: str) -> bool:
        return "ocs" in filename.lower()

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Наличие и цены"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист «{sheet_name}» не найден в файле {filepath}. "
                f"Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        # Первая строка — заголовки. Ищем индекс EAN128 по имени.
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ean_idx = _find_ean_column(header)
        if ean_idx is None:
            # Нормально: старые версии прайса могут не иметь EAN128,
            # gtin останется None — это не ошибка, просто меньше инфы.
            logger.info("В прайсе OCS колонка EAN128 не найдена — GTIN не будет заполняться.")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            cat_b        = str(_cell(row, _COL_CAT_B)  or "").strip()
            kind_c       = str(_cell(row, _COL_KIND_C) or "").strip()
            manufacturer = str(_cell(row, _COL_MAKER)  or "").strip() or None
            supplier_sku = str(_cell(row, _COL_SUPPLIER_SKU) or "").strip()
            mpn          = str(_cell(row, _COL_SKU)   or "").strip()
            name         = str(_cell(row, _COL_NAME)  or "").strip()
            price_raw    =     _cell(row, _COL_PRICE)
            currency_raw =     _cell(row, _COL_CURRENCY)
            stock_raw    =     _cell(row, _COL_STOCK)
            transit_raw  =     _cell(row, _COL_TRANSIT)
            ean_raw      =     _cell(row, ean_idx) if ean_idx is not None else None

            our_category = _resolve_category(cat_b, kind_c)

            price = _parse_price(price_raw)
            if price is None:
                # Без цены строка бесполезна — orchestrator её пропустит,
                # но для OCS мы и раньше так же делали через counters['errors'].
                continue

            currency = (str(currency_raw).strip().upper() if currency_raw else "RUB")[:3] or "RUB"

            # В прайсе OCS каталожный номер (MPN) был обязательным.
            # Сохраняем это правило: если MPN пуст — пропускаем строку.
            if not mpn:
                continue

            # У OCS supplier_sku (колонка E) может быть пустым — тогда
            # в supplier_prices.supplier_sku запишется NULL.
            yield PriceRow(
                supplier_sku=supplier_sku or "",
                mpn=mpn,
                gtin=_normalize_gtin(ean_raw),
                brand=manufacturer,
                raw_category=" | ".join(x for x in (cat_b, kind_c) if x),
                our_category=our_category,
                name=name,
                price=price,
                currency=currency,
                stock=_parse_int(stock_raw),
                transit=_parse_int(transit_raw),
                row_number=row_idx,
            )
