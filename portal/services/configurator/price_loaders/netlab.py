# Адаптер Netlab: чтение прайса дилерской категории D «DealerD.xlsx».
#
# Особенности формата:
#   - Лист «Цены» (есть ещё «Уцененные товары» — не берём, это розница
#     с произвольными уценками, в нашем каталоге не нужна).
#   - Сверху листа служебный блок (контакты, описание тарифов B/C/D/E/F/R,
#     текущий курс), затем последовательно идут «секции» товаров.
#   - В каждой секции своя строка-разделитель и своя строка заголовков:
#       «Бусиново | (пусто) | PartNumber | Артикул | Наименование |
#        B | C | D | E | F | R | РРЦ(Руб.) | Вес,кг | Объём,м^3 | Гарантия».
#     Первая такая строка заголовков — №21; дальше по всему листу
#     заголовки повторяются — определяем их по литералу «PartNumber»
#     в колонке C и пропускаем как обычные данные.
#   - Строка-разделитель категории: заполнена только колонка E
#     (Наименование), без PartNumber/Артикул и без числовых цен.
#     Категории однострочные («Материнские платы ASUS (для INTEL)»,
#     «SSD Kingston», «Корпуса CHIEFTEC» и т. п.) — иерархии нет,
#     поэтому маппинг строится по ключевым словам, а не по точному
#     совпадению строки. Серверные/внешние позиции отфильтровываются.
#   - C «PartNumber»  — MPN (партномер производителя).
#   - D «Артикул»     — внутренний SKU Netlab (supplier_sku).
#   - E «Наименование» — name.
#   - F..K (B/C/D/E/F/R) — шесть тарифов (B самый «дорогой» дилерский,
#     F — самый «дешёвый», R — розница). Берём колонку D (тариф D)
#     как стандартный дилерский. Все шесть колонок — в USD (явно
#     написано в служебной шапке листа).
#   - L «РРЦ(Руб.)»   — рекомендованная розничная в рублях. Заполнена
#     не у всех позиций; используем как fallback, если D-USD пуст.
#   - A «Бусиново»    — бинарный маркер остатка: «+» (есть) / «-» (нет).
#     Числовые остатки в этой колонке не встречаются, поэтому «+» → 5,
#     «-» → 0 (того же класса решение, что для буквенных маркеров
#     Treolan/Merlion).
#   - GTIN в прайсе Netlab нет.
#
# Опционально: путь к .zip-архиву распаковывается во временный каталог
# (Netlab прислал прайс именно в виде «dealerd.zip → DealerD.xlsx»).
# Если путь — обычный .xlsx, всё работает без изменений.

from __future__ import annotations

import logging
import os
import tempfile
import zipfile
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from portal.services.configurator.price_loaders.base import BasePriceLoader
from portal.services.configurator.price_loaders.models import PriceRow

logger = logging.getLogger(__name__)


# Индексы колонок на листе «Цены» (0-based). Соответствуют буквам
# Excel: A=0 «Бусиново», C=2 «PartNumber», D=3 «Артикул»,
# E=4 «Наименование», F=5 «B», G=6 «C», H=7 «D», L=11 «РРЦ(Руб.)».
_COL_STOCK_FLAG  = 0   # A — «+»/«-»
_COL_PARTNUMBER  = 2   # C — MPN
_COL_ARTICLE     = 3   # D — supplier_sku
_COL_NAME        = 4   # E — название (и текст разделителя категорий)
_COL_PRICE_D_USD = 7   # H — ценовая колонка тарифа «D» (USD)
_COL_PRICE_RRC   = 11  # L — РРЦ(Руб.)

HEADER_ROW = 21
DATA_START_ROW = 22

# Поля прайса узкие — крайние колонки иногда могут быть схлопнуты
# stream-режимом openpyxl (см. read_only). Принудительно читаем
# первые 15 колонок, чтобы не терять РРЦ.
_MAX_COL = 15


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


# Бинарные маркеры остатка у Netlab. Других значений в колонке A нет
# (~71k строк, только «+» и «-»). Без перевода в число конфигуратор
# (фильтр stock_qty > 0) вообще не увидел бы Netlab-прайс.
_NETLAB_QUAL_STOCK: dict[str, int] = {
    "+": 5,
    "-": 0,
}


def _parse_stock(value) -> int:
    """Остаток с учётом бинарных маркеров «+»/«-»."""
    if value is None:
        return 0
    s = str(value).strip()
    if not s:
        return 0
    if s in _NETLAB_QUAL_STOCK:
        return _NETLAB_QUAL_STOCK[s]
    return _parse_int(value)


def _normalize(s) -> str:
    if s is None:
        return ""
    # Excel часто отдаёт числовые SKU как float (11051003 → 11051003.0).
    # Между загрузками это даёт «дубликаты» одного и того же SKU
    # с разным написанием. Сворачиваем целочисленные float к строке
    # без хвостового «.0».
    if isinstance(s, float) and s.is_integer():
        return str(int(s))
    return str(s).strip()


def _is_repeated_header(row: tuple) -> bool:
    """Внутри листа Netlab заголовок «PartNumber|Артикул|Наименование|…»
    повторяется в начале каждой брендовой подсекции. По литералу
    «PartNumber» в колонке C точно отличаем такую строку от данных."""
    return _normalize(_cell(row, _COL_PARTNUMBER)) == "PartNumber"


def _is_category_separator(row: tuple) -> str | None:
    """Строка-разделитель категории: заполнена только колонка E
    (Наименование), без PartNumber/Артикул и без числовых цен."""
    name = _normalize(_cell(row, _COL_NAME))
    if not name:
        return None
    if _normalize(_cell(row, _COL_PARTNUMBER)) or _normalize(_cell(row, _COL_ARTICLE)):
        return None
    # Если в любой ценовой колонке есть число — это уже строка данных.
    for idx in (_COL_PRICE_D_USD, _COL_PRICE_RRC):
        if _parse_price(_cell(row, idx)) is not None:
            return None
    return name


# Маппинг категорий Netlab по ключевым словам.
#
# Прайс плоский, каждая «подсекция бренда» — отдельная строка-разделитель
# («Корпуса AEROCOOL», «SSD Kingston», «Память DDR4», «Видеокарты ASUS»).
# Энумерация всех ~150 строк нестабильна — Netlab регулярно добавляет
# и удаляет бренды, поэтому матчим по ключевым словам с явным списком
# исключений (серверные/внешние позиции, а также периферия типа
# «Охлаждающие подставки для ноутбуков»).
def _resolve_category(separator: str) -> str | None:
    s = separator.lower().replace("ё", "е")

    # Исключения — то, что НЕ наша категория, даже если ключевое слово совпало.
    if "серверн" in s:
        # «Серверные материнские платы», «Память серверная DDR4», «SSD ... серверные»
        return None
    if "внешн" in s:
        # «Внешние HDD/SSD», «Внешние контейнеры»
        return None
    # Брендовые секции серверных вендоров: HPE, Dell, IBM, Lenovo (server),
    # Huawei, Fujitsu, Supermicro, Chenbro, Procase, Gooxi — у них в этом
    # прайсе только серверная номенклатура (CPU/RAM/SSD/корпуса). Слово
    # «серверн» в строке-разделителе не пишется, поэтому фильтруем по бренду.
    _server_vendor_prefixes = (
        "hpe ", "hp ", "dell ", "ibm ", "lenovo ", "huawei ",
        "fujitsu ", "supermicro ", "chenbro ", "procase ", "gooxi ",
    )
    if s.startswith(_server_vendor_prefixes):
        return None
    if "подставк" in s:
        # «Охлаждающие подставки для ноутбуков»
        return None
    if "монтажн" in s or "электрик" in s:
        # «Блок питания/Монтажные материалы», «Блок питания/Электрика»
        return None
    if "монобл" in s and "корпус" in s:
        # «Корпуса под моноблоки PowerCool»
        return None
    if "пылев" in s or "фильтр" in s or "рельс" in s or "аксессуар" in s:
        # «Корпусные пылевые фильтры», «Рельсы для всех видов корпусов»
        return None

    if "процессор" in s:
        return "cpu"
    if "материнск" in s and "плат" in s:
        return "motherboard"
    if "видеокарт" in s or "видеокарта" in s:
        return "gpu"
    if "ddr" in s or "памят" in s:
        # На этом этапе уже отсеяли «серверная»/«для ноутбуков» — нет,
        # «для ноутбуков» НЕ отсеивали: SO-DIMM модули у Merlion тоже
        # помечены ram. Оставляем.
        return "ram"
    if "ssd" in s:
        return "storage"
    if ("жесткий" in s or "hdd" in s) and "диск" in s:
        return "storage"
    # «Блоки питания» проверяем РАНЬШЕ «корпус», иначе строка
    # «Блоки питания к корпусам» уйдёт в case вместо psu.
    # «блок» + «питани» вместо «блок питания» — чтобы захватывать
    # и единственное число, и множественное («Блоки питания»).
    if "блок" in s and "питани" in s:
        return "psu"
    if "корпус" in s:
        return "case"
    if "охлажда" in s or "вентилятор" in s or "кулер" in s:
        return "cooler"
    return None


def _open_workbook(filepath: str):
    """Возвращает (workbook, путь_к_временной_папке_или_None).
    Если filepath — .zip, распаковываем единственный .xlsx внутри
    во временный каталог. Вызывающий обязан удалить временный каталог
    через _cleanup_tmp.
    """
    lower = filepath.lower()
    if lower.endswith(".zip"):
        tmpdir = tempfile.mkdtemp(prefix="netlab_zip_")
        try:
            with zipfile.ZipFile(filepath) as zf:
                xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
                if not xlsx_names:
                    raise ValueError(
                        f"В архиве {filepath} нет ни одного .xlsx файла."
                    )
                if len(xlsx_names) > 1:
                    logger.warning(
                        "В архиве %s несколько .xlsx (%s) — используем первый.",
                        filepath, xlsx_names,
                    )
                inner_name = xlsx_names[0]
                zf.extract(inner_name, tmpdir)
                inner_path = os.path.join(tmpdir, inner_name)
            wb = load_workbook(inner_path, read_only=True, data_only=True)
            return wb, tmpdir
        except Exception:
            _cleanup_tmp(tmpdir)
            raise
    wb = load_workbook(filepath, read_only=True, data_only=True)
    return wb, None


def _cleanup_tmp(tmpdir: str | None) -> None:
    if not tmpdir:
        return
    try:
        for root, dirs, files in os.walk(tmpdir, topdown=False):
            for f in files:
                try:
                    os.remove(os.path.join(root, f))
                except OSError:
                    pass
            for d in dirs:
                try:
                    os.rmdir(os.path.join(root, d))
                except OSError:
                    pass
        os.rmdir(tmpdir)
    except OSError:
        pass


class NetlabLoader(BasePriceLoader):
    supplier_name = "Netlab"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "netlab" in name or "dealerd" in name or name.startswith("dealer_d")

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb, tmpdir = _open_workbook(filepath)
        try:
            sheet_name = "Цены"
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Лист «{sheet_name}» не найден в файле {filepath}. "
                    f"Доступные листы: {wb.sheetnames}"
                )
            ws = wb[sheet_name]

            # У реального DealerD.xlsx (~77k строк) элемент <dimension>
            # внутри XML листа повреждён и сообщает «A1:A1». В режиме
            # read_only openpyxl этому верит и не отдаёт ни одной строки.
            # `reset_dimensions()` форсирует сканирование всех строк
            # листа целиком — без него мы получаем «загружено 0 строк»
            # из 73k реальных позиций.
            ws.reset_dimensions()

            current_raw_category: str = ""
            current_our_category: str | None = None

            for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=DATA_START_ROW,
                    max_col=_MAX_COL,
                    values_only=True,
                ),
                start=DATA_START_ROW,
            ):
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue

                # Повторяющийся заголовок секции — пропускаем без сброса категории.
                if _is_repeated_header(row):
                    continue

                # Строка-разделитель категории.
                sep = _is_category_separator(row)
                if sep is not None:
                    current_raw_category = sep
                    current_our_category = _resolve_category(sep)
                    continue

                partnumber  = _normalize(_cell(row, _COL_PARTNUMBER))
                article     = _normalize(_cell(row, _COL_ARTICLE))
                name        = _normalize(_cell(row, _COL_NAME))
                price_usd   = _parse_price(_cell(row, _COL_PRICE_D_USD))
                price_rrc   = _parse_price(_cell(row, _COL_PRICE_RRC))
                stock       = _parse_stock(_cell(row, _COL_STOCK_FLAG))

                # Без артикула Netlab (D) или PartNumber (C) — у нас нет
                # ни supplier_sku, ни mpn, такая строка бесполезна.
                if not article and not partnumber:
                    continue
                if not article:
                    # У Netlab свой SKU присутствует у каждой реальной
                    # позиции. Пустота сигналит о хвостовом мусоре.
                    logger.warning(
                        "Netlab строка %d: пустой Артикул — строка пропущена.",
                        row_idx,
                    )
                    continue

                # Цена: приоритет — D-USD (стандартный дилерский тариф),
                # fallback — РРЦ в рублях.
                if price_usd is not None:
                    price = price_usd
                    currency = "USD"
                elif price_rrc is not None:
                    price = price_rrc
                    currency = "RUB"
                else:
                    continue

                yield PriceRow(
                    supplier_sku=article,
                    mpn=partnumber or None,
                    gtin=None,
                    brand=None,  # бренд у Netlab «зашит» в название категории
                    raw_category=current_raw_category,
                    our_category=current_our_category,
                    name=name,
                    price=price,
                    currency=currency,
                    stock=stock,
                    transit=0,  # отдельной колонки «транзит» в DealerD нет
                    row_number=row_idx,
                )
        finally:
            try:
                wb.close()
            except Exception:
                pass
            _cleanup_tmp(tmpdir)
