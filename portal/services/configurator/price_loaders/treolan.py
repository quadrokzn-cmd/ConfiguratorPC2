# Адаптер Treolan: чтение прайса «23_04_2026_catalog__1_.xlsx».
#
# Особенности формата:
#   - Лист «Каталог».
#   - Заголовки в строке 3; данные начинаются со строки 4.
#   - Категории НЕ в отдельных колонках, а в «разделительных» строках:
#     в такой строке заполнена ТОЛЬКО колонка A и её значение содержит
#     «->» (например «Комплектующие->Процессоры»). Парсер помнит
#     последнюю виденную категорию и навешивает её на следующие строки
#     с данными, пока не встретит новую строку-разделитель.
#   - A «Артикул»     — MPN (партномер производителя). Своего номера
#                        у Treolan нет, поэтому supplier_sku = mpn.
#   - B «Наименование» — name.
#   - C «Производитель» — brand.
#   - D «Склад»        — остаток.
#   - E «Транзит», F «Б.Тр.» — транзит (суммируем).
#   - G «Цена*»        — цена в USD.
#   - H «Цена руб.**»  — цена в RUB (приоритетная для системы).
#   - J «Код GTIN»     — штрихкод товара. КРИТИЧНО для Intel CPU:
#                         в Treolan артикул=S-Spec (например SRMBG),
#                         а в OCS — Order Code (CM8071512400F).
#                         Match по MPN не сработает, только по GTIN.

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from portal.services.configurator.price_loaders._qual_stock import TREOLAN_QUAL_STOCK
from portal.services.configurator.price_loaders.base import BasePriceLoader
from portal.services.configurator.price_loaders.models import PriceRow

logger = logging.getLogger(__name__)


# Путь категории от Treolan → our_category.
_CATEGORY_MAP: dict[str, str] = {
    "Комплектующие->Процессоры": "cpu",
    "Комплектующие->Материнские платы->Материнские платы для процесоров AMD клиентские":   "motherboard",
    "Комплектующие->Материнские платы->Материнские платы для процесоров Intel клиентские": "motherboard",
    "Комплектующие->Оперативная память->Память DDR3 клиентская": "ram",
    "Комплектующие->Оперативная память->Память DDR4 клиентская": "ram",
    "Комплектующие->Оперативная память->Память DDR5 клиентская": "ram",
    "Комплектующие->Видеокарты->Видеокарты на чипсетах NVIDIA":  "gpu",
    "Комплектующие->Видеокарты->Видеокарты на чипсетах ATI":     "gpu",
    "Комплектующие->Видеокарты->Видеокарты на чипсетах Intel":   "gpu",
    "Комплектующие->Твердотельные накопители SSD->SSD NVMe внутренние": "storage",
    "Комплектующие->Твердотельные накопители SSD->SSD SATA внутренние": "storage",
    "Комплектующие->Жесткие диски->HDD SATA внутренние":         "storage",
    "Комплектующие->Корпуса":                                    "case",
    "Комплектующие->БП для корпусов":                            "psu",
    "Комплектующие->Системы охлаждения":                         "cooler",
}


# Печатная техника Treolan (Этап 4 слияния, 2026-05-08).
# Структура путей под печать у Treolan: корень «Принтеры, сканеры, МФУ»;
# вторым уровнем идут «Принтеры», «МФУ», «Сканеры», «Аксессуары»,
# «Широкоформатные Принтеры/Плоттеры», «Широкоформатные МФУ» и т.д.
# Аксессуары/сканеры — out-of-scope (ignore). Запись printer/mfu в БД
# пока не подключена — orchestrator скипнет с pending_printers_mfu
# (Этап 6 даст таблицу `printers_mfu`).
_TREOLAN_PRINTER_ROOT = "Принтеры, сканеры, МФУ"
_TREOLAN_WIDE_PRINTER_PREFIXES: tuple[str, ...] = (
    f"{_TREOLAN_PRINTER_ROOT}->Широкоформатные Принтеры",
    f"{_TREOLAN_PRINTER_ROOT}->Широкоформатные Принтеры/Плоттеры",
)
_TREOLAN_WIDE_MFU_PREFIX = f"{_TREOLAN_PRINTER_ROOT}->Широкоформатные МФУ"


def _classify_treolan(path: str) -> str:
    """Категоризация по полному пути Treolan для печатных позиций.

    Возвращает 'printer'/'mfu'/'ignore'. 'ignore' — для сканеров,
    аксессуаров и неизвестных подкатегорий (для отладки прайса).
    Применяется только когда путь начинается с `_TREOLAN_PRINTER_ROOT`;
    иначе вызывающий код возвращает None (это «не наша» категория).
    """
    if any(path.startswith(p) for p in _TREOLAN_WIDE_PRINTER_PREFIXES):
        return "printer"
    if path.startswith(_TREOLAN_WIDE_MFU_PREFIX):
        return "mfu"
    parts = path.split("->")
    if len(parts) >= 2 and parts[0] == _TREOLAN_PRINTER_ROOT:
        middle = parts[1]
        if middle == "Принтеры":
            return "printer"
        if middle == "МФУ":
            return "mfu"
    logger.info("Treolan path=%r: классифицирован как ignore", path)
    return "ignore"


def _resolve_category(path: str) -> str | None:
    """Единый вход в категоризацию Treolan: ПК-карта → печатная карта → None.

    Возвращает 'cpu'/.../'cooler' (ПК), 'printer'/'mfu' (печать; orchestrator
    обрабатывает как pending_printers_mfu до Этапа 6) или None.
    """
    pc = _CATEGORY_MAP.get(path)
    if pc is not None:
        return pc
    if path.startswith(_TREOLAN_PRINTER_ROOT):
        cat = _classify_treolan(path)
        if cat in ("printer", "mfu"):
            return cat
    return None


_COL_ARTICLE    = 0  # A
_COL_NAME       = 1  # B
_COL_BRAND      = 2  # C
_COL_STOCK      = 3  # D
_COL_TRANSIT_1  = 4  # E
_COL_TRANSIT_2  = 5  # F
_COL_PRICE_USD  = 6  # G
_COL_PRICE_RUB  = 7  # H
_COL_GTIN       = 9  # J

HEADER_ROW = 3
DATA_START_ROW = 4


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


# Буквенные маркеры остатка Treolan («<10», «много», «>10», «>100»)
# теперь живут в shared-модуле _qual_stock — общем с REST-API парсером.
# Локальный alias оставлен для обратной совместимости тестов, которые
# импортировали _TREOLAN_QUAL_STOCK напрямую.
_TREOLAN_QUAL_STOCK = TREOLAN_QUAL_STOCK


def _parse_stock(value) -> int:
    """Остаток с учётом буквенных маркеров Treolan.

    Возвращает int >= 0. Приоритет — таблица маркеров, затем число.
    Нормализуем lower() и убираем пробелы, чтобы «< 10» и «<10» сошлись.
    """
    if value is None:
        return 0
    s = str(value).strip().lower().replace(" ", "")
    if not s:
        return 0
    if s in TREOLAN_QUAL_STOCK:
        return TREOLAN_QUAL_STOCK[s]
    return _parse_int(value)


def _normalize(s) -> str:
    return (str(s).strip() if s is not None else "")


def _normalize_gtin(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if "e" in s.lower():
        try:
            s = str(int(Decimal(s)))
        except InvalidOperation:
            return None
    digits = re.sub(r"\D", "", s)
    return digits or None


def _is_category_separator(row: tuple) -> str | None:
    """Строка-разделитель категории: заполнена ТОЛЬКО колонка A и
    её значение содержит «->». Возвращает эту строку или None, если
    это обычная строка данных."""
    a_val = _cell(row, _COL_ARTICLE)
    if a_val is None:
        return None
    a_str = str(a_val).strip()
    if not a_str or "->" not in a_str:
        return None
    # Проверяем, что другие значимые колонки пусты. Если заполнены
    # name/brand/price — это уже обычная строка данных.
    for idx in (_COL_NAME, _COL_BRAND, _COL_PRICE_USD, _COL_PRICE_RUB):
        if _normalize(_cell(row, idx)):
            return None
    return a_str


class TreolanLoader(BasePriceLoader):
    supplier_name = "Treolan"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "treolan" in name or "catalog" in name

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Каталог"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист «{sheet_name}» не найден в файле {filepath}. "
                f"Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        current_raw_category: str = ""
        current_our_category: str | None = None

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            # Обнаружили строку-разделитель → обновляем текущую категорию,
            # но сами данные из неё не выдаём.
            sep = _is_category_separator(row)
            if sep is not None:
                current_raw_category = sep
                current_our_category = _resolve_category(sep)
                continue

            article = _normalize(_cell(row, _COL_ARTICLE))
            name    = _normalize(_cell(row, _COL_NAME))
            brand   = _normalize(_cell(row, _COL_BRAND)) or None
            stock   = _parse_stock(_cell(row, _COL_STOCK))
            transit = _parse_stock(_cell(row, _COL_TRANSIT_1)) + _parse_stock(_cell(row, _COL_TRANSIT_2))
            price_usd = _parse_price(_cell(row, _COL_PRICE_USD))
            price_rub = _parse_price(_cell(row, _COL_PRICE_RUB))
            gtin = _normalize_gtin(_cell(row, _COL_GTIN))

            # Строка без артикула и без имени — пропускаем (бывает в хвосте).
            if not article and not name:
                continue
            # Нет артикула — орchestrator не сможет писать в supplier_prices
            # (нет supplier_sku), пропускаем с предупреждением.
            if not article:
                logger.warning(
                    "Treolan строка %d: пустой артикул — строка пропущена.",
                    row_idx,
                )
                continue

            if price_rub is not None:
                price = price_rub
                currency = "RUB"
            elif price_usd is not None:
                price = price_usd
                currency = "USD"
            else:
                continue

            yield PriceRow(
                # У Treolan артикул = и supplier_sku, и mpn.
                supplier_sku=article,
                mpn=article,
                gtin=gtin,
                brand=brand,
                raw_category=current_raw_category,
                our_category=current_our_category,
                name=name,
                price=price,
                currency=currency,
                stock=stock,
                transit=transit,
                row_number=row_idx,
            )
