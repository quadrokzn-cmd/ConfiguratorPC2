# Интеграционные тесты Excel-генератора и роутера экспорта (этап 8.1).
#
# Используют фикстуры test_web/conftest.py — реальная тестовая БД, клиенты
# с логином. exchange_rate мокаем, чтобы не ходить к ЦБ РФ во время тестов.

from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

import pytest
from openpyxl import load_workbook
from sqlalchemy import text as _t

from app.services import spec_service
from app.services.export import excel_builder


# ---- helpers: сид минимальных компонентов + вариант ---------------------

def _seed_user(db_session, *, login: str = "xl-manager") -> int:
    from app.auth import hash_password
    row = db_session.execute(
        _t(
            "INSERT INTO users (login, password_hash, role, name) "
            "VALUES (:l, :p, 'manager', 'XL Manager') RETURNING id"
        ),
        {"l": login, "p": hash_password("x")},
    ).first()
    db_session.commit()
    return int(row.id)


def _insert_cpu(db_session, *, sku: str = "CPU-SKU", gtin: str | None = "0101010101010") -> int:
    row = db_session.execute(
        _t(
            "INSERT INTO cpus "
            "  (sku, manufacturer, model, socket, cores, threads, "
            "   base_clock_ghz, turbo_clock_ghz, tdp_watts, gtin) "
            "VALUES "
            "  (:sku, 'Intel Corporation', 'Intel Core i5-12400F', 'LGA1700', "
            "   6, 12, 2.5, 4.4, 65, :gtin) RETURNING id"
        ),
        {"sku": sku, "gtin": gtin},
    ).first()
    db_session.commit()
    return int(row.id)


def _insert_ram(db_session, *, sku: str = "RAM-SKU") -> int:
    row = db_session.execute(
        _t(
            "INSERT INTO rams "
            "  (sku, manufacturer, model, memory_type, form_factor, "
            "   module_size_gb, modules_count, frequency_mhz, gtin) "
            "VALUES "
            "  (:sku, 'Kingston', 'Kingston 8GB DDR4', 'DDR4', 'DIMM', "
            "   8, 2, 3200, '0202020202020') RETURNING id"
        ),
        {"sku": sku},
    ).first()
    db_session.commit()
    return int(row.id)


def _make_query_with_cpu_ram(
    db_session,
    *,
    project_id: int,
    user_id: int,
    cpu_id: int,
    ram_id: int,
    manufacturer: str = "Intel",
    total_usd: float = 220.0,
    total_rub: float = 19800.0,
) -> int:
    build_result = {
        "status": "ok",
        "variants": [
            {
                "manufacturer": manufacturer,
                "path_used":    "default",
                "used_transit": False,
                "total_usd":    total_usd,
                "total_rub":    total_rub,
                "components": [
                    {
                        "category":      "cpu",
                        "component_id":  cpu_id,
                        "model":         "Intel Core i5-12400F",
                        "sku":           "BX8071512400F",
                        "manufacturer":  "Intel",
                        "quantity":      1,
                        "price_usd":     180.0,
                        "price_rub":     16200.0,
                        "supplier":      "OCS",
                        "supplier_sku":  "sup-cpu",
                    },
                    {
                        "category":      "ram",
                        "component_id":  ram_id,
                        "model":         "Kingston 8GB DDR4",
                        "sku":           "KVR32N22S8/8",
                        "manufacturer":  "Kingston",
                        "quantity":      2,
                        "price_usd":     20.0,
                        "price_rub":     1800.0,
                        "supplier":      "OCS",
                        "supplier_sku":  "sup-ram",
                    },
                ],
                "warnings": [],
            },
        ],
        "refusal_reason": None,
        "usd_rub_rate":   90.0,
        "fx_source":      "fallback",
    }
    row = db_session.execute(
        _t(
            "INSERT INTO queries "
            "  (project_id, user_id, raw_text, build_result_json, status, "
            "   cost_usd, cost_rub) "
            "VALUES (:pid, :uid, :rt, CAST(:br AS JSONB), 'ok', 0, 0) "
            "RETURNING id"
        ),
        {
            "pid": project_id, "uid": user_id,
            "rt":  "тест",
            "br":  json.dumps(build_result, ensure_ascii=False),
        },
    ).first()
    db_session.commit()
    return int(row.id)


# ---- тесты excel_builder -----------------------------------------------

def test_build_xlsx_single_config(db_session):
    """Один проект + одна конфигурация → валидный xlsx с comp-строкой
    и строками компонентов."""
    uid = _seed_user(db_session)
    pid = spec_service.create_empty_project(
        db_session, user_id=uid, name="Тестовый проект № 1",
    )
    cpu_id = _insert_cpu(db_session)
    ram_id = _insert_ram(db_session)
    qid = _make_query_with_cpu_ram(
        db_session, project_id=pid, user_id=uid,
        cpu_id=cpu_id, ram_id=ram_id,
    )
    spec_service.select_variant(
        db_session, project_id=pid, query_id=qid,
        manufacturer="Intel", quantity=1,
    )

    xlsx = excel_builder.build_project_xlsx(
        project_id=pid, db=db_session,
        rate=Decimal("92.5000"), rate_date=date(2026, 4, 10),
    )

    # Валидный xlsx открывается openpyxl-ем.
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active

    # Шапка проекта в строке 1 содержит имя проекта.
    assert "Тестовый проект № 1" in (ws["A1"].value or "")
    # Курс и дата.
    assert ws["O1"].value == "10.04.2026"
    assert ws["O2"].value == pytest.approx(92.5)

    # Собираем «наименования» всех строк: найдём comp и 2 компонента.
    names = [ws.cell(row=r, column=4).value for r in range(4, 20)]
    comp_row = next((v for v in names if v and "Системный блок" in v), None)
    assert comp_row is not None, f"Нет строки «Системный блок», names={names}"
    cpu_row = next((v for v in names if v and "i5-12400F" in v), None)
    ram_row = next((v for v in names if v and "Kingston" in v), None)
    assert cpu_row is not None
    assert ram_row is not None


def test_build_xlsx_two_configs_two_comp_blocks(db_session):
    """Два выбранных варианта → две «comp»-строки."""
    uid = _seed_user(db_session, login="xl-two")
    pid = spec_service.create_empty_project(
        db_session, user_id=uid, name="Два блока",
    )
    cpu_a = _insert_cpu(db_session, sku="CPU-A", gtin="0000000000001")
    cpu_b = _insert_cpu(db_session, sku="CPU-B", gtin="0000000000002")
    ram_a = _insert_ram(db_session, sku="RAM-A")
    ram_b = _insert_ram(db_session, sku="RAM-B")
    q1 = _make_query_with_cpu_ram(
        db_session, project_id=pid, user_id=uid,
        cpu_id=cpu_a, ram_id=ram_a, total_usd=500,
    )
    q2 = _make_query_with_cpu_ram(
        db_session, project_id=pid, user_id=uid,
        cpu_id=cpu_b, ram_id=ram_b, total_usd=800,
    )
    spec_service.select_variant(db_session, project_id=pid, query_id=q1,
                                manufacturer="Intel", quantity=2)
    spec_service.select_variant(db_session, project_id=pid, query_id=q2,
                                manufacturer="Intel", quantity=1)

    xlsx = excel_builder.build_project_xlsx(
        project_id=pid, db=db_session,
        rate=Decimal("90"), rate_date=date(2026, 4, 11),
    )
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active

    comp_names = [
        ws.cell(row=r, column=4).value
        for r in range(4, 30)
        if ws.cell(row=r, column=4).value
        and "Системный блок" in (ws.cell(row=r, column=4).value or "")
    ]
    assert len(comp_names) == 2, f"Ожидалось 2 comp-блока, нашли {comp_names}"


def test_build_xlsx_contains_gtin_and_sku(db_session):
    """Для компонентов с gtin/sku в шаблон попадают бар-код и артикул."""
    uid = _seed_user(db_session, login="xl-gtin")
    pid = spec_service.create_empty_project(db_session, user_id=uid, name="GTIN")
    cpu_id = _insert_cpu(db_session, sku="CPU-G", gtin="1234567890123")
    ram_id = _insert_ram(db_session, sku="RAM-G")
    qid = _make_query_with_cpu_ram(
        db_session, project_id=pid, user_id=uid,
        cpu_id=cpu_id, ram_id=ram_id,
    )
    spec_service.select_variant(
        db_session, project_id=pid, query_id=qid,
        manufacturer="Intel", quantity=1,
    )

    xlsx = excel_builder.build_project_xlsx(
        project_id=pid, db=db_session,
        rate=Decimal("90"), rate_date=date(2026, 4, 12),
    )
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active

    # Ищем строку CPU-компонента (не «Системный блок ...») по сочетанию
    # модели и short-description: «... · 6C/12T · ...».
    for r in range(4, 20):
        name = ws.cell(row=r, column=4).value or ""
        if "i5-12400F" in name and "6C/12T" in name:
            assert str(ws.cell(row=r, column=2).value or "") == "1234567890123"
            assert str(ws.cell(row=r, column=3).value or "") == "BX8071512400F"
            break
    else:
        pytest.fail("CPU-строка не найдена")


def test_build_xlsx_empty_spec_still_has_header(db_session):
    """Проект без спецификации — файл всё равно валидный, шапка и курс
    на месте, данных нет."""
    uid = _seed_user(db_session, login="xl-empty")
    pid = spec_service.create_empty_project(
        db_session, user_id=uid, name="Пустой",
    )
    xlsx = excel_builder.build_project_xlsx(
        project_id=pid, db=db_session,
        rate=Decimal("90"), rate_date=date(2026, 4, 13),
    )
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert "Пустой" in (ws["A1"].value or "")
    assert ws["O1"].value == "13.04.2026"
    # В строке 4 данных быть не должно.
    assert ws["D4"].value in (None, "")


# ---- тесты export_router -----------------------------------------------

def _mock_cbr(rate: str = "92.5000", d: date | None = None):
    """patch exchange_rate.get_usd_rate внутри export_router."""
    return patch(
        "app.routers.export_router.exchange_rate.get_usd_rate",
        return_value=(Decimal(rate), d or date(2026, 4, 10), "cache"),
    )


def test_excel_endpoint_returns_xlsx(
    db_session, manager_client, manager_user,
):
    cpu_id = _insert_cpu(db_session, sku="CPU-E", gtin="3333333333333")
    ram_id = _insert_ram(db_session, sku="RAM-E")
    pid = spec_service.create_empty_project(
        db_session, user_id=manager_user["id"], name="Эндпоинт",
    )
    qid = _make_query_with_cpu_ram(
        db_session, project_id=pid, user_id=manager_user["id"],
        cpu_id=cpu_id, ram_id=ram_id,
    )
    spec_service.select_variant(
        db_session, project_id=pid, query_id=qid,
        manufacturer="Intel", quantity=1,
    )

    with _mock_cbr():
        r = manager_client.get(f"/project/{pid}/export/excel")

    assert r.status_code == 200, r.text[:200]
    ct = r.headers["content-type"]
    assert "spreadsheetml" in ct
    cd = r.headers["content-disposition"]
    assert cd.startswith("attachment;")
    assert "filename*=UTF-8''" in cd
    # Бинарь валидно открывается.
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert "Эндпоинт" in (ws["A1"].value or "")


# Тест заглушки /export/kp (501) удалён: KP-эндпоинт реализован в этапе 8.2.
# Его поведение проверяется в tests/test_web/test_kp_builder.py.


def test_excel_endpoint_forbidden_for_other_user(
    db_session, app_client, manager_user, manager2_user,
):
    """Чужой проект — 403 и файл не отдаётся."""
    pid = spec_service.create_empty_project(
        db_session, user_id=manager_user["id"], name="Чужой",
    )
    # Логиним менеджера 2 — у него нет доступа к проекту менеджера 1.
    from tests.test_web.conftest import _login
    _login(app_client, manager2_user["login"], manager2_user["password"])
    with _mock_cbr():
        r = app_client.get(f"/project/{pid}/export/excel")
    assert r.status_code == 403


def test_excel_endpoint_404_for_missing_project(manager_client):
    with _mock_cbr():
        r = manager_client.get("/project/999999/export/excel")
    assert r.status_code == 404
