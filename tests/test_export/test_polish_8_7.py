# Юнит-тесты правок этапа 8.7:
# - Excel: высота строки 2 — стандартная (текст «Курс ЦБ» не обрезается).
# - KP: tblLayout=fixed, noWrap в шапке и в числовых колонках, ширины
#   tcW во всех строках совпадают с tblGrid (Word не перераспределяет
#   колонки автоматически).

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

import docx as _docx
from openpyxl import load_workbook

from app.services.export import excel_builder, kp_builder


_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


# =============================================================================
# Excel: высота строки 2
# =============================================================================

def _build_xlsx_minimal():
    """Минимальный xlsx без обращения к БД — нужен только для проверки
    высоты строки 2."""
    fake_project = {
        "id": 1, "name": "Тест 8.7",
        "created_at": datetime(2026, 4, 25, 10, 0),
        "author_login": "m", "author_name": "M",
    }
    with patch.object(excel_builder, "_load_project", return_value=fake_project), \
         patch(
             "app.services.export.excel_builder.spec_service.list_spec_items",
             return_value=[],
         ), \
         patch.object(excel_builder, "_collect_blocks", return_value=[]), \
         patch.object(excel_builder, "_fetch_gtin_map", return_value={}):
        xlsx = excel_builder.build_project_xlsx(
            project_id=1, db=None,
            rate=Decimal("92.5"), rate_date=date(2026, 4, 25),
        )
    return load_workbook(BytesIO(xlsx))


def test_excel_row2_height_normal():
    """Строка 2 содержит «Курс ЦБ» (N2) и значение курса (O2). Высота
    должна быть стандартной — None означает «по умолчанию», >=14
    означает явно заданную нормальную высоту. Сжатая высота 8pt
    обрезала бы текст «Курс ЦБ» сверху и снизу.
    """
    wb = _build_xlsx_minimal()
    ws = wb.active
    h = ws.row_dimensions[2].height
    assert h is None or h >= 14, (
        f"Высота строки 2 должна быть стандартной (None или >=14), получили {h!r}"
    )


def test_excel_row2_kurs_cells_present():
    """N2 и O2 заполнены — нельзя случайно стереть содержимое строки
    при правке высоты."""
    wb = _build_xlsx_minimal()
    ws = wb.active
    assert "Курс ЦБ" in (ws["N2"].value or "")
    assert isinstance(ws["O2"].value, (int, float)) and ws["O2"].value > 0


# =============================================================================
# KP: фикс ширин колонок и noWrap
# =============================================================================

def _fake_items(items_spec):
    out = []
    for i, (uu, qty, name) in enumerate(items_spec, start=1):
        out.append({
            "id": i, "query_id": 100 + i, "variant_manufacturer": "Intel",
            "quantity": qty, "position": i,
            "auto_name": name, "custom_name": None, "display_name": name,
            "unit_usd": uu, "unit_rub": 0.0,
            "total_usd": round(uu * qty, 2), "total_rub": 0.0,
            "created_at": None, "updated_at": None,
        })
    return out


def _mock_rate(value: str = "90"):
    return patch(
        "app.services.export.kp_builder.exchange_rate.get_usd_rate",
        return_value=(Decimal(value), date(2026, 4, 25), "cache"),
    )


def _build_kp(items_spec, markup=15):
    with patch(
        "app.services.export.kp_builder.spec_service.list_spec_items",
        return_value=_fake_items(items_spec),
    ), _mock_rate("90"):
        data = kp_builder.build_kp_docx(
            project_id=1, markup_percent=markup, db=None,
        )
    return _docx.Document(BytesIO(data))


def _kp_inner_tbl(doc):
    # Этап 9А.2.7: внешняя обёрточная таблица убрана; таблица позиций —
    # единственная в body.
    return doc.tables[0]._tbl


def test_kp_table_layout_fixed():
    """В tblPr есть <w:tblLayout w:type="fixed"/> — без этого Word
    перераспределяет ширины колонок, ломая шапку «Кол-во» в «ол-во»."""
    doc = _build_kp([(100.0, 1, "x")])
    tbl = _kp_inner_tbl(doc)
    tblPr = tbl.find(f"{_NS}tblPr")
    layout = tblPr.find(f"{_NS}tblLayout") if tblPr is not None else None
    assert layout is not None, "В tblPr нет tblLayout"
    assert layout.get(f"{_NS}type") == "fixed", (
        f"tblLayout должен быть fixed, получили {layout.get(f'{_NS}type')!r}"
    )


def test_kp_column_widths_match_grid():
    """Сумма gridCol w == tblW. Сумма tcW в каждой строке (header,
    data, total) тоже должна равняться общей grid width — иначе Word
    при отображении растянет/сожмёт ячейки самостоятельно.
    """
    doc = _build_kp([(100.0, 1, "Пример"), (200.0, 2, "Ещё")])
    tbl = _kp_inner_tbl(doc)

    # tblGrid sum
    grid = tbl.find(f"{_NS}tblGrid")
    grid_widths = [int(gc.get(f"{_NS}w")) for gc in grid.findall(f"{_NS}gridCol")]
    grid_total = sum(grid_widths)

    # tblW
    tbl_w = tbl.find(f"{_NS}tblPr").find(f"{_NS}tblW")
    assert int(tbl_w.get(f"{_NS}w")) == grid_total, (
        f"tblW {tbl_w.get(f'{_NS}w')!r} != сумме gridCol {grid_total}"
    )

    # Сумма tcW в каждой строке
    for tr_idx, tr in enumerate(tbl.findall(f"{_NS}tr")):
        row_total = 0
        for tc in tr.findall(f"{_NS}tc"):
            tcW = tc.find(f"{_NS}tcPr").find(f"{_NS}tcW")
            row_total += int(tcW.get(f"{_NS}w"))
        assert row_total == grid_total, (
            f"Строка {tr_idx}: сумма tcW={row_total} != grid_total={grid_total}"
        )


def test_kp_header_no_wrap():
    """Все ячейки шапки помечены <w:noWrap/> — заголовки «Цена с НДС
    (руб.)», «Сумма с НДС (руб.)» не должны разбиваться на 3 строки."""
    doc = _build_kp([(100.0, 1, "x")])
    tbl = _kp_inner_tbl(doc)
    header_tr = tbl.findall(f"{_NS}tr")[0]
    for i, tc in enumerate(header_tr.findall(f"{_NS}tc")):
        tcPr = tc.find(f"{_NS}tcPr")
        no_wrap = tcPr.find(f"{_NS}noWrap") if tcPr is not None else None
        assert no_wrap is not None, (
            f"В ячейке шапки {i} отсутствует <w:noWrap/>"
        )


def test_kp_numeric_cells_no_wrap():
    """Колонки «Цена» (tc[3]) и «Сумма» (tc[4]) во всех data-строках
    помечены <w:noWrap/>. Иначе число вида «42 064» переносится на
    две строки в узкой колонке.
    """
    doc = _build_kp([
        (100.0, 1, "Первый"),
        (200.0, 3, "Второй"),
    ])
    tbl = _kp_inner_tbl(doc)
    rows = tbl.findall(f"{_NS}tr")
    # Пропускаем header (rows[0]) и ИТОГО (rows[-1]) — только data.
    for tr_idx, tr in enumerate(rows[1:-1], start=1):
        tcs = tr.findall(f"{_NS}tc")
        for col_idx in (3, 4):
            tcPr = tcs[col_idx].find(f"{_NS}tcPr")
            no_wrap = tcPr.find(f"{_NS}noWrap") if tcPr is not None else None
            assert no_wrap is not None, (
                f"Data-строка {tr_idx}, колонка {col_idx}: нет <w:noWrap/>"
            )


def test_kp_total_row_cells_no_wrap():
    """Обе ячейки строки ИТОГО (label + value) — noWrap."""
    doc = _build_kp([(100.0, 1, "x")])
    tbl = _kp_inner_tbl(doc)
    total_tr = tbl.findall(f"{_NS}tr")[-1]
    for i, tc in enumerate(total_tr.findall(f"{_NS}tc")):
        tcPr = tc.find(f"{_NS}tcPr")
        no_wrap = tcPr.find(f"{_NS}noWrap") if tcPr is not None else None
        assert no_wrap is not None, (
            f"Ячейка ИТОГО {i}: ожидался <w:noWrap/>"
        )


def test_kp_total_row_has_min_height():
    """trHeight у строки ИТОГО задан (≥0.7см ≈ 397 twips), чтобы строка
    выделялась визуально."""
    doc = _build_kp([(100.0, 1, "x")])
    tbl = _kp_inner_tbl(doc)
    total_tr = tbl.findall(f"{_NS}tr")[-1]
    trPr = total_tr.find(f"{_NS}trPr")
    assert trPr is not None, "У строки ИТОГО нет trPr"
    trHeight = trPr.find(f"{_NS}trHeight")
    assert trHeight is not None, "У строки ИТОГО нет trHeight"
    assert int(trHeight.get(f"{_NS}val")) >= 397, (
        f"Высота строки ИТОГО должна быть ≥397 twips, "
        f"получили {trHeight.get(f'{_NS}val')!r}"
    )


def test_kp_grid_total_width_fits_a4_page():
    """Сумма ширин колонок не превышает 16см (9072 twips) — иначе
    таблица не помещается в полезную ширину A4 при стандартных полях
    25мм."""
    doc = _build_kp([(100.0, 1, "x")])
    tbl = _kp_inner_tbl(doc)
    grid = tbl.find(f"{_NS}tblGrid")
    total = sum(int(gc.get(f"{_NS}w")) for gc in grid.findall(f"{_NS}gridCol"))
    # 16см = 9072 twips. Допускаем небольшой запас погрешности до 9100.
    assert total <= 9100, (
        f"Общая ширина таблицы {total} twips превышает 16см (9072)"
    )


def test_kp_total_label_jc_right():
    """Текст «ИТОГО» в строке итога выровнен по правому краю — иначе
    он зрительно «уплывает» от значения справа."""
    doc = _build_kp([(100.0, 1, "x")])
    tbl = _kp_inner_tbl(doc)
    total_tr = tbl.findall(f"{_NS}tr")[-1]
    label_tc = total_tr.findall(f"{_NS}tc")[0]
    p = label_tc.find(f"{_NS}p")
    pPr = p.find(f"{_NS}pPr") if p is not None else None
    jc = pPr.find(f"{_NS}jc") if pPr is not None else None
    assert jc is not None and jc.get(f"{_NS}val") == "right", (
        f"«ИТОГО» должен быть jc=right, получили {jc.get(f'{_NS}val') if jc is not None else None!r}"
    )
