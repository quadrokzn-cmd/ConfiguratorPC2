"""Юнит-тесты Excel-выгрузки списка аукционов (Backlog #12, Фаза 4 плана
2026-05-13-auctions-excel-export.md).

Покрытие:
  - Пустая БД → файл с шапкой и autofilter, 0 data-строк.
  - 1 tender + 2 tender_items + primary match для одной → 2 строки;
    match-колонки пустые у строки без primary.
  - Гранулярность B: каждой tender_item — отдельная строка.
  - Hyperlink на tenders.url активен на строке с лотом.
  - Маржа % — формула в Excel с числовым форматом 0.00%.
  - Маржа RUB — формула в Excel с числовым форматом 0.00.
  - Сериализация ktru_codes_array (TEXT[]) → через запятую.
  - Сериализация flags_jsonb → ключи через запятую.
  - Фильтр status=['new'] → только лоты со статусом new.
  - Фильтр nmck_min → лоты с НМЦК ≥ порога.
  - Фильтр q (search) → ILIKE по reg_number / customer / region.
  - include_excluded_regions=False (default) → лоты с
    flags.excluded_by_region=true НЕ попадают.
  - Fallback курса 90.0 при пустой exchange_rates → rate_is_fallback=True.
  - Cap reached → флаг cap_reached=True (через monkeypatch _ROW_CAP).
  - default_filename → «Аукционы_YYYY-MM-DD.xlsx».
"""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import sessionmaker

from portal.services.auctions import excel_export
from portal.services.auctions.excel_export import (
    ExportReport,
    _COLUMNS,
    _HEADER_ROW,
    _DATA_START_ROW,
    _col_index,
    _col_letter,
    default_filename,
    export_auctions,
)
from portal.services.auctions_service import InboxFilters


MSK = timezone(timedelta(hours=3))


# ---------------------------------------------------------------
# Фикстуры
# ---------------------------------------------------------------

@pytest.fixture(autouse=True)
def _clean_auction_tables(db_engine):
    """Перед каждым тестом — чистые аукционные + smart-ingest-зависимые
    таблицы. exchange_rates тоже чистим, чтобы тесты fallback-курса не
    зависели от других тестов в сессии."""
    with db_engine.begin() as conn:
        conn.execute(text(
            "TRUNCATE TABLE matches, tender_items, tender_status, tenders, "
            "printers_mfu, supplier_prices, suppliers, exchange_rates "
            "RESTART IDENTITY CASCADE"
        ))
    yield


@pytest.fixture()
def session(db_engine):
    """SQLAlchemy-сессия для прямой передачи в export_auctions(db=...)."""
    Session = sessionmaker(bind=db_engine, autoflush=False, future=True)
    s = Session()
    try:
        yield s
    finally:
        s.close()


# ---------------------------------------------------------------
# Хелперы — вставка данных в БД
# ---------------------------------------------------------------

def _insert_tender(
    db_engine,
    reg_number: str = "0001",
    *,
    customer: str = "ФГБУ Тестовый Заказчик",
    customer_region: str = "Республика Татарстан",
    nmck_total: Decimal = Decimal("125000.00"),
    publish_date: datetime | None = None,
    submit_deadline: datetime | None = None,
    delivery_deadline: datetime | None = None,
    ktru_codes: list[str] | None = None,
    flags: dict | None = None,
    url: str | None = None,
) -> None:
    if publish_date is None:
        publish_date = datetime(2026, 4, 10, tzinfo=MSK)
    if submit_deadline is None:
        submit_deadline = datetime(2026, 4, 25, 9, 0, tzinfo=MSK)
    if delivery_deadline is None:
        delivery_deadline = datetime(2026, 7, 15, tzinfo=MSK)
    if ktru_codes is None:
        ktru_codes = ["26.20.18.000-00000069"]
    if flags is None:
        flags = {}
    if url is None:
        url = f"https://zakupki.gov.ru/{reg_number}"
    with db_engine.begin() as conn:
        conn.execute(text(
            """
            INSERT INTO tenders (reg_number, customer, customer_region,
                customer_contacts_jsonb, nmck_total, publish_date,
                submit_deadline, delivery_deadline, ktru_codes_array,
                url, raw_html, flags_jsonb)
            VALUES (:rn, :c, :r, '{}'::jsonb, :nmck, :pub, :sub, :dlv,
                CAST(:ktru AS TEXT[]), :url, '', CAST(:flags AS JSONB))
            """
        ), {
            "rn":    reg_number,
            "c":     customer,
            "r":     customer_region,
            "nmck":  nmck_total,
            "pub":   publish_date,
            "sub":   submit_deadline,
            "dlv":   delivery_deadline,
            "ktru":  "{" + ",".join(ktru_codes) + "}",
            "url":   url,
            "flags": json.dumps(flags, ensure_ascii=False),
        })


def _insert_tender_status(
    db_engine, reg_number: str, *, status: str = "new",
) -> None:
    with db_engine.begin() as conn:
        conn.execute(text(
            "INSERT INTO tender_status (tender_id, status) "
            "VALUES (:rn, :s)"
        ), {"rn": reg_number, "s": status})


def _insert_item(
    db_engine,
    *,
    tender_id: str,
    position_num: int = 1,
    name: str = "МФУ ч/б A4",
    qty: Decimal = Decimal("5"),
    nmck_per_unit: Decimal | None = Decimal("25000.00"),
    ktru_code: str = "26.20.18.000-00000069",
) -> int:
    with db_engine.begin() as conn:
        row = conn.execute(text(
            """
            INSERT INTO tender_items (tender_id, position_num, ktru_code,
                name, qty, unit, required_attrs_jsonb, nmck_per_unit)
            VALUES (:tid, :pos, :ktru, :n, :qty, 'шт', '{}'::jsonb, :ppu)
            RETURNING id
            """
        ), {
            "tid":  tender_id,
            "pos":  position_num,
            "ktru": ktru_code,
            "n":    name,
            "qty":  qty,
            "ppu":  nmck_per_unit,
        }).first()
        return int(row.id)


def _insert_sku(
    db_engine,
    *,
    sku: str = "pantum-bp1800",
    brand: str = "Pantum",
    name: str = "Pantum BP1800",
    category: str = "mfu",
    cost_base_rub: Decimal = Decimal("12000.00"),
) -> int:
    with db_engine.begin() as conn:
        row = conn.execute(text(
            """
            INSERT INTO printers_mfu (sku, brand, name, category, ktru_codes_array,
                attrs_jsonb, cost_base_rub)
            VALUES (:sku, :br, :n, :cat, CAST(:ktru AS TEXT[]),
                '{}'::jsonb, :cost)
            RETURNING id
            """
        ), {
            "sku":  sku,
            "br":   brand,
            "n":    name,
            "cat":  category,
            "ktru": "{26.20.18.000-00000069}",
            "cost": cost_base_rub,
        }).first()
        return int(row.id)


def _insert_match(
    db_engine, *, tender_item_id: int, sku_id: int,
    margin_rub: Decimal = Decimal("65000.00"),
    margin_pct: Decimal = Decimal("52.00"),
    price_total_rub: Decimal = Decimal("125000.00"),
    match_type: str = "primary",
) -> None:
    with db_engine.begin() as conn:
        conn.execute(text(
            """
            INSERT INTO matches (tender_item_id, nomenclature_id, match_type,
                price_total_rub, margin_rub, margin_pct)
            VALUES (:item, :sku, :mt, :ptr, :mr, :mp)
            """
        ), {
            "item": tender_item_id,
            "sku":  sku_id,
            "mt":   match_type,
            "ptr":  price_total_rub,
            "mr":   margin_rub,
            "mp":   margin_pct,
        })


def _filters(**kwargs) -> InboxFilters:
    return InboxFilters(**kwargs)


# ---------------------------------------------------------------
# 1. Базовые сценарии — структура файла
# ---------------------------------------------------------------

def test_empty_db_writes_header_only(tmp_path, session):
    out = tmp_path / "empty.xlsx"
    report = export_auctions(out, _filters(), db=session)

    assert out.exists()
    assert report.rows_count == 0
    assert not report.cap_reached

    wb = load_workbook(out)
    ws = wb.active
    assert ws.title == "Аукционы"
    # Шапка на строке 3
    assert ws.cell(row=_HEADER_ROW, column=1).value == _COLUMNS[0].title
    # Данных нет
    assert ws.cell(row=_DATA_START_ROW, column=2).value is None
    # autofilter
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith(f"A{_HEADER_ROW}")


def test_one_tender_two_items_one_with_match(tmp_path, db_engine, session):
    """1 tender + 2 items, primary match только у одной → 2 строки в Excel.

    Match-колонки у второй строки пустые. Гранулярность B подтверждена:
    одна строка per tender_item.
    """
    _insert_tender(db_engine, "0010")
    _insert_tender_status(db_engine, "0010", status="new")
    item1_id = _insert_item(
        db_engine, tender_id="0010", position_num=1,
        name="МФУ ч/б A4", qty=Decimal("5"),
        nmck_per_unit=Decimal("25000.00"),
    )
    _insert_item(
        db_engine, tender_id="0010", position_num=2,
        name="Принтер цветной A4", qty=Decimal("2"),
        nmck_per_unit=Decimal("18000.00"),
    )
    sku_id = _insert_sku(db_engine)
    _insert_match(
        db_engine, tender_item_id=item1_id, sku_id=sku_id,
    )

    out = tmp_path / "two-items.xlsx"
    report = export_auctions(out, _filters(), db=session)
    assert report.rows_count == 2

    wb = load_workbook(out)
    ws = wb.active

    # Две строки данных (position_num 1 и 2)
    pos_col = _col_index("№ позиции")
    pos_values = [ws.cell(row=_DATA_START_ROW + i, column=pos_col).value for i in range(2)]
    assert sorted([v for v in pos_values if v is not None]) == [1, 2]

    # Строка 1 (item1) — есть бренд (есть match), строка 2 — пусто.
    brand_col = _col_index("Бренд SKU")
    cost_col = _col_index("Cost base, ₽")
    for row_offset in range(2):
        row_idx = _DATA_START_ROW + row_offset
        pos = ws.cell(row=row_idx, column=pos_col).value
        brand = ws.cell(row=row_idx, column=brand_col).value
        cost = ws.cell(row=row_idx, column=cost_col).value
        if pos == 1:
            assert brand == "Pantum"
            assert cost is not None
        else:
            assert brand in (None, "")
            assert cost is None


def test_hyperlink_on_url(tmp_path, db_engine, session):
    """Колонка «Карточка zakupki» хранит hyperlink на tenders.url."""
    _insert_tender(db_engine, "0011", url="https://zakupki.gov.ru/0011-test")
    _insert_tender_status(db_engine, "0011")
    _insert_item(db_engine, tender_id="0011")

    out = tmp_path / "hyperlink.xlsx"
    export_auctions(out, _filters(), db=session)

    wb = load_workbook(out)
    ws = wb.active
    url_cell = ws.cell(row=_DATA_START_ROW, column=_col_index("Карточка zakupki"))
    assert url_cell.value == "Открыть"
    assert url_cell.hyperlink is not None
    assert url_cell.hyperlink.target == "https://zakupki.gov.ru/0011-test"


# ---------------------------------------------------------------
# 2. Формулы маржи
# ---------------------------------------------------------------

def test_margin_pct_is_formula_when_match_present(tmp_path, db_engine, session):
    """Маржа % — формула =(Price - Cost)/Cost, формат 0.00%."""
    _insert_tender(db_engine, "0020")
    _insert_tender_status(db_engine, "0020")
    item_id = _insert_item(
        db_engine, tender_id="0020",
        qty=Decimal("3"), nmck_per_unit=Decimal("40000"),
    )
    sku_id = _insert_sku(db_engine, cost_base_rub=Decimal("10000"))
    _insert_match(db_engine, tender_item_id=item_id, sku_id=sku_id)

    out = tmp_path / "margin.xlsx"
    export_auctions(out, _filters(), db=session)

    wb = load_workbook(out)
    ws = wb.active
    row = _DATA_START_ROW
    margin_pct_cell = ws.cell(row=row, column=_col_index("Маржа, %"))
    margin_rub_cell = ws.cell(row=row, column=_col_index("Маржа, ₽"))

    # Формула, не статика
    assert isinstance(margin_pct_cell.value, str) and margin_pct_cell.value.startswith("=")
    assert isinstance(margin_rub_cell.value, str) and margin_rub_cell.value.startswith("=")

    # Формула опирается на ячейки Cost / Price в этой же строке.
    price_letter = _col_letter("Цена за единицу, ₽")
    cost_letter = _col_letter("Cost base, ₽")
    qty_letter = _col_letter("Количество")
    assert f"{price_letter}{row}" in margin_pct_cell.value
    assert f"{cost_letter}{row}" in margin_pct_cell.value
    assert f"/{cost_letter}{row}" in margin_pct_cell.value
    assert f"{qty_letter}{row}" in margin_rub_cell.value

    # Формат
    assert margin_pct_cell.number_format == "0.00%"
    assert margin_rub_cell.number_format == "0.00"


def test_margin_empty_when_no_cost(tmp_path, db_engine, session):
    """Если у tender_item нет primary match (cost) — маржа-ячейки пусты."""
    _insert_tender(db_engine, "0021")
    _insert_tender_status(db_engine, "0021")
    _insert_item(db_engine, tender_id="0021")

    out = tmp_path / "no-margin.xlsx"
    export_auctions(out, _filters(), db=session)

    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=_DATA_START_ROW, column=_col_index("Маржа, %")).value is None
    assert ws.cell(row=_DATA_START_ROW, column=_col_index("Маржа, ₽")).value is None


# ---------------------------------------------------------------
# 3. Сериализация массивов и флагов
# ---------------------------------------------------------------

def test_ktru_array_serialized_comma_separated(tmp_path, db_engine, session):
    _insert_tender(
        db_engine, "0030",
        ktru_codes=["26.20.18.000-00000069", "26.20.16.120-00000013"],
    )
    _insert_tender_status(db_engine, "0030")
    _insert_item(db_engine, tender_id="0030")

    out = tmp_path / "ktru.xlsx"
    export_auctions(out, _filters(), db=session)
    wb = load_workbook(out)
    ws = wb.active
    ktru_cell = ws.cell(row=_DATA_START_ROW, column=_col_index("KTRU лота"))
    assert ktru_cell.value == "26.20.18.000-00000069,26.20.16.120-00000013"


def test_flags_serialized_as_truthy_keys(tmp_path, db_engine, session):
    _insert_tender(
        db_engine, "0031",
        flags={"below_nmck_min": True, "excluded_by_region": False, "no_positions_parsed": True},
    )
    _insert_tender_status(db_engine, "0031")
    _insert_item(db_engine, tender_id="0031")

    out = tmp_path / "flags.xlsx"
    export_auctions(out, _filters(), db=session)
    wb = load_workbook(out)
    ws = wb.active
    flags_cell = ws.cell(row=_DATA_START_ROW, column=_col_index("Флаги"))
    parts = (flags_cell.value or "").split(",")
    # Только truthy ключи — порядок не важен
    assert set(parts) == {"below_nmck_min", "no_positions_parsed"}


# ---------------------------------------------------------------
# 4. Фильтры
# ---------------------------------------------------------------

def test_filter_status(tmp_path, db_engine, session):
    """Фильтр statuses=['new'] → только tender'ы со статусом new."""
    _insert_tender(db_engine, "0040")
    _insert_tender_status(db_engine, "0040", status="new")
    _insert_item(db_engine, tender_id="0040")

    _insert_tender(db_engine, "0041")
    _insert_tender_status(db_engine, "0041", status="skipped")
    _insert_item(db_engine, tender_id="0041")

    out = tmp_path / "status-filter.xlsx"
    report = export_auctions(out, _filters(statuses=("new",)), db=session)
    assert report.rows_count == 1

    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=_DATA_START_ROW, column=_col_index("№ извещения")).value == "0040"


def test_filter_nmck_min(tmp_path, db_engine, session):
    _insert_tender(db_engine, "0050", nmck_total=Decimal("10000"))
    _insert_tender_status(db_engine, "0050")
    _insert_item(db_engine, tender_id="0050")

    _insert_tender(db_engine, "0051", nmck_total=Decimal("100000"))
    _insert_tender_status(db_engine, "0051")
    _insert_item(db_engine, tender_id="0051")

    out = tmp_path / "nmck-min.xlsx"
    report = export_auctions(
        out, _filters(nmck_min=Decimal("50000")), db=session,
    )
    assert report.rows_count == 1
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=_DATA_START_ROW, column=_col_index("№ извещения")).value == "0051"


def test_filter_search_ilike(tmp_path, db_engine, session):
    """ILIKE-фильтр по reg_number / customer / customer_region.

    Тестовая БД создана с LC_COLLATE='C' / LC_CTYPE='C' (см.
    `_ensure_worker_database_exists` в tests/conftest.py), поэтому ILIKE
    регистр-независим только для ASCII — для кириллицы нужен точный
    регистр. На prod locale системная, ILIKE «якут» совпадёт с «Якутия».
    """
    _insert_tender(db_engine, "0060", customer_region="Якутия")
    _insert_tender_status(db_engine, "0060")
    _insert_item(db_engine, tender_id="0060")

    _insert_tender(db_engine, "0061", customer_region="Республика Татарстан")
    _insert_tender_status(db_engine, "0061")
    _insert_item(db_engine, tender_id="0061")

    out = tmp_path / "search.xlsx"
    report = export_auctions(out, _filters(search="Якут"), db=session)
    assert report.rows_count == 1
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=_DATA_START_ROW, column=_col_index("№ извещения")).value == "0060"


def test_filter_excluded_by_region_default_hides(tmp_path, db_engine, session):
    """include_excluded_regions=False (default) → лоты с
    flags.excluded_by_region=true скрыты."""
    _insert_tender(
        db_engine, "0070", flags={"excluded_by_region": True},
    )
    _insert_tender_status(db_engine, "0070")
    _insert_item(db_engine, tender_id="0070")

    _insert_tender(db_engine, "0071", flags={})
    _insert_tender_status(db_engine, "0071")
    _insert_item(db_engine, tender_id="0071")

    out = tmp_path / "excluded.xlsx"
    report = export_auctions(out, _filters(), db=session)
    assert report.rows_count == 1
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=_DATA_START_ROW, column=_col_index("№ извещения")).value == "0071"


def test_filter_include_excluded_regions_shows_all(tmp_path, db_engine, session):
    _insert_tender(db_engine, "0080", flags={"excluded_by_region": True})
    _insert_tender_status(db_engine, "0080")
    _insert_item(db_engine, tender_id="0080")

    out = tmp_path / "excluded-included.xlsx"
    report = export_auctions(
        out, _filters(include_excluded_regions=True), db=session,
    )
    assert report.rows_count == 1


# ---------------------------------------------------------------
# 5. Курс и cap
# ---------------------------------------------------------------

def test_rate_fallback_when_exchange_rates_empty(tmp_path, db_engine, session):
    """Пустая exchange_rates → fallback 90.0 + flag rate_is_fallback=True."""
    _insert_tender(db_engine, "0090")
    _insert_tender_status(db_engine, "0090")
    _insert_item(db_engine, tender_id="0090")

    out = tmp_path / "fallback.xlsx"
    report = export_auctions(out, _filters(), db=session)
    assert report.rate_is_fallback is True
    assert report.rate_used == Decimal("90.0000")

    wb = load_workbook(out)
    ws = wb.active
    assert float(ws["B1"].value) == pytest.approx(90.0)


def test_rate_used_from_exchange_rates(tmp_path, db_engine, session):
    """Если в exchange_rates есть курс — используем его, не fallback."""
    with db_engine.begin() as conn:
        conn.execute(text(
            "INSERT INTO exchange_rates (rate_usd_rub, rate_date) "
            "VALUES (97.4523, CURRENT_DATE)"
        ))
    _insert_tender(db_engine, "0091")
    _insert_tender_status(db_engine, "0091")
    _insert_item(db_engine, tender_id="0091")

    out = tmp_path / "from-table.xlsx"
    report = export_auctions(out, _filters(), db=session)
    assert report.rate_is_fallback is False
    assert float(report.rate_used) == pytest.approx(97.4523, abs=1e-4)


def test_cap_reached_flag(tmp_path, db_engine, session, monkeypatch):
    """При cap=2 и 3 tender_items в БД → cap_reached=True, 2 строки."""
    monkeypatch.setattr(excel_export, "_ROW_CAP", 2)

    _insert_tender(db_engine, "0100")
    _insert_tender_status(db_engine, "0100")
    _insert_item(db_engine, tender_id="0100", position_num=1)
    _insert_item(db_engine, tender_id="0100", position_num=2)
    _insert_item(db_engine, tender_id="0100", position_num=3)

    out = tmp_path / "cap.xlsx"
    report = export_auctions(out, _filters(), db=session)
    assert report.cap_reached is True
    assert report.rows_count == 2


# ---------------------------------------------------------------
# 6. default_filename
# ---------------------------------------------------------------

def test_default_filename_uses_iso_date():
    name = default_filename(today=date(2026, 5, 17))
    assert name == "Аукционы_2026-05-17.xlsx"


def test_default_filename_today_when_none():
    name = default_filename()
    today = date.today().strftime("%Y-%m-%d")
    assert name == f"Аукционы_{today}.xlsx"


# ---------------------------------------------------------------
# 7. Filter summary в ExportReport
# ---------------------------------------------------------------

def test_filter_summary_in_report(tmp_path, db_engine, session):
    out = tmp_path / "summary.xlsx"
    report = export_auctions(
        out,
        _filters(
            statuses=("new", "in_review"),
            nmck_min=Decimal("10000"),
            search="квадро",
            urgent_only=True,
        ),
        db=session,
    )
    summary = report.filter_summary
    assert summary["statuses"] == ["new", "in_review"]
    assert summary["nmck_min"] == "10000"
    assert summary["search"] == "квадро"
    assert summary["urgent_only"] is True
    assert summary["print_only"] is False
    assert summary["include_excluded_regions"] is False
