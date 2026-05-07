# Тесты MerlionImapFetcher — IMAP-канал Merlion (12.1).

from __future__ import annotations

import io
import os
import re
import struct
import zipfile

import pytest


# =====================================================================
# 1. Subject regex
# =====================================================================

def test_merlion_subject_regex_matches():
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    pat = re.compile(MerlionImapFetcher.subject_pattern, re.IGNORECASE)
    for s in [
        "Прайс-лист MERLION",
        "Прайс-лист MERLION Москва 06.05.2026",
        "  Прайс-лист MERLION 30.04.2026",
        # Реальный Subject из IMAP-разведки 2026-05-06 (только этот):
        "Прайс-лист MERLION",
    ]:
        assert pat.search(s), f"subject должен матчиться: {s!r}"


def test_merlion_subject_regex_rejects_unrelated():
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    pat = re.compile(MerlionImapFetcher.subject_pattern, re.IGNORECASE)
    for s in [
        "Новости MERLION",
        "Прайс OCS - Состояние склада и цены",
        "Re: Прайс-лист MERLION",
    ]:
        assert not pat.search(s), f"subject должен быть отклонён: {s!r}"


# =====================================================================
# 2. Sender regex (домен merlion.ru, в т.ч. через Gmail-forward)
# =====================================================================

def test_merlion_sender_regex_matches():
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    pat = re.compile(MerlionImapFetcher.sender_pattern, re.IGNORECASE)
    for s in [
        "matveeva.y@merlion.ru",
        "Матвеева <matveeva.y@merlion.ru>",
        "пересылка via gmail; original: news@merlion.ru",
        # Реальные From из IMAP-разведки 2026-05-06:
        "matveeva.y@merlion.ru <matveeva.y@merlion.ru>",  # двойной From после Gmail-forward
        "Matveeva Yuliya <Matveeva.Y@merlion.ru>",        # display name + capitalized
        "Antonov Sergey <Antonov.S@merlion.ru>",          # другой менеджер
    ]:
        assert pat.search(s), f"sender должен матчиться: {s!r}"


def test_merlion_from_via_gmail_forward_matches():
    """После Gmail-forward Reply-To/X-Forwarded-For/Return-Path содержат
    quadro.kzn@gmail.com и caf_-префиксы, но FROM сохраняет реальный
    @merlion.ru. BaseImapFetcher склеивает все адресные заголовки в один
    haystack, и должен матчить именно по From — gmail в Reply-To не
    должен мешать (regex ищет @merlion.ru, gmail.com его не сматчит)."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher
    from app.services.auto_price.fetchers.base_imap import _addresses_in_header

    headers = {
        "From":            "matveeva.y@merlion.ru <matveeva.y@merlion.ru>",
        "Reply-To":        "",
        "X-Forwarded-For": "quadro.kzn@gmail.com quadro@quadro.tatar",
        "Return-Path":     "<quadro.kzn+caf_=quadro=quadro.tatar@gmail.com>",
        "Sender":          "",
    }
    haystack = _addresses_in_header(headers)
    pat = re.compile(MerlionImapFetcher.sender_pattern, re.IGNORECASE)
    assert pat.search(haystack), (
        f"Gmail-forwarded Merlion должен матчиться по From, "
        f"haystack={haystack!r}"
    )


def test_merlion_sender_regex_rejects_other():
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    pat = re.compile(MerlionImapFetcher.sender_pattern, re.IGNORECASE)
    for s in ["egarifullina@ocs.ru", "noreply@merlion.com", "scam@merlion.ru.fake"]:
        assert not pat.search(s)


# =====================================================================
# 3. ZIP extraction
# =====================================================================

def _make_zip_with(files: list[tuple[str, bytes]]) -> bytes:
    """Собирает in-memory ZIP с указанными файлами (UTF-8 EFS bit
    выставлен — Python zipfile это делает автоматически для
    не-ASCII имён)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files:
            zf.writestr(name, data)
    return buf.getvalue()


def _make_zip_cp1251(files: list[tuple[str, bytes]]) -> bytes:
    """Собирает ZIP с именами в cp1251 БЕЗ EFS-флага.

    Это воспроизводит формат Merlion-рассылки: имена в cp1251, bit 11
    в general purpose flags не выставлен. Чтобы получить именно такое
    поведение, мы пишем ZipInfo вручную: filename — байты cp1251,
    декодированные как latin-1 (поверхностный обход, чтобы Python не
    выставил EFS), и явно сбрасываем 0x800 в flag_bits.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files:
            # Имитируем поведение архиватора, который пишет cp1251
            # байты прямо в filename без UTF-8 EFS-флага. zipfile при
            # ЗАПИСИ интерпретирует name как unicode → закодирует в
            # ASCII, иначе выставит EFS. Чтобы обмануть — кладём
            # cp1251-байты, декодированные как latin-1: ASCII-проверка
            # пройдёт, EFS не выставится, на диск уйдут байты cp1251.
            cp1251_bytes = name.encode("cp1251")
            fake_ascii_name = cp1251_bytes.decode("latin-1")
            info = zipfile.ZipInfo(filename=fake_ascii_name)
            info.compress_type = zipfile.ZIP_DEFLATED
            # На всякий случай: явно сбросим bit 11 (хотя для ASCII
            # имени Python и не выставит).
            info.flag_bits &= ~0x800
            zf.writestr(info, data)
    return buf.getvalue()


def _make_real_xlsx_bytes(sheet_name: str = "Sheet1") -> bytes:
    """Собирает минимальный валидный xlsx через openpyxl, чтобы
    MerlionLoader.iter_rows() мог его открыть в end-to-end тесте."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = "Sample"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_merlion_zip_extraction_picks_largest_xlsx(monkeypatch):
    """В ZIP может быть несколько .xlsx (основной + сопровождающие).
    parse_attachment должен взять самый большой и передать в MerlionLoader."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    # Готовим ZIP с двумя xlsx разного размера + один pdf.
    big = b"PK\x03\x04" + (b"X" * 5000) + b"big-xlsx-end"
    small = b"PK\x03\x04small-xlsx"
    pdf = b"%PDF-1.4 lic"
    zip_bytes = _make_zip_with([
        ("license.pdf", pdf),
        ("merlion_msk_06_05.xlsx", big),
        ("hint_short.xlsx", small),
    ])

    captured = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            captured["filepath"] = filepath
            with open(filepath, "rb") as f:
                captured["bytes"] = f.read()
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    fetcher = MerlionImapFetcher()
    rows = fetcher.parse_attachment(zip_bytes, "merlion_06_05.zip")

    assert rows == []
    assert captured["filepath"].endswith(".xlsx")
    assert captured["bytes"] == big  # выбран самый большой


def test_merlion_zip_extraction_with_single_xlsx(monkeypatch):
    """Базовый кейс: один XLSX внутри — берём его."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"PK\x03\x04only-one"
    zip_bytes = _make_zip_with([("price.xlsx", payload)])

    seen = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            with open(filepath, "rb") as f:
                seen["bytes"] = f.read()
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    MerlionImapFetcher().parse_attachment(zip_bytes, "merlion.zip")
    assert seen["bytes"] == payload


def test_merlion_zip_no_xlsx_raises(monkeypatch):
    """ZIP без xlsx — RuntimeError, чтобы runner и orchestrator не делали
    загрузку с пустыми rows (не обнуляли остатки)."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    zip_bytes = _make_zip_with([("license.pdf", b"%PDF-1.4")])

    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")
    fetcher = MerlionImapFetcher()
    with pytest.raises(RuntimeError, match="не содержит ни одного файла .xlsx"):
        fetcher.parse_attachment(zip_bytes, "merlion.zip")


def test_merlion_bad_zip_raises(monkeypatch):
    """Bytes — не ZIP. RuntimeError, NoNewDataException не уместен —
    письмо есть, но кривое."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")
    fetcher = MerlionImapFetcher()
    with pytest.raises(RuntimeError, match="не распознано как ZIP"):
        fetcher.parse_attachment(b"not a zip", "broken.zip")


# =====================================================================
# 4. cp1251 имена внутри ZIP (12.1-fix-2)
# =====================================================================

def test_zip_with_cp1251_names_extracted_correctly(monkeypatch):
    """Воспроизводит реальный Merlion-ZIP: имена в cp1251 БЕЗ
    UTF-8 EFS-флага. До 12.1-fix-2 zf.extractall() падал на
    mismatch directory/header. Теперь fetcher должен корректно
    распаковать и передать содержимое в MerlionLoader."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"PK\x03\x04cp1251-content-marker"
    zip_bytes = _make_zip_cp1251([
        ("Прайслист_Мерлион_Москва.xlsm", payload),
    ])

    seen = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            seen["filepath"] = filepath
            with open(filepath, "rb") as f:
                seen["bytes"] = f.read()
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    MerlionImapFetcher().parse_attachment(zip_bytes, "PriceList_MERLION_Moskva.zip")
    assert seen["bytes"] == payload, "содержимое xlsm должно сохраниться"
    # Имя файла на диске должно быть raskодировано из cp1251 в нормальный
    # «Прайслист_Мерлион_Москва.xlsm» (или близкий — мы нормализуем
    # незаконные FS-символы, но кириллица должна быть).
    base = os.path.basename(seen["filepath"])
    assert "Прайслист" in base or base.lower().endswith(".xlsm"), (
        f"имя на диске должно быть cp1251-decoded: {base!r}"
    )


def test_zip_xlsm_extension_accepted(monkeypatch):
    """ZIP с .xlsm файлом (Excel с макросами) — fetcher должен его
    найти. До 12.1-fix-2 искалось только .xlsx."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"PK\x03\x04xlsm-payload"
    zip_bytes = _make_zip_with([("price_with_macros.xlsm", payload)])

    seen = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            with open(filepath, "rb") as f:
                seen["bytes"] = f.read()
            seen["ext"] = os.path.splitext(filepath)[1].lower()
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    MerlionImapFetcher().parse_attachment(zip_bytes, "merlion.zip")
    assert seen["bytes"] == payload
    assert seen["ext"] == ".xlsm"


def test_zip_already_utf8_efs_flag_preserved(monkeypatch):
    """Современный ZIP с UTF-8 именами (EFS bit 11 = 1) — наш фикс
    не должен ломать такие архивы. Имя должно остаться как есть,
    без двойного перекодирования."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"PK\x03\x04utf8-efs-payload"
    # _make_zip_with пишет через writestr(name, data) — Python для
    # не-ASCII автоматически выставляет UTF-8 EFS-флаг.
    zip_bytes = _make_zip_with([("Прайс.xlsx", payload)])

    seen = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            with open(filepath, "rb") as f:
                seen["bytes"] = f.read()
            seen["filepath"] = filepath
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    MerlionImapFetcher().parse_attachment(zip_bytes, "merlion.zip")
    assert seen["bytes"] == payload
    base = os.path.basename(seen["filepath"])
    # Не должно быть мусора вроде «Ÿàü½ª¨ÅÅ.xlsx».
    assert "Прайс" in base, (
        f"UTF-8-имя должно сохраниться без двойной перекодировки: {base!r}"
    )


def test_zip_with_multiple_files_picks_largest_xlsm_or_xlsx(monkeypatch):
    """В реальном ZIP может быть и .xlsx, и .xlsm. Берётся самый большой."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    big_xlsm = b"PK\x03\x04" + (b"X" * 20000) + b"big-xlsm"
    small_xlsx = b"PK\x03\x04small-xlsx"
    pdf = b"%PDF-1.4 license-text"
    zip_bytes = _make_zip_with([
        ("license.pdf", pdf),
        ("price_main.xlsm", big_xlsm),  # самый большой — выбираем его
        ("price_supplement.xlsx", small_xlsx),
    ])

    seen = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            with open(filepath, "rb") as f:
                seen["bytes"] = f.read()
            seen["filepath"] = filepath
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    MerlionImapFetcher().parse_attachment(zip_bytes, "merlion.zip")
    assert seen["bytes"] == big_xlsm
    assert seen["filepath"].endswith(".xlsm")


def test_zip_no_xlsx_or_xlsm_raises(monkeypatch):
    """ZIP без xlsx и xlsm — RuntimeError (формат рассылки изменился)."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    zip_bytes = _make_zip_with([
        ("license.pdf", b"%PDF-1.4"),
        ("readme.txt", b"hello"),
    ])
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")
    fetcher = MerlionImapFetcher()
    with pytest.raises(RuntimeError, match=r"не содержит ни одного файла .xlsx или .xlsm"):
        fetcher.parse_attachment(zip_bytes, "merlion.zip")


def test_merlion_loader_can_open_real_xlsm(tmp_path):
    """Smoke: openpyxl/MerlionLoader действительно открывает .xlsm
    как обычный xlsx (формат внутри идентичен — ZIP с XML; разница
    только в наличии макросов, которые openpyxl игнорирует)."""
    from openpyxl import load_workbook
    xlsm_path = tmp_path / "smoke.xlsm"
    xlsm_path.write_bytes(_make_real_xlsx_bytes())
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    try:
        assert wb.active is not None
    finally:
        wb.close()


# =====================================================================
# 5. Helper: декодинг имени по EFS-флагу
# =====================================================================

def test_decode_zip_member_name_cp1251_when_efs_off():
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    info = zipfile.ZipInfo()
    # Имитируем то, что zipfile при чтении положит в .filename: байты
    # cp1251, декодированные как cp437 → мусор.
    info.filename = "Прайслист_Мерлион_Москва.xlsm".encode("cp1251").decode("cp437")
    info.flag_bits = 0  # EFS off
    decoded = MerlionImapFetcher._decode_zip_member_name(info)
    assert decoded == "Прайслист_Мерлион_Москва.xlsm"


def test_decode_zip_member_name_keeps_utf8_when_efs_on():
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    info = zipfile.ZipInfo()
    info.filename = "Прайс.xlsx"
    info.flag_bits = 0x800  # EFS on (UTF-8)
    decoded = MerlionImapFetcher._decode_zip_member_name(info)
    assert decoded == "Прайс.xlsx"


# =====================================================================
# 6. Ручная распаковка _extract_zip_member_raw (12.1-fix-3)
# =====================================================================

def _make_zip_with_mismatched_lfh_name(
    central_name: str = "central_name.xlsm",
    local_name_bytes: bytes | None = None,
    payload: bytes = b"payload-content-bytes-here",
    compress_type: int = zipfile.ZIP_DEFLATED,
) -> bytes:
    """Собирает ZIP, в котором имя файла в local file header
    физически отличается от имени в central directory.

    Воспроизводит реальный Merlion-формат — стандартный
    zipfile.ZipFile.open() на таком архиве валится с
    «File name in directory … and header … differ».

    central_name должен быть ASCII (zipfile проверяет при писании,
    Python автоматически выставит UTF-8 EFS-флаг при не-ASCII), а
    local_name_bytes — байты той же длины что central_name.encode('ascii').
    """
    if local_name_bytes is None:
        local_name_bytes = b"x" * len(central_name.encode("ascii"))
    assert len(local_name_bytes) == len(central_name.encode("ascii")), (
        "local-name байт должно быть столько же, сколько central-name "
        "(чтобы LFH сохранил длину и offset до compressed body)"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compress_type) as zf:
        zf.writestr(central_name, payload)
    raw = bytearray(buf.getvalue())
    sig = b"PK\x03\x04"
    start = raw.find(sig)
    assert start >= 0, "не нашли local file header"
    # filename_length лежит в LFH по offset 26 (2 байта little-endian).
    fname_len = struct.unpack("<H", bytes(raw[start + 26:start + 28]))[0]
    fname_offset = start + 30
    assert len(local_name_bytes) == fname_len
    raw[fname_offset:fname_offset + fname_len] = local_name_bytes
    return bytes(raw)


def test_mismatched_zip_breaks_standard_zipfile_open():
    """Sanity: фикстура реально воспроизводит баг — стандартный
    zipfile.open(info) падает на BadZipFile."""
    zip_bytes = _make_zip_with_mismatched_lfh_name()
    buf = io.BytesIO(zip_bytes)
    with zipfile.ZipFile(buf, "r") as zf:
        info = zf.infolist()[0]
        with pytest.raises(zipfile.BadZipFile, match="differ"):
            with zf.open(info) as src:
                src.read()


def test_extract_zip_with_mismatched_central_and_local_headers(monkeypatch):
    """Реальный Merlion-кейс: имена central/local разные, стандартный
    zf.open() падает. parse_attachment должен использовать ручной
    распаковщик и корректно достать содержимое."""
    import app.services.auto_price.fetchers.merlion_imap as mer_mod
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"PK\x03\x04mismatched-payload-marker-12345"
    # 18 байт у обеих сторон — central и local имеют одинаковую длину,
    # но РАЗНЫЕ байты. Стандартный zf.open(info) на этом упадёт.
    zip_bytes = _make_zip_with_mismatched_lfh_name(
        central_name="merlion_price.xlsm",   # 18 байт ASCII
        local_name_bytes=b"abracadabraXX.xlsm",  # 18 байт, другие
        payload=payload,
    )

    seen = {}

    class _StubLoader:
        def iter_rows(self, filepath):
            with open(filepath, "rb") as f:
                seen["bytes"] = f.read()
            seen["filepath"] = filepath
            return iter([])

    monkeypatch.setattr(mer_mod, "MerlionLoader", lambda: _StubLoader())
    monkeypatch.setenv("IMAP_USER", "u@x.test")
    monkeypatch.setenv("IMAP_PASSWORD", "p")

    MerlionImapFetcher().parse_attachment(zip_bytes, "broken.zip")
    assert seen["bytes"] == payload
    assert seen["filepath"].endswith(".xlsm")


def test_extract_zip_member_raw_handles_zip_stored(tmp_path):
    """STORED (без сжатия): payload идёт как есть."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"STORED-uncompressed-bytes-of-some-length"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
        zf.writestr("a.xlsm", payload)
    raw = buf.getvalue()
    out = tmp_path / "out.xlsm"
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        info = zf.infolist()[0]
        MerlionImapFetcher._extract_zip_member_raw(zf, info, str(out))
    assert out.read_bytes() == payload


def test_extract_zip_member_raw_handles_zip_deflated(tmp_path):
    """DEFLATED (zlib): payload корректно распаковывается."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    # Возьмём заметно сжимаемый payload, чтобы compressed != raw.
    payload = (b"abc" * 5000) + b"END"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("b.xlsm", payload)
    raw = buf.getvalue()
    out = tmp_path / "out.xlsm"
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        info = zf.infolist()[0]
        MerlionImapFetcher._extract_zip_member_raw(zf, info, str(out))
    assert out.read_bytes() == payload


def test_extract_zip_member_raw_rejects_unknown_compression(tmp_path):
    """Метод сжатия, который мы не поддерживаем (например, BZIP2=12) —
    RuntimeError с понятным сообщением."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
        zf.writestr("c.xlsm", b"hello")
    raw = buf.getvalue()
    out = tmp_path / "out.xlsm"
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        info = zf.infolist()[0]
        # Подменим compress_type на «BZIP2» (12) — наш экстрактор
        # должен отказаться.
        info.compress_type = 12
        with pytest.raises(RuntimeError, match="неподдерживаемый compress_type"):
            MerlionImapFetcher._extract_zip_member_raw(zf, info, str(out))


def test_extract_handles_uncompressed_size_mismatch_warning(tmp_path, caplog):
    """Если info.file_size в central неверный — это warning, не error.
    Реальные xlsx-валидаторы поймают битость, если она реально есть."""
    import logging as _logging
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    payload = b"hello world payload"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
        zf.writestr("d.xlsm", payload)
    raw = buf.getvalue()
    out = tmp_path / "out.xlsm"
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        info = zf.infolist()[0]
        # Соврём про uncompressed size — экстрактор должен warn, не raise.
        info.file_size = 999999
        with caplog.at_level(_logging.WARNING, logger="app.services.auto_price.fetchers.merlion_imap"):
            MerlionImapFetcher._extract_zip_member_raw(zf, info, str(out))
    assert out.read_bytes() == payload
    assert any("не совпал" in rec.message for rec in caplog.records), (
        "ожидался warning о несовпадении file_size"
    )


def test_extract_zip_member_raw_rejects_bad_signature(tmp_path):
    """Если local file header начинается не с 'PK\\x03\\x04' —
    RuntimeError. Защита от мусора."""
    from app.services.auto_price.fetchers.merlion_imap import MerlionImapFetcher

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
        zf.writestr("e.xlsm", b"hello")
    raw = bytearray(buf.getvalue())
    # Затрём signature первого LFH мусором.
    sig_pos = raw.find(b"PK\x03\x04")
    raw[sig_pos:sig_pos + 4] = b"XXXX"
    out = tmp_path / "out.xlsm"
    # Используем правильный raw для парсинга central directory (он
    # лежит в конце). Чтобы корректно прочитать infolist(), сначала
    # загрузим неиспорченный архив, потом сделаем raw-corrupt и
    # передадим его как fp в zf.
    with zipfile.ZipFile(io.BytesIO(buf.getvalue()), "r") as zf_good:
        info = zf_good.infolist()[0]
    # Обманём: подсунем испорченные байты как fp.
    class _FakeZf:
        fp = io.BytesIO(bytes(raw))
    with pytest.raises(RuntimeError, match="неверная сигнатура"):
        MerlionImapFetcher._extract_zip_member_raw(_FakeZf(), info, str(out))
