"""Тесты Excel-импорта каталога (Фаза 3 плана 2026-05-13).

Fixture-файлы генерируются inline через openpyxl — не зависим от
параллельного Export-чата. Покрытие сценариев из DoD брифа:
  - id есть → UPDATE редактируемых полей;
  - id пустой + name заполнен → INSERT нового товара;
  - id есть, не найден в БД → skip + предупреждение;
  - валидационные ошибки в одной строке не валят валидные строки;
  - read-only колонки (цены, поставщик, даты) игнорируются с записью в
    отчёт;
  - audit_log пишется при успехе и при провале.

Тесты UI-эндпоинта /databases/catalog-excel/upload/{kind} проверяют
доступы (admin only) и JSON-структуру ответа.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import text

from portal.services.catalog.excel_import import (
    PC_SHEETS,
    PRINTER_SHEETS,
    import_components_pc,
    import_printers_mfu,
)


# ---------------------------------------------------------------------------
# Хелперы для построения xlsx-фикстур
# ---------------------------------------------------------------------------


def _write_header_rows(ws, headers: list[str]) -> None:
    """Пишет служебную строку 1 (курс), пустую строку 2 и шапку в строке 3
    — так же, как это будет делать Фаза 2 экспорта (план фиксирован)."""
    ws.cell(row=1, column=1, value="Курс ЦБ (USD→RUB)")
    ws.cell(row=1, column=2, value=97.45)
    # Row 2 — пустая.
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=h)


def _make_cpu_xlsx(tmp_path: Path, rows: list[dict], *, include_ro: bool = False) -> Path:
    """Строит минимальный xlsx с одним листом CPU."""
    wb = Workbook()
    # Дефолтный лист — переименуем под CPU; остальные 7 листов добавлять
    # не обязательно (import пропускает отсутствующие).
    ws = wb.active
    ws.title = PC_SHEETS["cpu"]["sheet_name"]

    headers = [
        "id", "model", "manufacturer", "sku", "gtin", "is_hidden",
        "socket", "cores", "threads", "base_clock_ghz", "turbo_clock_ghz",
        "tdp_watts", "has_integrated_graphics", "memory_type", "package_type",
        "process_nm", "l3_cache_mb", "max_memory_freq", "release_year",
    ]
    if include_ro:
        headers += ["Цена min, USD", "Цена min, RUB", "Поставщик (min)", "Цена обновлена"]

    _write_header_rows(ws, headers)

    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, h in enumerate(headers, start=1):
            v = row_data.get(h)
            if v is not None:
                ws.cell(row=row_idx, column=col_idx, value=v)

    path = tmp_path / "cpu.xlsx"
    wb.save(path)
    return path


def _make_printer_xlsx(tmp_path: Path, rows: list[dict]) -> Path:
    """Минимальный xlsx с листом «Принтеры»."""
    wb = Workbook()
    ws = wb.active
    ws.title = PRINTER_SHEETS["printer"]["sheet_name"]

    headers = [
        "id", "sku", "mpn", "gtin", "brand", "name", "category",
        "ktru_codes_array", "is_hidden", "cost_base_rub", "margin_pct_target",
        # attrs:
        "print_speed_ppm", "colorness", "max_format", "duplex",
        "resolution_dpi", "network_interface", "usb",
        "starter_cartridge_pages", "print_technology",
        "weight_kg", "box_width_cm", "box_height_cm", "box_depth_cm",
    ]
    _write_header_rows(ws, headers)
    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, h in enumerate(headers, start=1):
            v = row_data.get(h)
            if v is not None:
                ws.cell(row=row_idx, column=col_idx, value=v)

    path = tmp_path / "printer.xlsx"
    wb.save(path)
    return path


def _insert_cpu(db_session, *, model: str = "Intel Core i5-12400F", **kw) -> int:
    """Удобный INSERT для тестов. id возвращает."""
    row = db_session.execute(
        text(
            "INSERT INTO cpus (model, manufacturer, sku, socket, cores, threads, "
            "  base_clock_ghz, turbo_clock_ghz, tdp_watts, has_integrated_graphics, "
            "  memory_type, package_type) "
            "VALUES (:model, :manufacturer, :sku, :socket, :cores, :threads, "
            "  :base, :turbo, :tdp, :igpu, :mem, :pkg) "
            "RETURNING id"
        ),
        {
            "model": model,
            "manufacturer": kw.get("manufacturer", "Intel"),
            "sku":          kw.get("sku", "SKU-1"),
            "socket":       kw.get("socket", "LGA1700"),
            "cores":        kw.get("cores", 6),
            "threads":      kw.get("threads", 12),
            "base":         kw.get("base", 2.5),
            "turbo":        kw.get("turbo", 4.4),
            "tdp":          kw.get("tdp", 65),
            "igpu":         kw.get("igpu", True),
            "mem":          kw.get("mem", "DDR4"),
            "pkg":          kw.get("pkg", "BOX"),
        },
    ).first()
    db_session.commit()
    return int(row.id)


def _insert_printer(
    db_session,
    *,
    sku: str = "PRN-1",
    brand: str = "Pantum",
    name: str = "P3010DW",
    category: str = "printer",
    attrs: dict | None = None,
) -> int:
    row = db_session.execute(
        text(
            "INSERT INTO printers_mfu (sku, brand, name, category, attrs_jsonb) "
            "VALUES (:sku, :brand, :name, :category, CAST(:attrs AS JSONB)) "
            "RETURNING id"
        ),
        {
            "sku": sku, "brand": brand, "name": name, "category": category,
            "attrs": json.dumps(attrs or {}, ensure_ascii=False),
        },
    ).first()
    db_session.commit()
    return int(row.id)


# ---------------------------------------------------------------------------
# Чистка таблиц (тесты могут писать в cpus / printers_mfu).
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _clean_catalog(db_engine):
    with db_engine.begin() as conn:
        conn.execute(text(
            "TRUNCATE TABLE "
            "  printers_mfu, cpus, motherboards, rams, gpus, storages, "
            "  cases, psus, coolers "
            "RESTART IDENTITY CASCADE"
        ))
    yield


# ---------------------------------------------------------------------------
# 1. UPDATE существующей строки по id
# ---------------------------------------------------------------------------


def test_pc_update_existing_cpu_by_id(db_session, tmp_path):
    cpu_id = _insert_cpu(db_session, model="Intel Core i5-12400F", cores=6)

    xlsx = _make_cpu_xlsx(tmp_path, [
        {
            "id": cpu_id,
            "model": "Intel Core i5-12400F",
            "manufacturer": "Intel",
            "sku": "SKU-1",
            "is_hidden": False,
            "socket": "LGA1700",
            "cores": 6,
            "threads": 12,
            "base_clock_ghz": 2.5,
            "turbo_clock_ghz": 4.4,
            "tdp_watts": 100,         # ← изменено
            "has_integrated_graphics": True,
            "memory_type": "DDR5",    # ← изменено
            "package_type": "BOX",
        },
    ])

    report = import_components_pc(xlsx, user_id=1, session=db_session)

    assert report.updated == 1
    assert report.inserted == 0
    assert report.skipped == 0
    assert report.error_count == 0

    row = db_session.execute(
        text("SELECT tdp_watts, memory_type FROM cpus WHERE id = :id"),
        {"id": cpu_id},
    ).first()
    assert row.tdp_watts == 100
    assert row.memory_type == "DDR5"


# ---------------------------------------------------------------------------
# 2. INSERT новой строки (id пустой)
# ---------------------------------------------------------------------------


def test_pc_insert_new_cpu_when_id_blank(db_session, tmp_path):
    xlsx = _make_cpu_xlsx(tmp_path, [
        {
            # id отсутствует
            "model": "AMD Ryzen 5 7600X",
            "manufacturer": "AMD",
            "sku": "100-100000593WOF",
            "is_hidden": False,
            "socket": "AM5",
            "cores": 6,
            "threads": 12,
            "base_clock_ghz": 4.7,
            "turbo_clock_ghz": 5.3,
            "tdp_watts": 105,
            "has_integrated_graphics": True,
            "memory_type": "DDR5",
            "package_type": "BOX",
        },
    ])

    report = import_components_pc(xlsx, user_id=1, session=db_session)

    assert report.inserted == 1
    assert report.updated == 0
    assert report.error_count == 0

    row = db_session.execute(
        text("SELECT model, manufacturer, socket FROM cpus WHERE model = :m"),
        {"m": "AMD Ryzen 5 7600X"},
    ).first()
    assert row is not None
    assert row.manufacturer == "AMD"
    assert row.socket == "AM5"


# ---------------------------------------------------------------------------
# 3. id есть, но в БД не найден → skip + warning
# ---------------------------------------------------------------------------


def test_pc_skip_unknown_id_with_warning(db_session, tmp_path):
    xlsx = _make_cpu_xlsx(tmp_path, [
        {
            "id": 99999,
            "model": "Phantom CPU",
            "manufacturer": "X",
            "socket": "X",
            "cores": 1, "threads": 1,
            "base_clock_ghz": 1.0, "turbo_clock_ghz": 1.0,
            "tdp_watts": 10, "has_integrated_graphics": False,
            "memory_type": "DDR4", "package_type": "OEM",
        },
    ])
    report = import_components_pc(xlsx, user_id=1, session=db_session)

    assert report.skipped == 1
    assert report.updated == 0
    assert report.inserted == 0
    assert any("99999" in w for w in report.warnings)


# ---------------------------------------------------------------------------
# 4. Валидационная ошибка в одной строке не валит другие
# ---------------------------------------------------------------------------


def test_pc_validation_error_does_not_break_valid_rows(db_session, tmp_path):
    valid_id = _insert_cpu(db_session, model="Intel i5", sku="OK")

    xlsx = _make_cpu_xlsx(tmp_path, [
        {
            # ОШИБКА: cores не int
            "id": valid_id,
            "model": "Intel i5",
            "manufacturer": "Intel",
            "socket": "LGA1700",
            "cores": "не-число",
            "threads": 12,
            "base_clock_ghz": 2.5, "turbo_clock_ghz": 4.4,
            "tdp_watts": 65, "has_integrated_graphics": True,
            "memory_type": "DDR4", "package_type": "BOX",
        },
        {
            # Валидная строка — INSERT
            "model": "AMD Ryzen 7 7700X",
            "manufacturer": "AMD",
            "socket": "AM5",
            "cores": 8, "threads": 16,
            "base_clock_ghz": 4.5, "turbo_clock_ghz": 5.4,
            "tdp_watts": 105, "has_integrated_graphics": True,
            "memory_type": "DDR5", "package_type": "BOX",
        },
    ])

    report = import_components_pc(xlsx, user_id=1, session=db_session)

    assert report.error_count == 1, report.errors
    assert "cores" in report.errors[0].message
    # Валидная строка должна была вставиться.
    assert report.inserted == 1
    row = db_session.execute(
        text("SELECT model FROM cpus WHERE model = :m"),
        {"m": "AMD Ryzen 7 7700X"},
    ).first()
    assert row is not None


# ---------------------------------------------------------------------------
# 5. Read-only колонки игнорируются + warning в отчёт
# ---------------------------------------------------------------------------


def test_pc_readonly_columns_ignored_and_warning(db_session, tmp_path):
    cpu_id = _insert_cpu(db_session, model="Intel i5", sku="OK")

    xlsx = _make_cpu_xlsx(
        tmp_path,
        [{
            "id": cpu_id,
            "model": "Intel i5",
            "manufacturer": "Intel",
            "socket": "LGA1700",
            "cores": 6, "threads": 12,
            "base_clock_ghz": 2.5, "turbo_clock_ghz": 4.4,
            "tdp_watts": 65, "has_integrated_graphics": True,
            "memory_type": "DDR4", "package_type": "BOX",
            # read-only колонки — должны игнорироваться:
            "Цена min, USD": 999.99,
            "Цена min, RUB": 99999.99,
            "Поставщик (min)": "Несуществующий поставщик",
            "Цена обновлена": "2026-05-14",
        }],
        include_ro=True,
    )

    report = import_components_pc(xlsx, user_id=1, session=db_session)

    assert report.updated == 1
    assert report.error_count == 0

    # Главное — read-only колонки нигде не отразились (нет колонки price/
    # supplier в cpus, поэтому INSERT/UPDATE их обойти не мог бы — но
    # проверяем явно, что warning о пропуске присутствует).
    ro_warning = [w for w in report.warnings if "read-only columns ignored" in w]
    assert ro_warning, report.warnings
    assert "Цена min, USD" in ro_warning[0]


def test_pc_availability_columns_are_readonly(db_session, tmp_path):
    """Колонки наличия (Склад/Транзит/Поставщиков) — read-only: значения
    в Excel игнорируются на импорте, имена попадают в общий ro-warning."""
    cpu_id = _insert_cpu(db_session, model="Intel i5", sku="AVAIL-RO")

    # Самостоятельно строим лист с колонками наличия (хелпер _make_cpu_xlsx
    # их не пишет — это новые колонки, не вошедшие в исходную фикстуру).
    wb = Workbook()
    ws = wb.active
    ws.title = PC_SHEETS["cpu"]["sheet_name"]
    headers = [
        "id", "model", "manufacturer", "sku", "gtin", "is_hidden",
        "socket", "cores", "threads", "base_clock_ghz", "turbo_clock_ghz",
        "tdp_watts", "has_integrated_graphics", "memory_type", "package_type",
        "process_nm", "l3_cache_mb", "max_memory_freq", "release_year",
        # Все ro-колонки, включая три новых:
        "Цена min, USD", "Цена min, RUB", "Поставщик (min)", "Цена обновлена",
        "Склад, шт", "Транзит, шт", "Поставщиков, шт",
    ]
    # Шапка через хелпер; на нём же модуль-уровневые row1/row2/row3.
    ws.cell(row=1, column=1, value="Курс ЦБ (USD→RUB)")
    ws.cell(row=1, column=2, value=97.45)
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=h)
    row_data = {
        "id": cpu_id,
        "model": "Intel i5",
        "manufacturer": "Intel",
        "socket": "LGA1700",
        "cores": 6, "threads": 12,
        "base_clock_ghz": 2.5, "turbo_clock_ghz": 4.4,
        "tdp_watts": 65, "has_integrated_graphics": True,
        "memory_type": "DDR4", "package_type": "BOX",
        # «Якобы правки» в ro-колонках наличия — должны быть проигнорированы.
        "Склад, шт": 99999,
        "Транзит, шт": 88888,
        "Поставщиков, шт": 77,
    }
    for col_idx, h in enumerate(headers, start=1):
        v = row_data.get(h)
        if v is not None:
            ws.cell(row=4, column=col_idx, value=v)
    path = tmp_path / "cpu_avail.xlsx"
    wb.save(path)

    report = import_components_pc(path, user_id=1, session=db_session)

    assert report.updated == 1
    assert report.error_count == 0

    # Имена трёх новых ro-колонок попали в общий ro-warning.
    ro_warning = [w for w in report.warnings if "read-only columns ignored" in w]
    assert ro_warning, report.warnings
    assert "Склад, шт"       in ro_warning[0]
    assert "Транзит, шт"     in ro_warning[0]
    assert "Поставщиков, шт" in ro_warning[0]


# ---------------------------------------------------------------------------
# 6. Полностью пустые строки → пропуск без ошибки
# ---------------------------------------------------------------------------


def test_pc_completely_empty_rows_not_an_error(db_session, tmp_path):
    cpu_id = _insert_cpu(db_session, model="Intel i5")

    # Лист с одной валидной строкой UPDATE и одной полностью пустой строкой.
    xlsx = _make_cpu_xlsx(tmp_path, [
        {
            "id": cpu_id,
            "model": "Intel i5",
            "manufacturer": "Intel",
            "socket": "LGA1700",
            "cores": 6, "threads": 12,
            "base_clock_ghz": 2.5, "turbo_clock_ghz": 4.4,
            "tdp_watts": 65, "has_integrated_graphics": True,
            "memory_type": "DDR4", "package_type": "BOX",
        },
        {},  # пустая
    ])

    report = import_components_pc(xlsx, user_id=1, session=db_session)
    assert report.error_count == 0
    assert report.updated == 1


# ---------------------------------------------------------------------------
# 7. Принтеры: per-key merge attrs_jsonb
# ---------------------------------------------------------------------------


def test_printers_attrs_per_key_merge(db_session, tmp_path):
    pid = _insert_printer(
        db_session,
        sku="PRN-MERGE",
        attrs={
            "print_speed_ppm": 30,
            "colorness":       "ч/б",
            "max_format":      "A4",
            "duplex":          "yes",
            "resolution_dpi":  1200,
            "network_interface": ["LAN"],
            "usb":             "yes",
            "starter_cartridge_pages": 1500,
            "print_technology": "лазерная",
        },
    )

    # В Excel — только colorness и max_format заполнены; остальные пустые
    # → НЕ должны затирать существующие значения (per-key merge).
    xlsx = _make_printer_xlsx(tmp_path, [{
        "id": pid,
        "sku": "PRN-MERGE",
        "brand": "Pantum",
        "name": "P3010DW",
        "category": "printer",
        "colorness": "цветной",   # ← изменено
        "max_format": "A3",        # ← изменено
        # остальные attrs ячейки пустые
    }])

    report = import_printers_mfu(xlsx, user_id=1, session=db_session)
    assert report.updated == 1
    assert report.error_count == 0

    row = db_session.execute(
        text("SELECT attrs_jsonb FROM printers_mfu WHERE id = :id"),
        {"id": pid},
    ).first()
    attrs = row.attrs_jsonb
    if isinstance(attrs, str):
        attrs = json.loads(attrs)
    assert attrs["colorness"] == "цветной"
    assert attrs["max_format"] == "A3"
    # Не затёрто:
    assert attrs["print_speed_ppm"] == 30
    assert attrs["duplex"] == "yes"
    assert attrs["network_interface"] == ["LAN"]


# ---------------------------------------------------------------------------
# 8. Принтеры: INSERT нового с attrs
# ---------------------------------------------------------------------------


def test_printers_insert_with_attrs(db_session, tmp_path):
    xlsx = _make_printer_xlsx(tmp_path, [{
        "sku": "PRN-NEW",
        "brand": "HP",
        "name": "LaserJet 1020",
        "category": "printer",
        "colorness": "ч/б",
        "max_format": "A4",
        "duplex": "no",
        "print_speed_ppm": 18,
    }])

    report = import_printers_mfu(xlsx, user_id=1, session=db_session)
    assert report.inserted == 1
    assert report.error_count == 0

    row = db_session.execute(
        text("SELECT brand, attrs_jsonb FROM printers_mfu WHERE sku = :s"),
        {"s": "PRN-NEW"},
    ).first()
    assert row.brand == "HP"
    attrs = row.attrs_jsonb
    if isinstance(attrs, str):
        attrs = json.loads(attrs)
    assert attrs["colorness"] == "ч/б"
    assert attrs["print_speed_ppm"] == 18


# ---------------------------------------------------------------------------
# 9. Принтеры: 'n/a' пишется как строка 'n/a', enum валидируется
# ---------------------------------------------------------------------------


def test_printers_na_marker_and_enum_validation(db_session, tmp_path):
    xlsx = _make_printer_xlsx(tmp_path, [
        # Валидная строка с n/a-маркерами.
        {
            "sku": "PRN-NA",
            "brand": "Generic",
            "name": "GP-100",
            "category": "printer",
            "colorness": "n/a",
            "max_format": "A4",
        },
        # Строка с НЕВАЛИДНЫМ enum — пойдёт в errors, остальные применятся.
        {
            "sku": "PRN-BAD",
            "brand": "Generic",
            "name": "GP-200",
            "category": "printer",
            "colorness": "розовый",    # ← invalid
            "max_format": "A4",
        },
    ])

    report = import_printers_mfu(xlsx, user_id=1, session=db_session)
    # Валидная строка должна была вставиться.
    assert report.inserted == 1
    assert report.error_count == 1
    assert "colorness" in report.errors[0].message

    row = db_session.execute(
        text("SELECT attrs_jsonb FROM printers_mfu WHERE sku = :s"),
        {"s": "PRN-NA"},
    ).first()
    assert row is not None
    attrs = row.attrs_jsonb
    if isinstance(attrs, str):
        attrs = json.loads(attrs)
    assert attrs["colorness"] == "n/a"


# ---------------------------------------------------------------------------
# 10. Массивы (supported_form_factors у case)
# ---------------------------------------------------------------------------


def test_pc_arrays_split_by_comma(db_session, tmp_path):
    # Минимальный CASE-лист, чтобы проверить TEXT[] серилизацию.
    wb = Workbook()
    ws = wb.active
    ws.title = PC_SHEETS["case"]["sheet_name"]
    headers = [
        "id", "model", "manufacturer", "sku", "gtin", "is_hidden",
        "supported_form_factors", "has_psu_included",
    ]
    _write_header_rows(ws, headers)
    for col_idx, h in enumerate(headers, start=1):
        v = {
            "model": "DeepCool MATREXX 55",
            "manufacturer": "DeepCool",
            "is_hidden": False,
            "supported_form_factors": "ATX,mATX,ITX",
            "has_psu_included": False,
        }.get(h)
        if v is not None:
            ws.cell(row=4, column=col_idx, value=v)
    path = tmp_path / "case.xlsx"
    wb.save(path)

    report = import_components_pc(path, user_id=1, session=db_session)
    assert report.inserted == 1
    assert report.error_count == 0

    row = db_session.execute(
        text("SELECT supported_form_factors FROM cases WHERE model = :m"),
        {"m": "DeepCool MATREXX 55"},
    ).first()
    assert set(row.supported_form_factors) == {"ATX", "mATX", "ITX"}


# ---------------------------------------------------------------------------
# UI-эндпоинт: доступы и audit_log
# ---------------------------------------------------------------------------


pytestmark_endpoint = []


def _login_via_portal(client, login: str, password: str) -> None:
    """Login helper — повторяет логику tests/test_portal/conftest._login_via_portal,
    но импортирован сюда напрямую (test_catalog не использует conftest portal'а)."""
    import re
    r = client.get("/login")
    assert r.status_code == 200
    token_match = re.search(r'name="csrf_token" value="([^"]+)"', r.text)
    assert token_match
    token = token_match.group(1)
    r = client.post(
        "/login",
        data={"login": login, "password": password, "csrf_token": token},
    )
    assert r.status_code in (302, 303)


@pytest.fixture()
def portal_client_local():
    from fastapi.testclient import TestClient
    from portal.main import app
    with TestClient(app, follow_redirects=False) as c:
        yield c


@pytest.fixture()
def admin_user_local(db_session):
    """Создаём admin'а локально, не зависим от portal/conftest.py."""
    import json as _json
    from shared.auth import hash_password
    row = db_session.execute(
        text(
            "INSERT INTO users (login, password_hash, role, name, permissions) "
            "VALUES (:l, :p, :r, :n, CAST(:perms AS JSONB)) RETURNING id"
        ),
        {
            "l": "admin_x", "p": hash_password("admin-pass"),
            "r": "admin", "n": "Admin", "perms": _json.dumps({}),
        },
    ).first()
    db_session.commit()
    return {"id": int(row.id), "login": "admin_x", "password": "admin-pass"}


@pytest.fixture()
def manager_user_local(db_session):
    import json as _json
    from shared.auth import hash_password
    row = db_session.execute(
        text(
            "INSERT INTO users (login, password_hash, role, name, permissions) "
            "VALUES (:l, :p, :r, :n, CAST(:perms AS JSONB)) RETURNING id"
        ),
        {
            "l": "manager_x", "p": hash_password("man-pass"),
            "r": "manager", "n": "Man", "perms": _json.dumps({"configurator": True}),
        },
    ).first()
    db_session.commit()
    return {"id": int(row.id), "login": "manager_x", "password": "man-pass"}


@pytest.fixture(autouse=True)
def _truncate_users_audit(db_engine):
    """Очищаем users и audit_log между тестами эндпоинта."""
    with db_engine.begin() as conn:
        conn.execute(text(
            "TRUNCATE TABLE audit_log, users RESTART IDENTITY CASCADE"
        ))
    yield


def test_endpoint_admin_can_upload_and_audit_written(
    portal_client_local, admin_user_local, db_session, tmp_path,
):
    _login_via_portal(
        portal_client_local, admin_user_local["login"], admin_user_local["password"],
    )

    cpu_id = _insert_cpu(db_session, model="Intel i5", sku="OK")
    xlsx = _make_cpu_xlsx(tmp_path, [{
        "id": cpu_id,
        "model": "Intel i5",
        "manufacturer": "Intel",
        "socket": "LGA1700",
        "cores": 6, "threads": 12,
        "base_clock_ghz": 2.5, "turbo_clock_ghz": 4.4,
        "tdp_watts": 88,   # ← новое значение
        "has_integrated_graphics": True,
        "memory_type": "DDR4", "package_type": "BOX",
    }])

    # Достаём csrf_token со страницы /login (он же годится для AJAX —
    # см. auth.get_csrf_token, токен per-session).
    r = portal_client_local.get("/")
    import re
    token_match = re.search(r'name="csrf_token" value="([^"]+)"', r.text)
    # Если на главной нет токена (не у всех страниц форма) — берём со
    # страницы /settings/users (она админская и точно даст csrf).
    if not token_match:
        r2 = portal_client_local.get("/settings/users")
        token_match = re.search(r'name="csrf_token" value="([^"]+)"', r2.text)
    assert token_match, "не удалось извлечь csrf_token"
    csrf = token_match.group(1)

    with open(xlsx, "rb") as fh:
        resp = portal_client_local.post(
            "/databases/catalog-excel/upload/pc",
            data={"csrf_token": csrf},
            files={"uploaded_file": ("cpu.xlsx", fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 200, resp.text
    payload = resp.json()
    assert payload["updated"] == 1
    assert payload["errors_count"] == 0
    assert payload["saved_path"].endswith(".xlsx")

    # Файл реально сохранился.
    saved = Path(payload["saved_path"])
    assert saved.exists()
    # Зачищаем за собой.
    saved.unlink(missing_ok=True)

    # БД действительно обновилась.
    row = db_session.execute(
        text("SELECT tdp_watts FROM cpus WHERE id = :id"),
        {"id": cpu_id},
    ).first()
    assert row.tdp_watts == 88

    # audit_log пишется.
    audit = db_session.execute(
        text(
            "SELECT action, target_type, target_id, payload FROM audit_log "
            "WHERE action = :a ORDER BY id DESC LIMIT 1"
        ),
        {"a": "catalog_excel_import"},
    ).first()
    assert audit is not None
    assert audit.target_type == "catalog_excel"
    assert audit.target_id == "pc"
    payload_db = audit.payload
    if isinstance(payload_db, str):
        payload_db = json.loads(payload_db)
    assert payload_db["target"] == "pc"
    assert payload_db["updated"] == 1


def test_endpoint_manager_gets_403(portal_client_local, manager_user_local, tmp_path):
    _login_via_portal(
        portal_client_local, manager_user_local["login"], manager_user_local["password"],
    )

    # Берём csrf-токен любого доступного для менеджера экрана.
    r = portal_client_local.get("/")
    import re
    m = re.search(r'name="csrf_token" value="([^"]+)"', r.text)
    csrf = m.group(1) if m else ""

    xlsx_path = tmp_path / "empty.xlsx"
    Workbook().save(xlsx_path)

    with open(xlsx_path, "rb") as fh:
        resp = portal_client_local.post(
            "/databases/catalog-excel/upload/pc",
            data={"csrf_token": csrf},
            files={"uploaded_file": ("empty.xlsx", fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 403


def test_endpoint_rejects_invalid_kind(portal_client_local, admin_user_local, tmp_path):
    _login_via_portal(
        portal_client_local, admin_user_local["login"], admin_user_local["password"],
    )
    r = portal_client_local.get("/")
    import re
    m = re.search(r'name="csrf_token" value="([^"]+)"', r.text)
    if not m:
        r2 = portal_client_local.get("/settings/users")
        m = re.search(r'name="csrf_token" value="([^"]+)"', r2.text)
    csrf = m.group(1)

    xlsx_path = tmp_path / "x.xlsx"
    Workbook().save(xlsx_path)

    with open(xlsx_path, "rb") as fh:
        resp = portal_client_local.post(
            "/databases/catalog-excel/upload/unknown_kind",
            data={"csrf_token": csrf},
            files={"uploaded_file": ("x.xlsx", fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 400
