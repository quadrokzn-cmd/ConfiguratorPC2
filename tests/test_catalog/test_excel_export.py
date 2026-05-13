# Тесты Excel-выгрузки каталога (Фаза 2 плана 2026-05-13).
#
# Проверяем именно структуру файла глазами openpyxl-а — это контракт
# Фазы 3 (importer), а не текст в ячейках.
#
# Чек-лист (см. DoD в промте):
#   - 8 листов для PC и 2 для printers, имена совпадают с _SheetSpec;
#   - служебная строка 1 содержит «Курс ЦБ …» в A1 и число в B1;
#   - на шапке (строка 3) висит autofilter;
#   - первая колонка (id) скрыта;
#   - для USD-товара RUB-ячейка содержит формулу «=…*$B$1»;
#   - для RUB-товара RUB-ячейка — статическое число, USD пустой;
#   - TEXT[]-колонка сериализуется через запятую;
#   - attrs_jsonb-ключи у printers_mfu пишутся в свои колонки;
#   - exchange_rates пустой → fallback 90.0 (rate_is_fallback=True).

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

from portal.services.catalog.excel_export import (
    _FALLBACK_RATE,
    default_filename,
    export_components_pc,
    export_printers_mfu,
)


# ---------------------------------------------------------------------
# Утилиты: вставка фикстурных данных
# ---------------------------------------------------------------------

def _insert_supplier(session, name: str = "Поставщик-Тест") -> int:
    row = session.execute(
        text(
            "INSERT INTO suppliers (name, is_active) "
            "VALUES (:n, TRUE) RETURNING id"
        ),
        {"n": name},
    ).first()
    session.commit()
    return int(row.id)


def _insert_cpu(
    session, model: str = "Intel Core i5-12400F", *, manufacturer: str = "Intel",
) -> int:
    row = session.execute(
        text(
            "INSERT INTO cpus (model, manufacturer, sku, gtin, socket, "
            "  cores, threads, base_clock_ghz, turbo_clock_ghz, tdp_watts, "
            "  has_integrated_graphics, memory_type, package_type) "
            "VALUES (:m, :mfr, 'CPU-SKU', '1234567890123', 'LGA1700', "
            "  6, 12, 2.5, 4.4, 65, FALSE, 'DDR4+DDR5', 'BOX') RETURNING id"
        ),
        {"m": model, "mfr": manufacturer},
    ).first()
    session.commit()
    return int(row.id)


def _insert_case(session, supported: list[str]) -> int:
    # Postgres ARRAY-литерал: ['ATX','mATX']
    row = session.execute(
        text(
            "INSERT INTO cases (model, manufacturer, supported_form_factors, "
            "  has_psu_included) "
            "VALUES ('Корпус ABC', 'Generic', CAST(:arr AS TEXT[]), FALSE) "
            "RETURNING id"
        ),
        {"arr": "{" + ",".join(supported) + "}"},
    ).first()
    session.commit()
    return int(row.id)


def _insert_supplier_price(
    session, *, supplier_id: int, category: str, component_id: int,
    price: Decimal, currency: str = "USD", stock: int = 5,
    transit: int = 0,
    updated_at: datetime | None = None,
) -> None:
    session.execute(
        text(
            "INSERT INTO supplier_prices "
            "  (supplier_id, category, component_id, price, currency, "
            "   stock_qty, transit_qty, updated_at) "
            "VALUES (:sid, :cat, :cid, :p, :cur, :st, :tr, "
            "  COALESCE(:upd, NOW()))"
        ),
        {
            "sid": supplier_id, "cat": category, "cid": component_id,
            "p": str(price), "cur": currency, "st": stock, "tr": transit,
            "upd": updated_at,
        },
    )
    session.commit()


def _insert_exchange_rate(
    session, rate: Decimal, rate_date: date,
) -> None:
    session.execute(
        text(
            "INSERT INTO exchange_rates (rate_date, rate_usd_rub, source) "
            "VALUES (:d, :r, 'cbr')"
        ),
        {"d": rate_date, "r": str(rate)},
    )
    session.commit()


def _insert_printer(
    session, *, sku: str, brand: str, name: str, category: str,
    attrs: dict | None = None, ktru: list[str] | None = None,
    attrs_source: str = "manual",
) -> int:
    import json
    row = session.execute(
        text(
            "INSERT INTO printers_mfu "
            "  (sku, brand, name, category, ktru_codes_array, attrs_jsonb, "
            "   attrs_source) "
            "VALUES (:sku, :brand, :name, :cat, CAST(:ktru AS TEXT[]), "
            "  CAST(:attrs AS JSONB), :src) "
            "RETURNING id"
        ),
        {
            "sku": sku, "brand": brand, "name": name, "cat": category,
            "ktru": "{" + ",".join(ktru or []) + "}",
            "attrs": json.dumps(attrs or {}, ensure_ascii=False),
            "src": attrs_source,
        },
    ).first()
    session.commit()
    return int(row.id)


# ---------------------------------------------------------------------
# Тесты структуры файла
# ---------------------------------------------------------------------

EXPECTED_PC_SHEETS = [
    "CPU", "Motherboard", "RAM", "GPU",
    "Storage", "Case", "PSU", "Cooler",
]

EXPECTED_PRINTERS_SHEETS = ["Принтеры", "МФУ"]


def test_export_components_pc_creates_8_sheets(tmp_path: Path, db_session):
    """Базовый смоук: файл создан, 8 листов с правильными именами."""
    out = tmp_path / "pc.xlsx"
    report = export_components_pc(out, db=db_session)

    assert out.exists()
    assert report.total_rows == 0
    assert list(report.sheet_counts.keys()) == EXPECTED_PC_SHEETS
    assert report.rate_is_fallback is True
    assert report.rate_used == _FALLBACK_RATE

    wb = load_workbook(out)
    assert wb.sheetnames == EXPECTED_PC_SHEETS


def test_export_printers_mfu_creates_2_sheets(tmp_path: Path, db_session):
    """Смоук: файл создан, 2 листа («Принтеры», «МФУ»)."""
    out = tmp_path / "pr.xlsx"
    report = export_printers_mfu(out, db=db_session)

    assert out.exists()
    assert list(report.sheet_counts.keys()) == EXPECTED_PRINTERS_SHEETS

    wb = load_workbook(out)
    assert wb.sheetnames == EXPECTED_PRINTERS_SHEETS


def test_each_pc_sheet_has_rate_cell_and_autofilter(tmp_path: Path, db_session):
    """На каждом из 8 листов есть служебная строка 1 (курс) и autofilter."""
    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    for sheet_name in EXPECTED_PC_SHEETS:
        ws = wb[sheet_name]
        assert ws["A1"].value == "Курс ЦБ (USD→RUB)", sheet_name
        # B1 — число (fallback 90.0 при пустой exchange_rates).
        assert isinstance(ws["B1"].value, (int, float)), sheet_name
        assert float(ws["B1"].value) == pytest.approx(float(_FALLBACK_RATE))

        # autofilter крепится к строке 3 (шапка) — проверяем что ref начинается
        # с "A3:" и заканчивается на "3".
        assert ws.auto_filter.ref is not None, sheet_name
        assert ws.auto_filter.ref.startswith("A3:"), ws.auto_filter.ref
        assert ws.auto_filter.ref.endswith("3"), ws.auto_filter.ref


def test_id_column_is_hidden(tmp_path: Path, db_session):
    """Первая колонка (id) должна быть скрытой на всех листах."""
    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    for sheet_name in EXPECTED_PC_SHEETS:
        ws = wb[sheet_name]
        # column_dimensions['A'].hidden = True
        assert ws.column_dimensions["A"].hidden is True, sheet_name
        # При этом заголовок A3 — 'id' (даже если колонка скрыта).
        assert ws.cell(row=3, column=1).value == "id"


# ---------------------------------------------------------------------
# Тесты RUB-формулы и сериализации
# ---------------------------------------------------------------------

def test_usd_priced_cpu_produces_rub_formula(tmp_path: Path, db_session):
    """У товара с USD-ценой колонка «Цена min, RUB» содержит формулу."""
    sid = _insert_supplier(db_session, "USDSup")
    cpu_id = _insert_cpu(db_session, "USD Test CPU")
    _insert_supplier_price(
        db_session, supplier_id=sid, category="cpu",
        component_id=cpu_id, price=Decimal("100.00"), currency="USD",
    )
    _insert_exchange_rate(db_session, Decimal("95.5"), date.today())

    out = tmp_path / "pc.xlsx"
    report = export_components_pc(out, db=db_session)

    # rate взят из БД, не fallback.
    assert report.rate_is_fallback is False
    assert report.rate_used == Decimal("95.5000")

    wb = load_workbook(out)
    ws = wb["CPU"]
    # На листе одна строка данных — строка 4.
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    usd_col = headers["Цена min, USD"]
    rub_col = headers["Цена min, RUB"]
    supplier_col = headers["Поставщик (min)"]

    usd_cell = ws.cell(row=4, column=usd_col)
    rub_cell = ws.cell(row=4, column=rub_col)
    supplier_cell = ws.cell(row=4, column=supplier_col)

    # USD-цена — число.
    assert float(usd_cell.value) == pytest.approx(100.0)

    # RUB-цена — формула с абсолютной ссылкой на $B$1.
    assert isinstance(rub_cell.value, str)
    assert rub_cell.value.startswith("=")
    assert "$B$1" in rub_cell.value
    # Формула ссылается на ту же строку 4.
    assert rub_cell.value.endswith("*$B$1")
    # Префикс — буква колонки USD + "4".
    from openpyxl.utils import get_column_letter
    expected_prefix = f"={get_column_letter(usd_col)}4*"
    assert rub_cell.value.startswith(expected_prefix), rub_cell.value

    # Поставщик заполнен.
    assert supplier_cell.value == "USDSup"


def test_rub_priced_storage_writes_static_number(tmp_path: Path, db_session):
    """У товара только с RUB-ценой: USD пустой, RUB — статика."""
    sid = _insert_supplier(db_session, "RUBSup")
    # storages — без NOT NULL после миграции 002, минимальный INSERT.
    storage_id = db_session.execute(
        text(
            "INSERT INTO storages (model, manufacturer, storage_type) "
            "VALUES ('Storage X', 'Vendor', 'SSD') RETURNING id"
        )
    ).first().id
    db_session.commit()
    _insert_supplier_price(
        db_session, supplier_id=sid, category="storage",
        component_id=storage_id, price=Decimal("9999.99"), currency="RUB",
    )

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["Storage"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    usd_cell = ws.cell(row=4, column=headers["Цена min, USD"])
    rub_cell = ws.cell(row=4, column=headers["Цена min, RUB"])

    assert usd_cell.value is None
    # Без USD-предложения — RUB пишется как число, не формула.
    assert isinstance(rub_cell.value, (int, float))
    assert float(rub_cell.value) == pytest.approx(9999.99)


def test_array_column_serialized_with_commas(tmp_path: Path, db_session):
    """TEXT[]-колонка `supported_form_factors` сериализуется как 'ATX,mATX,ITX'."""
    case_id = _insert_case(db_session, ["ATX", "mATX", "ITX"])

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["Case"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    cell = ws.cell(row=4, column=headers["supported_form_factors"])
    assert cell.value == "ATX,mATX,ITX"
    # И id (скрытая) — это случайно вставленный id.
    id_cell = ws.cell(row=4, column=headers["id"])
    assert id_cell.value == case_id


def test_inactive_supplier_excluded_from_min_price(tmp_path: Path, db_session):
    """Неактивный поставщик не должен попадать в min-цену."""
    sid_active = _insert_supplier(db_session, "Active")
    sid_inactive = _insert_supplier(db_session, "Inactive")
    # Делаем второго неактивным.
    db_session.execute(
        text("UPDATE suppliers SET is_active = FALSE WHERE id = :sid"),
        {"sid": sid_inactive},
    )
    db_session.commit()

    cpu_id = _insert_cpu(db_session)
    # Неактивный — дешевле, но не должен быть выбран.
    _insert_supplier_price(
        db_session, supplier_id=sid_inactive, category="cpu",
        component_id=cpu_id, price=Decimal("50.00"), currency="USD",
    )
    _insert_supplier_price(
        db_session, supplier_id=sid_active, category="cpu",
        component_id=cpu_id, price=Decimal("120.00"), currency="USD",
    )

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["CPU"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    usd_cell = ws.cell(row=4, column=headers["Цена min, USD"])
    supplier_cell = ws.cell(row=4, column=headers["Поставщик (min)"])
    assert float(usd_cell.value) == pytest.approx(120.0)
    assert supplier_cell.value == "Active"


def test_out_of_stock_offers_excluded_from_min_price(tmp_path: Path, db_session):
    """Предложения с stock=0 AND transit=0 не считаются активными."""
    sid = _insert_supplier(db_session, "S")
    cpu_id = _insert_cpu(db_session)
    # Дешёвое, но без остатка.
    _insert_supplier_price(
        db_session, supplier_id=sid, category="cpu",
        component_id=cpu_id, price=Decimal("50.00"), currency="USD",
        stock=0,
    )

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["CPU"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=4, column=headers["Цена min, USD"]).value is None
    assert ws.cell(row=4, column=headers["Поставщик (min)"]).value is None


# ---------------------------------------------------------------------
# Колонки наличия (Склад / Транзит / Поставщиков)
# ---------------------------------------------------------------------

def test_availability_sums_across_active_suppliers(tmp_path: Path, db_session):
    """Склад/Транзит/Поставщиков суммируются по всем активным поставщикам
    с активными предложениями (stock>0 OR transit>0)."""
    sid1 = _insert_supplier(db_session, "S1")
    sid2 = _insert_supplier(db_session, "S2")
    cpu_id = _insert_cpu(db_session, "CPU-Avail")

    # S1: stock=7, transit=3
    _insert_supplier_price(
        db_session, supplier_id=sid1, category="cpu",
        component_id=cpu_id, price=Decimal("100.00"), currency="USD",
        stock=7, transit=3,
    )
    # S2: stock=2, transit=10
    _insert_supplier_price(
        db_session, supplier_id=sid2, category="cpu",
        component_id=cpu_id, price=Decimal("105.00"), currency="USD",
        stock=2, transit=10,
    )

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["CPU"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    # Все три колонки наличия присутствуют.
    assert "Склад, шт"       in headers
    assert "Транзит, шт"     in headers
    assert "Поставщиков, шт" in headers

    assert ws.cell(row=4, column=headers["Склад, шт"]).value       == 9   # 7+2
    assert ws.cell(row=4, column=headers["Транзит, шт"]).value     == 13  # 3+10
    assert ws.cell(row=4, column=headers["Поставщиков, шт"]).value == 2


def test_availability_excludes_inactive_supplier(tmp_path: Path, db_session):
    """Неактивный поставщик не входит ни в одну из трёх агрегатных колонок."""
    sid_active   = _insert_supplier(db_session, "Live")
    sid_inactive = _insert_supplier(db_session, "Dead")
    db_session.execute(
        text("UPDATE suppliers SET is_active = FALSE WHERE id = :sid"),
        {"sid": sid_inactive},
    )
    db_session.commit()
    cpu_id = _insert_cpu(db_session, "CPU-InactiveAvail")

    _insert_supplier_price(
        db_session, supplier_id=sid_active, category="cpu",
        component_id=cpu_id, price=Decimal("100.00"), currency="USD",
        stock=4, transit=1,
    )
    _insert_supplier_price(
        db_session, supplier_id=sid_inactive, category="cpu",
        component_id=cpu_id, price=Decimal("80.00"), currency="USD",
        stock=999, transit=999,
    )

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["CPU"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=4, column=headers["Склад, шт"]).value       == 4
    assert ws.cell(row=4, column=headers["Транзит, шт"]).value     == 1
    assert ws.cell(row=4, column=headers["Поставщиков, шт"]).value == 1


def test_availability_excludes_out_of_stock_rows(tmp_path: Path, db_session):
    """Строка с stock=0 AND transit=0 не считается активной — не входит в
    суммы и не увеличивает счётчик поставщиков."""
    sid_a = _insert_supplier(db_session, "A")
    sid_b = _insert_supplier(db_session, "B")
    cpu_id = _insert_cpu(db_session, "CPU-OutOfStock")

    # A — активен: stock=5
    _insert_supplier_price(
        db_session, supplier_id=sid_a, category="cpu",
        component_id=cpu_id, price=Decimal("100.00"), currency="USD",
        stock=5, transit=0,
    )
    # B — out-of-stock: stock=0 AND transit=0 (хотя цена есть)
    _insert_supplier_price(
        db_session, supplier_id=sid_b, category="cpu",
        component_id=cpu_id, price=Decimal("90.00"), currency="USD",
        stock=0, transit=0,
    )

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["CPU"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=4, column=headers["Склад, шт"]).value       == 5
    assert ws.cell(row=4, column=headers["Транзит, шт"]).value     == 0
    assert ws.cell(row=4, column=headers["Поставщиков, шт"]).value == 1


def test_availability_empty_when_no_active_offers(tmp_path: Path, db_session):
    """Без активных предложений все три колонки пустые (не ноль) — означает
    «у этой позиции нет ни одной активной строки в supplier_prices»."""
    _insert_cpu(db_session, "CPU-NoOffers")

    out = tmp_path / "pc.xlsx"
    export_components_pc(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["CPU"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    # Пустые ячейки — None у openpyxl.
    assert ws.cell(row=4, column=headers["Склад, шт"]).value       is None
    assert ws.cell(row=4, column=headers["Транзит, шт"]).value     is None
    assert ws.cell(row=4, column=headers["Поставщиков, шт"]).value is None


def test_availability_mfu_uses_mfu_category(tmp_path: Path, db_session):
    """Лист «МФУ» суммирует наличие по supplier_prices.category='mfu',
    а не 'printer'. Регрессия мини-этапа 2026-05-13 (orchestrator-fix)
    проверена для цен; здесь — то же для наличия."""
    sid = _insert_supplier(db_session, "MFU-AvailSup")
    pid_mfu = _insert_printer(
        db_session, sku="M-AVAIL", brand="Pantum", name="Pantum BM5100",
        category="mfu",
    )
    _insert_supplier_price(
        db_session, supplier_id=sid, category="mfu",
        component_id=pid_mfu, price=Decimal("450.00"), currency="USD",
        stock=12, transit=4,
    )
    # А вот «грязная» MFU-строка с category='printer' (легаси до 0038)
    # — должна игнорироваться на листе «МФУ».
    _insert_supplier_price(
        db_session, supplier_id=sid, category="printer",
        component_id=pid_mfu, price=Decimal("999.00"), currency="USD",
        stock=999, transit=999,
    )

    out = tmp_path / "pr.xlsx"
    export_printers_mfu(out, db=db_session)

    wb = load_workbook(out)
    ws_m = wb["МФУ"]
    headers_m = {ws_m.cell(row=3, column=c).value: c
                 for c in range(1, ws_m.max_column + 1)}
    assert ws_m.cell(row=4, column=headers_m["Склад, шт"]).value       == 12
    assert ws_m.cell(row=4, column=headers_m["Транзит, шт"]).value     == 4
    assert ws_m.cell(row=4, column=headers_m["Поставщиков, шт"]).value == 1


# ---------------------------------------------------------------------
# Тесты printers_mfu
# ---------------------------------------------------------------------

def test_printers_filter_by_category(tmp_path: Path, db_session):
    """«Принтеры» — только category='printer', «МФУ» — только 'mfu'."""
    pid_p = _insert_printer(
        db_session, sku="P-1", brand="HP", name="HP LaserJet",
        category="printer",
        attrs={"colorness": "ч/б", "max_format": "A4", "duplex": "yes"},
        ktru=["26.20.16.120-00000001"],
    )
    pid_m = _insert_printer(
        db_session, sku="M-1", brand="Pantum", name="Pantum MFU",
        category="mfu",
        attrs={"colorness": "цветной", "max_format": "A3"},
    )

    out = tmp_path / "pr.xlsx"
    report = export_printers_mfu(out, db=db_session)
    assert report.sheet_counts["Принтеры"] == 1
    assert report.sheet_counts["МФУ"] == 1

    wb = load_workbook(out)
    ws_p = wb["Принтеры"]
    headers = {ws_p.cell(row=3, column=c).value: c
               for c in range(1, ws_p.max_column + 1)}
    # На листе «Принтеры» строка одна, и это HP, id=pid_p.
    assert ws_p.cell(row=4, column=headers["id"]).value == pid_p
    assert ws_p.cell(row=4, column=headers["sku"]).value == "P-1"
    assert ws_p.cell(row=4, column=headers["brand"]).value == "HP"
    # attrs_jsonb-ключи попали в свои колонки.
    assert ws_p.cell(row=4, column=headers["colorness"]).value == "ч/б"
    assert ws_p.cell(row=4, column=headers["max_format"]).value == "A4"
    assert ws_p.cell(row=4, column=headers["duplex"]).value == "yes"
    # KTRU-массив через запятую.
    assert ws_p.cell(row=4, column=headers["ktru_codes_array"]).value == (
        "26.20.16.120-00000001"
    )

    ws_m = wb["МФУ"]
    headers_m = {ws_m.cell(row=3, column=c).value: c
                 for c in range(1, ws_m.max_column + 1)}
    assert ws_m.cell(row=4, column=headers_m["id"]).value == pid_m
    assert ws_m.cell(row=4, column=headers_m["sku"]).value == "M-1"
    assert ws_m.cell(row=4, column=headers_m["max_format"]).value == "A3"


def test_mfu_prices_use_mfu_category(tmp_path: Path, db_session):
    """Лист «МФУ» подтягивает min-цену по supplier_prices.category='mfu',
    а не 'printer'. Регрессия мини-этапа 2026-05-13: orchestrator писал
    все printers_mfu-строки в category='printer' независимо от
    printers_mfu.category, и лист «МФУ» оставался без цен.
    """
    sid = _insert_supplier(db_session, "MFUSup")
    # Принтер с ценой category='printer' — лист «Принтеры» должен подхватить.
    pid_printer = _insert_printer(
        db_session, sku="P-PRINT", brand="HP", name="HP LaserJet",
        category="printer",
    )
    _insert_supplier_price(
        db_session, supplier_id=sid, category="printer",
        component_id=pid_printer, price=Decimal("250.00"), currency="USD",
    )
    # МФУ с ценой category='mfu' — лист «МФУ» должен подхватить.
    pid_mfu = _insert_printer(
        db_session, sku="M-MFU", brand="Pantum", name="Pantum BM5100",
        category="mfu",
    )
    _insert_supplier_price(
        db_session, supplier_id=sid, category="mfu",
        component_id=pid_mfu, price=Decimal("450.00"), currency="USD",
    )

    out = tmp_path / "pr.xlsx"
    export_printers_mfu(out, db=db_session)

    wb = load_workbook(out)
    ws_p = wb["Принтеры"]
    headers_p = {ws_p.cell(row=3, column=c).value: c
                 for c in range(1, ws_p.max_column + 1)}
    assert float(
        ws_p.cell(row=4, column=headers_p["Цена min, USD"]).value
    ) == pytest.approx(250.0)
    assert ws_p.cell(row=4, column=headers_p["Поставщик (min)"]).value == "MFUSup"

    ws_m = wb["МФУ"]
    headers_m = {ws_m.cell(row=3, column=c).value: c
                 for c in range(1, ws_m.max_column + 1)}
    assert float(
        ws_m.cell(row=4, column=headers_m["Цена min, USD"]).value
    ) == pytest.approx(450.0)
    assert ws_m.cell(row=4, column=headers_m["Поставщик (min)"]).value == "MFUSup"


def test_mfu_price_with_printer_category_is_ignored(tmp_path: Path, db_session):
    """Если в supplier_prices ошибочно осталась MFU-строка с category='printer'
    (старые данные до миграции 0038), лист «МФУ» её НЕ показывает —
    защищает от регрессии, если backfill не был накачен."""
    sid = _insert_supplier(db_session, "LegacySup")
    pid_mfu = _insert_printer(
        db_session, sku="M-LEGACY", brand="HP", name="HP MFP",
        category="mfu",
    )
    # Симулируем баг: MFU-строка лежит с category='printer'.
    _insert_supplier_price(
        db_session, supplier_id=sid, category="printer",
        component_id=pid_mfu, price=Decimal("300.00"), currency="USD",
    )

    out = tmp_path / "pr.xlsx"
    export_printers_mfu(out, db=db_session)

    wb = load_workbook(out)
    ws_m = wb["МФУ"]
    headers_m = {ws_m.cell(row=3, column=c).value: c
                 for c in range(1, ws_m.max_column + 1)}
    assert ws_m.cell(row=4, column=headers_m["Цена min, USD"]).value is None
    assert ws_m.cell(row=4, column=headers_m["Поставщик (min)"]).value is None


def test_printer_dimension_attrs_written(tmp_path: Path, db_session):
    """Опциональные ключи габаритов (weight_kg/box_*) попадают в Excel."""
    _insert_printer(
        db_session, sku="P-DIM", brand="HP", name="HP Big",
        category="printer",
        attrs={
            "weight_kg": 14.2, "box_width_cm": 50.0,
            "box_height_cm": 40.0, "box_depth_cm": 35.5,
        },
    )

    out = tmp_path / "pr.xlsx"
    export_printers_mfu(out, db=db_session)

    wb = load_workbook(out)
    ws = wb["Принтеры"]
    headers = {ws.cell(row=3, column=c).value: c
               for c in range(1, ws.max_column + 1)}
    assert float(ws.cell(row=4, column=headers["weight_kg"]).value) == 14.2
    assert float(ws.cell(row=4, column=headers["box_width_cm"]).value) == 50.0


# ---------------------------------------------------------------------
# default_filename
# ---------------------------------------------------------------------

def test_default_filename_pc():
    fn = default_filename("pc", today=date(2026, 5, 14))
    assert fn == "Комплектующие_ПК_2026-05-14.xlsx"


def test_default_filename_printers():
    fn = default_filename("printers", today=date(2026, 5, 14))
    assert fn == "Печатная_техника_2026-05-14.xlsx"


def test_default_filename_invalid():
    with pytest.raises(ValueError):
        default_filename("xxx", today=date(2026, 5, 14))


# HTTP-тесты эндпоинта /databases/catalog-excel/download/{target} живут в
# tests/test_portal/test_catalog_excel.py — там доступны фикстуры
# admin_portal_client / manager_portal_client / portal_client.
