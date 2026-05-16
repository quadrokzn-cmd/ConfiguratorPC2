"""HTTP-тесты роута GET /auctions/excel (Backlog #12, Фаза 4 плана
2026-05-13-auctions-excel-export.md).

Юнит-тесты сервиса лежат в tests/test_auctions/test_excel_export.py —
здесь только HTTP-слой: коды ответов, MIME, имя файла, audit_log,
доступы (anonymous / no-perm / viewer / admin) и наличие кнопки
«Скачать Excel» на /auctions.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import text

from tests.test_portal.auctions_fixtures import (
    auctions_no_perm,    # noqa: F401 — pytest fixture
    auctions_viewer,     # noqa: F401
    insert_match,
    insert_printer_mfu,
    insert_tender,
    insert_tender_item,
    login_as,
)


_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ---------------------------------------------------------------
# Доступы
# ---------------------------------------------------------------

def test_anonymous_redirect_to_login(portal_client):
    r = portal_client.get("/auctions/excel")
    assert r.status_code in (302, 303), r.status_code
    assert "/login" in r.headers.get("location", "")


def test_manager_without_auctions_perm_403(portal_client, auctions_no_perm):
    login_as(portal_client, auctions_no_perm)
    r = portal_client.get("/auctions/excel")
    assert r.status_code == 403, r.status_code


def test_auctions_viewer_200(portal_client, auctions_viewer):
    """Менеджер с правом 'auctions' скачивает Excel."""
    login_as(portal_client, auctions_viewer)
    r = portal_client.get("/auctions/excel")
    assert r.status_code == 200, r.text[:200]
    assert r.headers["content-type"] == _XLSX_MIME
    cd = r.headers["content-disposition"].lower()
    assert "filename" in cd

    wb = load_workbook(BytesIO(r.content))
    assert wb.active.title == "Аукционы"


def test_admin_200(admin_portal_client):
    r = admin_portal_client.get("/auctions/excel")
    assert r.status_code == 200, r.text[:200]


# ---------------------------------------------------------------
# Содержимое файла под фильтрами через query string
# ---------------------------------------------------------------

def test_query_filter_status_passes_through(
    portal_client, auctions_viewer, db_session,
):
    """status=skipped в query → только skipped-лоты в файле."""
    insert_tender(db_session, reg_number="ut-01", status="new")
    insert_tender_item(db_session, tender_id="ut-01")
    insert_tender(db_session, reg_number="ut-02", status="skipped")
    insert_tender_item(db_session, tender_id="ut-02")

    login_as(portal_client, auctions_viewer)
    r = portal_client.get("/auctions/excel?status=skipped")
    assert r.status_code == 200

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    # Шапка на строке 3, данные с 4-й. Ожидаем 1 data-строку = ut-02.
    reg_col_letter = "B"  # № извещения — вторая колонка
    values = [
        ws[f"{reg_col_letter}{row}"].value for row in range(4, 10)
    ]
    non_empty = [v for v in values if v]
    assert non_empty == ["ut-02"]


# ---------------------------------------------------------------
# Audit log
# ---------------------------------------------------------------

def test_audit_log_written_on_download(
    portal_client, auctions_viewer, db_session,
):
    """Каждое успешное скачивание → запись auctions_excel_export."""
    insert_tender(db_session, reg_number="au-01")
    insert_tender_item(db_session, tender_id="au-01")

    login_as(portal_client, auctions_viewer)
    r = portal_client.get("/auctions/excel")
    assert r.status_code == 200

    row = db_session.execute(text(
        "SELECT action, target_type, user_login, payload "
        "FROM audit_log "
        "WHERE action = 'auctions_excel_export' "
        "ORDER BY id DESC LIMIT 1"
    )).first()
    assert row is not None, "audit_log запись auctions_excel_export не найдена"
    assert row.action == "auctions_excel_export"
    assert row.target_type == "auctions_excel"
    assert row.user_login == auctions_viewer["login"]

    payload = row.payload
    assert payload["rows_count"] >= 1
    assert "filter_summary" in payload
    assert "rate_used" in payload
    assert payload["rate_fallback"] is True  # exchange_rates пуст в тестах
    assert payload["cap_reached"] is False
    # filter_summary содержит ключи всех 7 фильтров inbox
    summary = payload["filter_summary"]
    for key in (
        "statuses", "nmck_min", "nmck_max", "search",
        "urgent_only", "print_only", "include_excluded_regions",
    ):
        assert key in summary, f"filter_summary должен содержать '{key}'"


# ---------------------------------------------------------------
# Кнопка на странице /auctions
# ---------------------------------------------------------------

def test_inbox_page_has_download_button(portal_client, auctions_viewer):
    """На /auctions есть кнопка «Скачать Excel» с правильной ссылкой."""
    login_as(portal_client, auctions_viewer)
    r = portal_client.get("/auctions")
    assert r.status_code == 200
    assert 'data-testid="download-auctions-excel"' in r.text
    # Ссылка ведёт на /auctions/excel (без query при пустых фильтрах)
    assert 'href="/auctions/excel"' in r.text


def test_inbox_page_button_preserves_query(portal_client, auctions_viewer):
    """Кнопка «Скачать Excel» наследует текущие фильтры из URL."""
    login_as(portal_client, auctions_viewer)
    r = portal_client.get("/auctions?status=new&q=якут")
    assert r.status_code == 200
    # Ожидаем, что href содержит ?status=new&q=якут (либо URL-кодированно)
    assert "/auctions/excel?status=new" in r.text or \
           "/auctions/excel?status=new&amp;q=%D1%8F%D0%BA%D1%83%D1%82" in r.text
