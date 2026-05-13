# HTTP-тесты эндпоинта /databases/catalog-excel/download/{target}
# (Фаза 2 плана plans/2026-05-13-catalog-excel-export-import.md).
#
# Юнит-тесты сервиса лежат в tests/test_catalog/test_excel_export.py —
# здесь только то, что относится к HTTP-слою: коды ответов, MIME-тип,
# запись в audit_log, доступы (admin/manager/anonymous).

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import text


_EXPECTED_PC_SHEETS = [
    "CPU", "Motherboard", "RAM", "GPU",
    "Storage", "Case", "PSU", "Cooler",
]

_EXPECTED_PRINTERS_SHEETS = ["Принтеры", "МФУ"]


def test_download_pc_returns_xlsx(admin_portal_client):
    """GET /databases/catalog-excel/download/pc → 200, xlsx, 8 листов."""
    r = admin_portal_client.get("/databases/catalog-excel/download/pc")
    assert r.status_code == 200, r.text[:200]
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename" in r.headers["content-disposition"].lower()

    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == _EXPECTED_PC_SHEETS


def test_download_printers_returns_xlsx(admin_portal_client):
    r = admin_portal_client.get("/databases/catalog-excel/download/printers")
    assert r.status_code == 200, r.text[:200]
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == _EXPECTED_PRINTERS_SHEETS


def test_download_invalid_target_404(admin_portal_client):
    r = admin_portal_client.get("/databases/catalog-excel/download/foo")
    assert r.status_code == 404


def test_download_manager_403(manager_portal_client):
    """Менеджер не должен скачивать — только админ."""
    r = manager_portal_client.get("/databases/catalog-excel/download/pc")
    assert r.status_code == 403


def test_download_anonymous_redirect(portal_client):
    """Без логина — редирект на /login."""
    r = portal_client.get("/databases/catalog-excel/download/pc")
    assert r.status_code in (302, 303), r.status_code
    assert "/login" in r.headers.get("location", "")


def test_download_writes_audit_log(admin_portal_client, db_session):
    """После скачивания в audit_log появляется запись catalog_excel_export."""
    r = admin_portal_client.get("/databases/catalog-excel/download/pc")
    assert r.status_code == 200

    row = db_session.execute(
        text(
            "SELECT action, target_type, target_id, payload "
            "FROM audit_log "
            "WHERE action = 'catalog_excel_export' "
            "ORDER BY id DESC LIMIT 1"
        )
    ).first()
    assert row is not None, "audit_log: запись catalog_excel_export не найдена"
    assert row.action == "catalog_excel_export"
    assert row.target_type == "catalog_excel"
    assert row.target_id == "pc"
    # payload — JSONB, psycopg2 разворачивает в dict.
    assert row.payload["target"] == "pc"
    assert "rows_count" in row.payload
    assert "sheet_counts" in row.payload


# ---------------------------------------------------------------------------
# UI-страница /databases/catalog-excel (Фаза 4)
# ---------------------------------------------------------------------------


def test_page_admin_200(admin_portal_client):
    """GET /databases/catalog-excel админу отдаёт 200 со страницей."""
    r = admin_portal_client.get("/databases/catalog-excel")
    assert r.status_code == 200, r.text[:200]
    body = r.text
    # На странице есть обе карточки и обе кнопки скачивания.
    assert "Комплектующие ПК" in body
    assert "Печатная техника" in body
    assert "/databases/catalog-excel/download/pc" in body
    assert "/databases/catalog-excel/download/printers" in body
    # data-testid карточек — стабильные якоря для будущих UI-тестов.
    assert 'data-testid="catalog-card-pc"' in body
    assert 'data-testid="catalog-card-printers"' in body


def test_page_manager_403(manager_portal_client):
    """Менеджер не должен видеть страницу — только админ."""
    r = manager_portal_client.get("/databases/catalog-excel")
    assert r.status_code == 403


def test_page_anonymous_redirect(portal_client):
    r = portal_client.get("/databases/catalog-excel")
    assert r.status_code in (302, 303), r.status_code
    assert "/login" in r.headers.get("location", "")


def test_page_history_shows_recent_export(admin_portal_client, db_session):
    """После скачивания запись audit_log появляется на странице как «Экспорт»."""
    # Сначала вызываем download — он пишет audit_log row.
    r1 = admin_portal_client.get("/databases/catalog-excel/download/pc")
    assert r1.status_code == 200

    r2 = admin_portal_client.get("/databases/catalog-excel")
    assert r2.status_code == 200
    body = r2.text
    # Бейдж «Экспорт» должен присутствовать в строке истории для pc-карточки.
    assert "Экспорт" in body
    # data-testid конкретной audit-строки (id из БД).
    row = db_session.execute(
        text(
            "SELECT id FROM audit_log "
            "WHERE action = 'catalog_excel_export' AND target_id = 'pc' "
            "ORDER BY id DESC LIMIT 1"
        )
    ).first()
    assert row is not None
    assert f'data-testid="audit-row-pc-{row.id}"' in body


def test_page_history_separated_by_kind(admin_portal_client, db_session):
    """История pc-карточки и printers-карточки не перемешиваются.

    Делаем по одному download для каждого target и проверяем, что в HTML
    audit-row с правильным target_id рендерится под своей карточкой.
    """
    admin_portal_client.get("/databases/catalog-excel/download/pc")
    admin_portal_client.get("/databases/catalog-excel/download/printers")

    rows = db_session.execute(
        text(
            "SELECT id, target_id FROM audit_log "
            "WHERE action = 'catalog_excel_export' "
            "ORDER BY id DESC LIMIT 2"
        )
    ).all()
    by_kind = {r.target_id: r.id for r in rows}
    assert {"pc", "printers"}.issubset(by_kind.keys())

    body = admin_portal_client.get("/databases/catalog-excel").text
    assert f'data-testid="audit-row-pc-{by_kind["pc"]}"' in body
    assert f'data-testid="audit-row-printers-{by_kind["printers"]}"' in body
    # Перекрёстных совпадений не должно быть.
    assert f'data-testid="audit-row-pc-{by_kind["printers"]}"' not in body
    assert f'data-testid="audit-row-printers-{by_kind["pc"]}"' not in body
