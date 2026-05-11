# Фикстуры для тестов портала (этап 9Б.1).
#
# DB-инфраструктура (db_engine, db_session, миграции 001..018) живёт
# в корневом `tests/conftest.py`. Здесь только portal-специфичные
# фикстуры: чистка таблиц перед каждым тестом, создание пользователей
# и TestClient портала.

from __future__ import annotations

import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text


@pytest.fixture(autouse=True)
def _clean_tables(db_engine):
    """Перед каждым тестом — пустые таблицы, чтобы предыдущий тест
    не мешал.

    Этап 9a: добавлены аукционные «таблицы данных» (matches, tender_items,
    tender_status, tenders, printers_mfu). Lookup-таблицы (settings,
    excluded_regions, ktru_watchlist) НЕ TRUNCATE-аются — у них seed
    из миграции 030/034, и тесты на чтение полагаются на эти значения.
    После тестов, которые мутируют lookup'ы, фикстура восстанавливает
    дефолты UPSERT'ом.
    """
    with db_engine.begin() as conn:
        conn.execute(text(
            "TRUNCATE TABLE audit_log, sent_emails, specification_items, queries, "
            "projects, daily_budget_log, users, api_usage_log, exchange_rates, "
            "matches, tender_status, tender_items, tenders, printers_mfu "
            "RESTART IDENTITY CASCADE"
        ))
        # Возвращаем seed-значения settings (тесты могут менять их).
        conn.execute(text(
            "INSERT INTO settings (key, value) VALUES "
            "  ('margin_threshold_pct', '15'), "
            "  ('nmck_min_rub', '30000'), "
            "  ('max_price_per_unit_rub', '300000'), "
            "  ('contract_reminder_days', '3'), "
            "  ('deadline_alert_hours', '24'), "
            "  ('auctions_ingest_enabled', 'true') "
            "ON CONFLICT (key) DO UPDATE "
            "  SET value = EXCLUDED.value, updated_at = NOW()"
        ))
        # excluded_regions: возвращаем все 7 в excluded=TRUE (дефолт).
        conn.execute(text(
            "UPDATE excluded_regions SET excluded = TRUE"
        ))
        # ktru_watchlist: только 2 зонтика активны (как в миграции 030).
        conn.execute(text(
            "UPDATE ktru_watchlist "
            "  SET is_active = (code IN "
            "    ('26.20.18.000-00000001', '26.20.16.120-00000001'))"
        ))
    yield


# ---- Создание пользователей --------------------------------------------


def _create_user(
    session,
    *,
    login: str,
    password: str,
    role: str,
    name: str,
    permissions: dict | None = None,
) -> int:
    from shared.auth import hash_password
    perms = permissions if permissions is not None else (
        {} if role == "admin" else {"configurator": True}
    )
    row = session.execute(
        text(
            "INSERT INTO users (login, password_hash, role, name, permissions) "
            "VALUES (:l, :p, :r, :n, CAST(:perms AS JSONB)) RETURNING id"
        ),
        {
            "l":     login,
            "p":     hash_password(password),
            "r":     role,
            "n":     name,
            "perms": json.dumps(perms, ensure_ascii=False),
        },
    ).first()
    session.commit()
    return int(row.id)


@pytest.fixture()
def admin_user(db_session):
    uid = _create_user(db_session, login="admin", password="admin-pass",
                       role="admin", name="Администратор")
    return {"id": uid, "login": "admin", "password": "admin-pass"}


@pytest.fixture()
def manager_user(db_session):
    """Менеджер с дефолтными правами (configurator: True)."""
    uid = _create_user(db_session, login="manager1", password="manager-pass",
                       role="manager", name="Менеджер 1")
    return {"id": uid, "login": "manager1", "password": "manager-pass"}


@pytest.fixture()
def manager_user_no_perms(db_session):
    """Менеджер без единого разрешённого модуля."""
    uid = _create_user(
        db_session, login="manager_empty", password="manager-pass",
        role="manager", name="Менеджер без прав", permissions={},
    )
    return {"id": uid, "login": "manager_empty", "password": "manager-pass"}


# ---- TestClient'ы портала ---------------------------------------------

@pytest.fixture()
def portal_client():
    """TestClient портала без залогиненной сессии."""
    from portal.main import app
    with TestClient(app, follow_redirects=False) as c:
        yield c


def extract_csrf(html: str) -> str:
    import re
    m = re.search(r'name="csrf_token" value="([^"]+)"', html)
    assert m, "csrf_token не найден на странице"
    return m.group(1)


def _login_via_portal(client: TestClient, login: str, password: str) -> None:
    r = client.get("/login")
    assert r.status_code == 200, r.status_code
    token = extract_csrf(r.text)
    r = client.post(
        "/login",
        data={"login": login, "password": password, "csrf_token": token},
    )
    assert r.status_code in (302, 303), f"login failed: {r.status_code} {r.text[:200]}"


# Алиас под старое имя из tests/test_web/conftest.py — _login.
# После UI-4 (Путь B): app_client/portal_client — это portal-клиент,
# логин выполняется напрямую в нём (без копирования cookie).
_login = _login_via_portal


@pytest.fixture()
def admin_portal_client(portal_client, admin_user):
    _login_via_portal(portal_client, admin_user["login"], admin_user["password"])
    return portal_client


@pytest.fixture()
def manager_portal_client(portal_client, manager_user):
    _login_via_portal(portal_client, manager_user["login"], manager_user["password"])
    return portal_client


# ---- Утилиты для тестов конфигуратора (бывший tests/test_web/conftest.py) ----


def parse_query_submit_redirect(location: str) -> tuple[int, int]:
    """POST /configurator/query редиректит на /configurator/project/{pid}?highlight={qid}.
    Возвращает (project_id, query_id). UI-4 (Путь B): URL переехал на /configurator/."""
    from urllib.parse import urlsplit, parse_qs
    parts = urlsplit(location)
    pid = int(parts.path.rsplit("/", 1)[1])
    qid = int(parse_qs(parts.query).get("highlight", ["0"])[0])
    return pid, qid


def qid_from_submit_redirect(location: str) -> int:
    """Только query_id — для тестов, которым не нужен project_id."""
    return parse_query_submit_redirect(location)[1]


@pytest.fixture()
def mock_process_query(monkeypatch):
    """Мокает process_query из portal.routers.configurator.{main,projects}.

    UI-4 (Путь B, 2026-05-11): main_router и project_router переехали из
    app/routers в portal/routers/configurator. Импорты обновлены.

    По умолчанию возвращает «успешный» FinalResponse с одним Intel-вариантом.
    Тест может подменить возвращаемое значение через mock.return_value.
    """
    from unittest.mock import MagicMock
    from portal.routers.configurator import main as configurator_main
    from portal.routers.configurator import projects as configurator_projects
    from portal.services.configurator.engine.schema import (
        BuildRequest, BuildResult, ComponentChoice, CPURequirements,
        GPURequirements, RAMRequirements, StorageRequirements,
        SupplierOffer, Variant,
    )
    from portal.services.configurator.nlu.schema import FinalResponse, ParsedRequest

    fake_variant = Variant(
        manufacturer="Intel",
        components=[
            ComponentChoice(
                category="cpu",
                component_id=1,
                model="Intel Core i5-12400F",
                sku="BX8071512400F",
                manufacturer="Intel",
                chosen=SupplierOffer(
                    supplier="Поставщик А", price_usd=180, price_rub=16200, stock=10,
                ),
            ),
            ComponentChoice(
                category="ram",
                component_id=2,
                model="Kingston 16GB DDR4",
                sku=None,
                manufacturer="Kingston",
                chosen=SupplierOffer(
                    supplier="Поставщик Б", price_usd=40, price_rub=3600, stock=5,
                ),
            ),
        ],
        total_usd=220,
        total_rub=19800,
    )
    fake_result = BuildResult(
        status="ok",
        variants=[fake_variant],
        refusal_reason=None,
        usd_rub_rate=90.0,
        fx_source="fallback",
    )
    fake_req = BuildRequest()
    fake_parsed = ParsedRequest(is_empty=False, purpose="office", budget_usd=300)

    default_resp = FinalResponse(
        kind="ok",
        interpretation="Офисный ПК до $300.",
        formatted_text="[Фиктивный форматированный ответ]",
        build_request=fake_req,
        build_result=fake_result,
        parsed=fake_parsed,
        resolved=[],
        warnings=[],
        cost_usd=0.0,
    )

    mock = MagicMock(return_value=default_resp)
    monkeypatch.setattr(configurator_main, "process_query", mock)
    monkeypatch.setattr(configurator_projects, "process_query", mock)
    return mock


# UI-4: алиасы fixture'ов под старые имена из tests/test_web/conftest.py —
# чтобы перенесённые тесты test_configurator_*.py продолжали работать без
# массового переписывания. app_client/admin_client/manager_client теперь
# используют portal-клиент (конфигуратор переехал в portal/configurator/*).

@pytest.fixture()
def app_client(portal_client):
    return portal_client


@pytest.fixture()
def admin_client(admin_portal_client):
    return admin_portal_client


@pytest.fixture()
def manager_client(manager_portal_client):
    return manager_portal_client


@pytest.fixture()
def manager2_user(db_session):
    """Второй менеджер с дефолтными правами (для тестов «чужой проект»)."""
    uid = _create_user(db_session, login="manager2", password="manager-pass",
                       role="manager", name="Менеджер 2")
    return {"id": uid, "login": "manager2", "password": "manager-pass"}


@pytest.fixture()
def manager_no_perms(db_session):
    """Алиас manager_user_no_perms (имя из test_web/conftest.py)."""
    uid = _create_user(
        db_session, login="manager_empty", password="manager-pass",
        role="manager", name="Без доступа",
        permissions={},
    )
    return {"id": uid, "login": "manager_empty", "password": "manager-pass"}


# ---- Минимальный Merlion-XLSX для тестов /admin/price-uploads --------
#
# Дублирует основу test_price_loaders/conftest.py:make_merlion_xlsx,
# но без сложной API — здесь нужны простые сценарии. Вынесено в общий
# portal-conftest, чтобы тест test_admin_price_uploads.py не лез в чужой
# conftest.

@pytest.fixture()
def make_merlion_xlsx(tmp_path):
    """Фабрика минимального Merlion-XLSM. Параметры см. тест."""
    from openpyxl import Workbook

    def _make(rows: list[dict], *, name: str = "Прайслист_Мерлион.xlsm") -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Price List"
        ws.cell(row=1, column=1, value="Шапка Merlion")
        headers = [
            "Группа 1", "Группа 2", "Группа 3", "Бренд", "Номер", "Ext код",
            "Код производителя", "Наименование", "Валюта", "Цена",
            "Цена(руб)", "Доступно", "Ожидаемый приход", "На складе поставщика",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=11, column=col_idx, value=h)
        for i, r in enumerate(rows, start=12):
            ws.cell(row=i, column=1,  value=r.get("g1"))
            ws.cell(row=i, column=2,  value=r.get("g2"))
            ws.cell(row=i, column=3,  value=r.get("g3"))
            ws.cell(row=i, column=4,  value=r.get("brand"))
            ws.cell(row=i, column=5,  value=r.get("number"))
            ws.cell(row=i, column=7,  value=r.get("mpn"))
            ws.cell(row=i, column=8,  value=r.get("name"))
            ws.cell(row=i, column=10, value=r.get("price_usd"))
            ws.cell(row=i, column=11, value=r.get("price_rub"))
            ws.cell(row=i, column=12, value=r.get("stock"))
        path = Path(tmp_path) / name
        wb.save(path)
        return str(path)

    return _make
