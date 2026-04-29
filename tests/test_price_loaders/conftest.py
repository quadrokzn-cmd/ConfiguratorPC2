# Фикстуры для тестов пакета price_loaders.
#
# DB-инфраструктура (db_engine, db_session, миграции 001..018) — в
# корневом `tests/conftest.py`. Здесь только чистка таблиц компонентов/
# поставщиков/прайсов перед каждым тестом и фабрики Excel-моков.

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import text


@pytest.fixture(autouse=True)
def _clean_component_tables(db_engine):
    """Перед каждым тестом — пустые таблицы компонентов/цен/поставщиков."""
    with db_engine.begin() as conn:
        conn.execute(text(
            "TRUNCATE TABLE unmapped_supplier_items, price_uploads, "
            "supplier_prices, suppliers, "
            "cpus, motherboards, rams, gpus, storages, cases, psus, coolers, "
            "component_field_sources "
            "RESTART IDENTITY CASCADE"
        ))
    yield


# ---- Хелперы для построения Excel-моков -----------------------------------


def _save_workbook(wb: Workbook, tmp_path: Path, name: str) -> str:
    """Сохраняет workbook во временный файл и возвращает его путь."""
    path = tmp_path / name
    wb.save(path)
    return str(path)


@pytest.fixture()
def make_merlion_xlsx(tmp_path: Path):
    """Фабрика: принимает список строк (dict с колонками Merlion) и
    возвращает путь к .xlsm с правильной структурой:
      - лист «Price List»;
      - строка 11 = заголовки;
      - строка 12+ = данные.
    """
    def _make(rows: list[dict], *, name: str = "Прайслист_Мерлион_Москва.xlsm") -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Price List"

        # Строки 1..10 — служебные (пустые или любой мусор). Их парсер не читает.
        ws.cell(row=1, column=1, value="Служебная шапка Merlion")

        # Строка 11 — заголовки. Порядок важен, индексы колонок в парсере хардкодятся.
        headers = [
            "Группа 1", "Группа 2", "Группа 3",   # A B C
            "Бренд",                               # D
            "Номер",                               # E (supplier_sku)
            "Ext код",                             # F (пропуск)
            "Код производителя",                   # G (MPN)
            "Наименование",                        # H
            "Валюта",                              # I (пропуск)
            "Цена",                                # J (USD)
            "Цена(руб)",                           # K (RUB)
            "Доступно",                            # L
            "Ожидаемый приход",                    # M
            "На складе поставщика",                # N
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=11, column=col_idx, value=h)

        # Строки данных.
        for i, r in enumerate(rows, start=12):
            ws.cell(row=i, column=1,  value=r.get("g1"))
            ws.cell(row=i, column=2,  value=r.get("g2"))
            ws.cell(row=i, column=3,  value=r.get("g3"))
            ws.cell(row=i, column=4,  value=r.get("brand"))
            ws.cell(row=i, column=5,  value=r.get("number"))
            ws.cell(row=i, column=7,  value=r.get("mpn"))
            ws.cell(row=i, column=8,  value=r.get("name"))
            ws.cell(row=i, column=10, value=r.get("price_usd"))
            ws.cell(row=i, column=11, value=r.get("price_rub"))
            ws.cell(row=i, column=12, value=r.get("stock"))
            ws.cell(row=i, column=13, value=r.get("transit_1"))
            ws.cell(row=i, column=14, value=r.get("transit_2"))

        return _save_workbook(wb, tmp_path, name)

    return _make


@pytest.fixture()
def make_treolan_xlsx(tmp_path: Path):
    """Фабрика: принимает последовательность «узлов» (строк данных и
    строк-разделителей категорий) и возвращает путь к .xlsx.

    Узел может быть:
      {"category": "Комплектующие->Процессоры"}  — строка-разделитель;
      {                                            — строка товара:
          "article":  "SRMBG",
          "name":     "...",
          "brand":    "Intel",
          "stock":    5, "transit_1":0, "transit_2":0,
          "price_usd": 200, "price_rub": 18000,
          "gtin":     "5032037260466",
      }
    """
    def _make(items: list[dict], *, name: str = "23_04_2026_catalog__1_.xlsx") -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог"

        # Строки 1-2 — служебные.
        ws.cell(row=1, column=1, value="Каталог Treolan")

        # Строка 3 — заголовки.
        headers = [
            "Артикул",        # A
            "Наименование",   # B
            "Производитель",  # C
            "Склад",          # D
            "Транзит",        # E
            "Б.Тр.",          # F
            "Цена*",          # G (USD)
            "Цена руб.**",    # H (RUB)
            "--",             # I (пропуск)
            "Код GTIN",       # J
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=h)

        # Данные начиная со строки 4.
        row_num = 4
        for it in items:
            if "category" in it:
                # Строка-разделитель: заполнена только колонка A.
                ws.cell(row=row_num, column=1, value=it["category"])
            else:
                ws.cell(row=row_num, column=1, value=it.get("article"))
                ws.cell(row=row_num, column=2, value=it.get("name"))
                ws.cell(row=row_num, column=3, value=it.get("brand"))
                ws.cell(row=row_num, column=4, value=it.get("stock"))
                ws.cell(row=row_num, column=5, value=it.get("transit_1"))
                ws.cell(row=row_num, column=6, value=it.get("transit_2"))
                ws.cell(row=row_num, column=7, value=it.get("price_usd"))
                ws.cell(row=row_num, column=8, value=it.get("price_rub"))
                ws.cell(row=row_num, column=10, value=it.get("gtin"))
            row_num += 1

        return _save_workbook(wb, tmp_path, name)

    return _make


@pytest.fixture()
def make_ocs_xlsx(tmp_path: Path):
    """Фабрика минимального OCS-прайса: лист «Наличие и цены» с одной
    строкой заголовков и N строк данных. Поддерживает опциональную
    колонку EAN128, если она нужна тесту.
    """
    def _make(rows: list[dict], *, name: str = "OCS_price.xlsx",
              with_ean: bool = True) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Наличие и цены"

        headers = [
            "Категория A", "Категория B", "Вид оборудования", "Производитель",
            "Номенклатурный номер",    # E (supplier_sku)
            "Доп код",                 # F
            "Каталожный номер",        # G (MPN)
            "Наименование",            # H
            "Цена",                    # I
            "Валюта",                  # J
            "Скидка",                  # K
            "Остаток",                 # L
            "пусто", "пусто", "пусто", "пусто", "пусто",
            "Транзит",                 # R (col 17 0-based)
        ]
        if with_ean:
            headers.append("EAN128")   # col 18

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)

        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1,  value=r.get("cat_a"))
            ws.cell(row=i, column=2,  value=r.get("cat_b"))
            ws.cell(row=i, column=3,  value=r.get("kind_c"))
            ws.cell(row=i, column=4,  value=r.get("maker"))
            ws.cell(row=i, column=5,  value=r.get("supplier_sku"))
            ws.cell(row=i, column=7,  value=r.get("mpn"))
            ws.cell(row=i, column=8,  value=r.get("name"))
            ws.cell(row=i, column=9,  value=r.get("price"))
            ws.cell(row=i, column=10, value=r.get("currency"))
            ws.cell(row=i, column=12, value=r.get("stock"))
            ws.cell(row=i, column=18, value=r.get("transit"))
            if with_ean:
                ws.cell(row=i, column=19, value=r.get("ean"))

        return _save_workbook(wb, tmp_path, name)

    return _make
