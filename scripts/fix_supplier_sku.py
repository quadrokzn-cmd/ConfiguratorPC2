# Одноразовая миграция данных: правильные значения supplier_sku
# в supplier_prices из OCS-прайса.
#
# Зачем это нужно. До фикса price_loader писал в supplier_prices.supplier_sku
# каталожный номер производителя (колонка G Excel), т.е. то же значение,
# что и в <таблица>.sku. Из-за этого в UI менеджер видел дубликат «Артикул»
# и «№ у поставщика». Правильное значение supplier_sku — номенклатурный
# номер OCS (колонка E), например «1000659869».
#
# Скрипт не трогает цены, остатки, наличие компонентов и их каталожные
# номера — только supplier_prices.supplier_sku. Повторный запуск безопасен
# (идемпотентен): если значение уже верное, UPDATE просто ничего не поменяет.
#
# Использование:
#   python scripts/fix_supplier_sku.py <путь-к-прайсу-ocs.xlsx>
#   python scripts/fix_supplier_sku.py <путь> --dry-run

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8")
        except Exception:
            pass

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import SessionLocal
from portal.services.configurator.price_loader import (
    _CATEGORY_MAP,
    _COL_CAT_B,
    _COL_KIND_C,
    _COL_SKU,
    _COL_SUPPLIER_SKU,
    _cell,
    _resolve_mapping,
)


def _iter_price_rows(filepath: str):
    """Перебирает строки листа «Наличие и цены» как (row_idx, row)."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = "Наличие и цены"
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Лист «{sheet_name}» не найден в файле {filepath}. "
            f"Доступные листы: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        yield row_idx, row


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Перезаливает supplier_prices.supplier_sku из прайса OCS "
                    "(колонка E), не меняя цены и остатки.",
    )
    ap.add_argument("filepath", help="путь к .xlsx-прайсу OCS")
    ap.add_argument("--dry-run", action="store_true",
                    help="только показать, сколько строк будет обновлено, без записи")
    args = ap.parse_args()

    if not os.path.isfile(args.filepath):
        print(f"Файл не найден: {args.filepath}", file=sys.stderr)
        return 2

    session = SessionLocal()
    try:
        row = session.execute(
            text("SELECT id FROM suppliers WHERE name='OCS' LIMIT 1")
        ).fetchone()
        if row is None:
            print("Поставщик OCS не найден в БД. Сначала загрузите прайс через "
                  "`python scripts/load_price.py`.", file=sys.stderr)
            return 3
        supplier_id = int(row.id)

        stats = {"considered": 0, "matched": 0, "updated": 0, "skipped": 0, "absent": 0}

        for row_idx, row in _iter_price_rows(args.filepath):
            cat_b  = str(_cell(row, _COL_CAT_B) or "").strip()
            kind_c = str(_cell(row, _COL_KIND_C) or "").strip()
            supplier_sku = str(_cell(row, _COL_SUPPLIER_SKU) or "").strip()  # E
            sku          = str(_cell(row, _COL_SKU) or "").strip()           # G

            mapping = _resolve_mapping(cat_b, kind_c)
            if mapping is None or not sku:
                continue
            stats["considered"] += 1
            table, category = mapping
            supplier_sku_or_none = supplier_sku or None

            # Находим компонент в его таблице по каталожному sku.
            comp = session.execute(
                text(f"SELECT id FROM {table} WHERE sku = :sku LIMIT 1"),
                {"sku": sku},
            ).fetchone()
            if comp is None:
                stats["absent"] += 1
                continue
            component_id = int(comp.id)
            stats["matched"] += 1

            if args.dry_run:
                # Считаем, сколько _реально_ изменится значение
                current = session.execute(
                    text(
                        "SELECT supplier_sku FROM supplier_prices "
                        "WHERE supplier_id = :sid AND category = :cat "
                        "AND component_id = :cid"
                    ),
                    {"sid": supplier_id, "cat": category, "cid": component_id},
                ).fetchone()
                if current is None:
                    stats["skipped"] += 1
                elif (current.supplier_sku or "") != (supplier_sku_or_none or ""):
                    stats["updated"] += 1
                continue

            result = session.execute(
                text(
                    "UPDATE supplier_prices "
                    "SET supplier_sku = :sup_sku, updated_at = NOW() "
                    "WHERE supplier_id = :sid "
                    "  AND category = :cat "
                    "  AND component_id = :cid "
                    "  AND (supplier_sku IS DISTINCT FROM :sup_sku)"
                ),
                {
                    "sup_sku": supplier_sku_or_none,
                    "sid":     supplier_id,
                    "cat":     category,
                    "cid":     component_id,
                },
            )
            if result.rowcount and result.rowcount > 0:
                stats["updated"] += 1
            else:
                # либо значение уже правильное, либо записи в supplier_prices
                # просто нет (не загружена). Это не ошибка — пропускаем тихо.
                stats["skipped"] += 1

        if not args.dry_run:
            session.commit()

        print("Готово.")
        print(f"  строк прайса рассмотрено:        {stats['considered']}")
        print(f"  компонент найден в БД:           {stats['matched']}")
        print(f"  компонент отсутствует:           {stats['absent']}")
        print(f"  supplier_sku {'будет обновлён' if args.dry_run else 'обновлён'}: {stats['updated']}")
        print(f"  пропущено (уже верно / нет записи): {stats['skipped']}")
        if args.dry_run:
            print("  [dry-run] изменения НЕ применены.")
        return 0
    finally:
        session.close()


if __name__ == "__main__":
    sys.exit(main())
