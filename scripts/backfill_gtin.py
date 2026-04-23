# Одноразовый скрипт: перезаполняет колонку gtin у существующих
# компонентов по данным из прайса OCS (колонка EAN128).
#
# Зачем: до этапа 7 GTIN в БД не хранился. На этапе 7 появилось
# сопоставление по GTIN — особенно критичное для Intel CPU (OCS хранит
# Order Code, Treolan — S-Spec, GTIN одинаков). Этот скрипт нужен,
# чтобы один раз подтянуть GTIN у 3 040 компонентов, уже загруженных
# с OCS.
#
# Что делает:
#   1. Открывает Excel OCS (--file путь).
#   2. Находит колонку EAN128 по заголовку (не по индексу — стабильнее
#      при смене формата прайса).
#   3. Для каждой строки определяет нашу таблицу через маппинг категорий
#      (колонка B/C) и пытается UPDATE <table> SET gtin = :gtin
#      WHERE sku = :mpn AND (gtin IS NULL OR gtin = '').
#
# Что НЕ делает:
#   - не создаёт новые компоненты;
#   - не трогает supplier_prices;
#   - не пишет в price_uploads;
#   - не переопределяет уже проставленные GTIN (считаем их «руками
#     админа» и более надёжными).
#
# Идемпотентен: повторный запуск ничего не ломает.

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import SessionLocal
from app.services.enrichment.base import ALLOWED_TABLES, CATEGORY_TO_TABLE
from app.services.price_loaders.ocs import (
    _COL_CAT_B, _COL_KIND_C, _COL_SKU,
    _find_ean_column, _normalize_gtin, _resolve_category,
)

logger = logging.getLogger(__name__)


def _run(filepath: str) -> dict:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = "Наличие и цены"
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Лист «{sheet_name}» не найден в файле {filepath}. "
            f"Доступные листы: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    ean_idx = _find_ean_column(header)
    if ean_idx is None:
        raise ValueError(
            "В заголовке прайса OCS нет колонки EAN128/EAN/GTIN — "
            "backfill невозможен."
        )

    counters = {
        "total_rows":        0,
        "no_ean_in_price":   0,
        "updated":           0,
        "already_had_gtin":  0,
        "not_found_in_db":   0,
        "errors":            0,
    }

    session = SessionLocal()
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            cat_b  = str(row[_COL_CAT_B]  or "").strip() if len(row) > _COL_CAT_B  else ""
            kind_c = str(row[_COL_KIND_C] or "").strip() if len(row) > _COL_KIND_C else ""
            mpn    = str(row[_COL_SKU]    or "").strip() if len(row) > _COL_SKU    else ""
            ean    = row[ean_idx] if len(row) > ean_idx else None

            our_category = _resolve_category(cat_b, kind_c)
            if our_category is None or not mpn:
                continue

            counters["total_rows"] += 1

            gtin = _normalize_gtin(ean)
            if not gtin:
                counters["no_ean_in_price"] += 1
                continue

            table = CATEGORY_TO_TABLE[our_category]
            assert table in ALLOWED_TABLES, f"Недопустимая таблица: {table}"

            try:
                # Смотрим, что с компонентом сейчас.
                res = session.execute(
                    text(f"SELECT id, gtin FROM {table} WHERE sku = :sku LIMIT 1"),
                    {"sku": mpn},
                ).first()
                if res is None:
                    counters["not_found_in_db"] += 1
                    continue
                if res.gtin:
                    counters["already_had_gtin"] += 1
                    continue
                session.execute(
                    text(f"UPDATE {table} SET gtin = :g WHERE id = :id"),
                    {"g": gtin, "id": res.id},
                )
                counters["updated"] += 1
            except Exception as exc:
                counters["errors"] += 1
                logger.error("Строка %d (sku=%r): %s", row_idx, mpn, exc)

        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()

    return counters


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(name)s  %(message)s",
    )

    parser = argparse.ArgumentParser(
        description="Backfill GTIN в таблицах компонентов по данным из прайса OCS."
    )
    parser.add_argument("--file", required=True, help="Путь к Excel-файлу прайса OCS.")
    args = parser.parse_args()

    if not Path(args.file).is_file():
        print(f"ОШИБКА: файл не найден: {args.file}", file=sys.stderr)
        return 2

    print(f"Backfill GTIN из прайса OCS: {args.file}")

    try:
        counters = _run(args.file)
    except ValueError as exc:
        print(f"ОШИБКА: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:
        logger.exception("Необработанное исключение при backfill GTIN")
        print(f"Критическая ошибка: {exc}", file=sys.stderr)
        return 1

    print()
    print("Готово.")
    print(f"  Всего строк из нужных категорий:  {counters['total_rows']}")
    print(f"  Заполнено GTIN:                    {counters['updated']}")
    print(f"  Уже был GTIN (не трогали):         {counters['already_had_gtin']}")
    print(f"  В прайсе нет EAN128:               {counters['no_ean_in_price']}")
    print(f"  Компонент в БД не найден по sku:   {counters['not_found_in_db']}")
    print(f"  Ошибок:                            {counters['errors']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
