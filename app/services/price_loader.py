# Загрузка прайс-листов поставщиков в таблицу supplier_prices.

import logging
import os
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import SessionLocal

logger = logging.getLogger(__name__)

# Маппинг (колонка B, колонка C) → (таблица БД, значение category).
# None в позиции C означает «без уточнения видом оборудования».
_CATEGORY_MAP = {
    ("Процессоры",                       None): ("cpus",         "cpu"),
    ("Материнские платы",                None): ("motherboards", "motherboard"),
    ("Оперативная память",               None): ("rams",         "ram"),
    ("Видеокарты",                       None): ("gpus",         "gpu"),
    ("Накопители информации", "Жёсткие диски"):  ("storages",    "storage"),
    ("Накопители информации", "Твердотельные накопители"): ("storages", "storage"),
    ("Корпуса",                          None): ("cases",        "case"),
    ("Блоки питания",                    None): ("psus",         "psu"),
    ("Системы охлаждения для ПК", "Воздушное охлаждение для процессоров"):
        ("coolers", "cooler"),
    ("Системы охлаждения для ПК",
     "Системы жидкостного охлаждения «всё-в-одном» для процессоров"):
        ("coolers", "cooler"),
}

# Белый список имён таблиц — исключает подстановку произвольных значений
_ALLOWED_TABLES = frozenset({"cpus", "motherboards", "rams", "gpus", "storages", "cases", "psus", "coolers"})

# Индексы нужных колонок (0-based) в прайсе OCS.
#
# ВАЖНО про sku vs supplier_sku:
#   - Колонка E («Номенклатурный номер», индекс 4) — внутренний код OCS,
#     например '1000659869'. Это supplier_sku: номер, по которому
#     менеджер оформляет заказ у этого конкретного поставщика.
#   - Колонка G («Каталожный номер», индекс 6) — партномер производителя,
#     например 'CM8071504650609' (Intel). Это sku: каталожный артикул,
#     общий для всех поставщиков и для базы компонентов.
#
# Раньше мы брали только колонку G и писали её И в cpus/gpus/...sku,
# И в supplier_prices.supplier_sku — в итоге supplier_sku дублировал sku.
_COL_CAT_A        = 0
_COL_CAT_B        = 1
_COL_KIND_C       = 2
_COL_MAKER        = 3
_COL_SUPPLIER_SKU = 4    # E: номенклатурный номер OCS → supplier_prices.supplier_sku
_COL_SKU          = 6    # G: каталожный номер производителя → <table>.sku
_COL_NAME         = 7
_COL_PRICE        = 8
_COL_CURRENCY     = 9
_COL_STOCK        = 11
_COL_TRANSIT      = 17
_MIN_COLS         = _COL_TRANSIT + 1   # строка должна содержать хотя бы 18 колонок


def _cell(row: tuple, idx: int):
    """Безопасное чтение ячейки: возвращает None, если строка короче."""
    return row[idx] if len(row) > idx else None


def _get_or_create_supplier(session) -> int:
    """Возвращает id поставщика OCS, создавая запись при необходимости."""
    row = session.execute(
        text("SELECT id FROM suppliers WHERE name = 'OCS' LIMIT 1")
    ).fetchone()
    if row:
        return row.id
    result = session.execute(
        text("INSERT INTO suppliers (name, is_active) VALUES ('OCS', TRUE) RETURNING id")
    ).fetchone()
    return result.id


def _resolve_mapping(cat_b: str, kind_c: str):
    """Определяет (table, category) по колонкам B и C прайса.

    Сначала ищет точное совпадение с уточнением вида, затем без него.
    Возвращает кортеж или None, если строка не относится к нужным категориям.
    """
    b = cat_b.strip() if cat_b else ""
    c = kind_c.strip() if kind_c else ""

    # Точное совпадение с видом оборудования
    hit = _CATEGORY_MAP.get((b, c if c else None))
    if hit:
        return hit

    # Попытка найти по одной категории без уточнения
    return _CATEGORY_MAP.get((b, None))


def _parse_price(value) -> Decimal | None:
    """Парсит цену из ячейки: терпит пробелы и запятые-разделители.

    Возвращает Decimal > 0 или None при ошибке.
    """
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    """Преобразует значение ячейки в целое. Пустота или ошибка → 0."""
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


def _find_or_create_component(
    session, table: str, name: str, manufacturer: str, sku: str
) -> tuple[int, bool]:
    """Ищет компонент по SKU; при отсутствии создаёт минимальную запись.

    Возвращает (component_id, is_new).
    Имя таблицы подставляется только из _ALLOWED_TABLES.
    """
    assert table in _ALLOWED_TABLES, f"Недопустимая таблица: {table}"

    row = session.execute(
        text(f"SELECT id FROM {table} WHERE sku = :sku LIMIT 1"),
        {"sku": sku},
    ).fetchone()
    if row:
        return row.id, False

    result = session.execute(
        text(
            f"INSERT INTO {table} (model, manufacturer, sku) "
            "VALUES (:model, :manufacturer, :sku) RETURNING id"
        ),
        {"model": name, "manufacturer": manufacturer, "sku": sku},
    ).fetchone()
    return result.id, True


def _upsert_price(
    session,
    supplier_id: int,
    category: str,
    component_id: int,
    supplier_sku: str | None,
    price: Decimal,
    currency: str,
    stock_qty: int,
    transit_qty: int,
) -> None:
    """Вставляет или обновляет строку цены через ON CONFLICT."""
    session.execute(
        text(
            "INSERT INTO supplier_prices "
            "    (supplier_id, category, component_id, supplier_sku, "
            "     price, currency, stock_qty, transit_qty, updated_at) "
            "VALUES "
            "    (:supplier_id, :category, :component_id, :supplier_sku, "
            "     :price, :currency, :stock_qty, :transit_qty, NOW()) "
            "ON CONFLICT (supplier_id, category, component_id) DO UPDATE SET "
            "    supplier_sku = EXCLUDED.supplier_sku, "
            "    price        = EXCLUDED.price, "
            "    currency     = EXCLUDED.currency, "
            "    stock_qty    = EXCLUDED.stock_qty, "
            "    transit_qty  = EXCLUDED.transit_qty, "
            "    updated_at   = NOW()"
        ),
        {
            "supplier_id":  supplier_id,
            "category":     category,
            "component_id": component_id,
            "supplier_sku": supplier_sku,
            "price":        price,
            "currency":     currency,
            "stock_qty":    stock_qty,
            "transit_qty":  transit_qty,
        },
    )


def _record_upload(session, supplier_id: int, filename: str, counters: dict) -> tuple[int, str]:
    """Записывает итоги загрузки в price_uploads.

    Возвращает (upload_id, status).
    """
    updated      = counters["updated"]
    added        = counters["added"]
    skipped      = counters["skipped"]
    errors       = counters["errors"]
    rows_matched = updated + added

    if rows_matched == 0 and (counters["processed"] > 0 or counters["total_rows"] > 0):
        status = "failed"
    elif errors > 0 or skipped > 0:
        status = "partial"
    else:
        status = "success"

    notes = f"updated={updated}, added={added}, skipped={skipped}, errors={errors}"

    result = session.execute(
        text(
            "INSERT INTO price_uploads "
            "    (supplier_id, filename, rows_total, rows_matched, rows_unmatched, status, notes) "
            "VALUES "
            "    (:supplier_id, :filename, :rows_total, :rows_matched, :rows_unmatched, :status, :notes) "
            "RETURNING id"
        ),
        {
            "supplier_id":    supplier_id,
            "filename":       filename,
            "rows_total":     counters["processed"],
            "rows_matched":   rows_matched,
            "rows_unmatched": skipped + errors,
            "status":         status,
            "notes":          notes,
        },
    ).fetchone()
    return result.id, status


def _save_failed_upload(filepath: str, counters: dict) -> None:
    """Фиксирует критический сбой в price_uploads отдельной сессией.

    Вызывается после rollback основной сессии, чтобы факт провала сохранился.
    """
    session = SessionLocal()
    try:
        row = session.execute(
            text("SELECT id FROM suppliers WHERE name = 'OCS' LIMIT 1")
        ).fetchone()
        if row is None:
            # Поставщик не был создан до сбоя — нечего записывать
            return
        session.execute(
            text(
                "INSERT INTO price_uploads "
                "    (supplier_id, filename, rows_total, rows_matched, rows_unmatched, status, notes) "
                "VALUES "
                "    (:supplier_id, :filename, :rows_total, 0, :rows_unmatched, 'failed', :notes)"
            ),
            {
                "supplier_id":    row.id,
                "filename":       os.path.basename(filepath),
                "rows_total":     counters.get("processed", 0),
                "rows_unmatched": counters.get("skipped", 0) + counters.get("errors", 0),
                "notes":          "Критическая ошибка при загрузке",
            },
        )
        session.commit()
    except Exception:
        # Ошибки здесь не должны затенять исходное исключение загрузки.
        # Логируем максимально безопасно — через repr, а не через форматирование.
        try:
            logger.error(
                "Не удалось сохранить запись о провальной загрузке. "
                "Служебная ошибка подавлена, чтобы не затенять исходное исключение."
            )
            session.rollback()
        except Exception:
            pass
    finally:
        session.close()


def load_ocs_price(filepath: str) -> dict:
    """Загружает прайс-лист OCS из Excel и обновляет таблицы компонентов и цен.

    Возвращает словарь со статистикой:
        total_rows, processed, updated, added, skipped, errors, status, upload_id.
    """
    # Открываем файл и проверяем наличие нужного листа
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = "Наличие и цены"
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Лист «{sheet_name}» не найден в файле {filepath}. "
            f"Доступные листы: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    counters = {
        "total_rows": 0,
        "processed":  0,
        "updated":    0,
        "added":      0,
        "skipped":    0,
        "errors":     0,
    }

    session = SessionLocal()
    try:
        supplier_id = _get_or_create_supplier(session)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем полностью пустые строки (типичный хвост листа)
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            cat_b        = str(_cell(row, _COL_CAT_B) or "").strip()
            kind_c       = str(_cell(row, _COL_KIND_C) or "").strip()
            manufacturer = str(_cell(row, _COL_MAKER) or "").strip()
            supplier_sku = str(_cell(row, _COL_SUPPLIER_SKU) or "").strip()  # E
            sku          = str(_cell(row, _COL_SKU) or "").strip()           # G
            name         = str(_cell(row, _COL_NAME) or "").strip()
            price_raw    = _cell(row, _COL_PRICE)
            currency_raw = _cell(row, _COL_CURRENCY)
            stock_raw    = _cell(row, _COL_STOCK)
            transit_raw  = _cell(row, _COL_TRANSIT)

            counters["total_rows"] += 1

            # Определяем целевую таблицу и категорию по маппингу
            mapping = _resolve_mapping(cat_b, kind_c)
            if mapping is None:
                counters["skipped"] += 1
                continue

            table, category = mapping

            # Каталожный номер производителя — ключ для сопоставления с таблицей
            # компонентов. Без него строка бесполезна.
            if not sku:
                counters["skipped"] += 1
                continue
            # Если OCS не указал свой номенклатурный номер — в supplier_sku
            # оставим NULL, чтобы не врать менеджеру.
            supplier_sku_or_none = supplier_sku or None

            counters["processed"] += 1

            # Парсим цену — строки с нечитаемой или нулевой ценой не обрабатываем
            price = _parse_price(price_raw)
            if price is None:
                logger.warning(
                    "Строка %d (sku=%r): не удалось разобрать цену (%r), строка пропущена",
                    row_idx, supplier_sku, price_raw,
                )
                counters["errors"] += 1
                continue

            # Валюта: обрезаем до 3 символов, чтобы гарантированно влезло в VARCHAR(3).
            # Пустое или некорректное значение трактуем как RUB.
            currency = str(currency_raw).strip().upper() if currency_raw else "RUB"
            if not currency:
                currency = "RUB"
            currency = currency[:3]

            stock_qty   = _parse_int(stock_raw)
            transit_qty = _parse_int(transit_raw)

            # SAVEPOINT на каждую строку: если здесь что-то упадёт,
            # откатим только эту строку, а не всю загрузку целиком.
            savepoint = session.begin_nested()
            try:
                # Компонент ищется/создаётся по каталожному номеру (sku, колонка G):
                # это глобальный артикул производителя, общий для всех поставщиков.
                component_id, is_new = _find_or_create_component(
                    session, table, name, manufacturer, sku
                )
                # В supplier_prices пишется номенклатурный номер OCS (supplier_sku,
                # колонка E) — именно по нему менеджер оформляет заказ у OCS.
                _upsert_price(
                    session, supplier_id, category, component_id,
                    supplier_sku_or_none, price, currency, stock_qty, transit_qty,
                )
                savepoint.commit()
                if is_new:
                    counters["added"] += 1
                else:
                    counters["updated"] += 1

            except Exception as exc:
                savepoint.rollback()
                logger.error(
                    "Строка %d (sku=%r, supplier_sku=%r): ошибка при записи в БД — %s",
                    row_idx, sku, supplier_sku, exc,
                )
                counters["errors"] += 1

        upload_id, status = _record_upload(
            session, supplier_id, os.path.basename(filepath), counters
        )
        session.commit()

    except Exception:
        session.rollback()
        # Сохраняем факт провала в отдельной сессии — основная уже откатилась
        _save_failed_upload(filepath, counters)
        session.close()
        raise

    session.close()

    result = {
        "total_rows": counters["total_rows"],
        "processed":  counters["processed"],
        "updated":    counters["updated"],
        "added":      counters["added"],
        "skipped":    counters["skipped"],
        "errors":     counters["errors"],
        "status":     status,
        "upload_id":  upload_id,
    }

    # Авто-хук обогащения новых SKU через OpenAI Web Search (этап 2.5В).
    # Включается флагом OPENAI_ENRICH_AUTO_HOOK в .env. Сам хук не бросает
    # исключения наружу и не влияет на результат загрузки прайса.
    try:
        from app.services.enrichment.openai_search.hooks import auto_enrich_new_skus
        auto_enrich_new_skus(added_new=counters["added"])
    except Exception as exc:
        logger.warning("auto-hook OpenAI пропущен из-за ошибки: %s", exc)

    return result
