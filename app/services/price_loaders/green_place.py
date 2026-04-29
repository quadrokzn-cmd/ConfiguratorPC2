# Адаптер «Green Place»: чтение прайс-листа «Price_GP_<номер>_<дата>.xlsx».
#
# Юридически Green Place связан с Merlion (тот же холдинг), но в нашей
# системе — отдельный supplier_id с собственным прайс-листом и (в
# перспективе) собственным контактным email.
#
# Особенности формата:
#   - Лист «Worksheet».
#   - Строка 1 — заголовки; данные начинаются со строки 2.
#   - Категории НЕ строками-разделителями, а тремя колонками
#     (как у Merlion): «Группа 1» → «Группа 2» → «Группа 3».
#     Полный путь идёт в raw_category, маппинг — по всей тройке.
#   - A=0 «Но»             — внутренний код Green Place (supplier_sku).
#   - B=1 «Наименование»   — name.
#   - C=2 «Бренд»          — brand.
#   - D=3 «PRT Номер»      — MPN.
#   - E=4 «Группа 1»       — раздел.
#   - F=5 «Группа 2»       — подраздел.
#   - G=6 «Группа 3»       — подгруппа (может быть пустой).
#   - H=7 «Доступно»       — основной остаток.
#   - I=8 «На складе»      — складской остаток (≈ дублирует «Доступно»;
#                             берём «Доступно» как stock).
#   - J=9 «В транзите, ближайшее» — текстовая отметка типа «04.05.2026 (200+)».
#   - K=10 «В транзите, всего»     — числовой транзит, его и берём.
#   - L=11 «Цена, USD».
#   - M=12 «Цена, РУБ».
#   - GTIN в прайсе нет.
#
# Цена: приоритет RUB (M), fallback — USD (L). Совпадает с
# OCS/Merlion/Treolan-логикой.
#
# По наблюдению на реальном прайсе (~1300 позиций) основная масса
# Green Place — серверное и сетевое железо. Для нашего конфигуратора
# реально полезны только две тройки категорий — потребительские CPU
# (включая «Оборудование для геймеров»). Остальные ветки оставлены
# не смаппленными (our_category=None) и orchestrator их пропустит,
# но ошибки не будет — позиция просто не попадёт в каталог.

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.services.price_loaders.base import BasePriceLoader
from app.services.price_loaders.models import PriceRow

logger = logging.getLogger(__name__)


# (Группа 1, Группа 2, Группа 3) → our_category.
# Группа 3 может быть пустой строкой — это допустимый случай в прайсе.
_CATEGORY_MAP: dict[tuple[str, str, str], str] = {
    ("Комплектующие для компьютеров", "Процессоры",  "Прочие"): "cpu",
    ("Оборудование для геймеров",     "Процессоры",  ""):       "cpu",
    # Память DDR* у Green Place промаркирована как «Server Memory»
    # внутри секции «Комплектующие для компьютеров» — выглядит как
    # серверная RAM, в каталог такие модули не подходят. Поэтому
    # отдельной строки для «Память оперативная / Server Memory» нет:
    # эти позиции уйдут в skipped, и так и должно быть.
}


_COL_SUPPLIER_SKU = 0   # A
_COL_NAME         = 1   # B
_COL_BRAND        = 2   # C
_COL_MPN          = 3   # D — PRT Номер
_COL_GROUP_1      = 4   # E
_COL_GROUP_2      = 5   # F
_COL_GROUP_3      = 6   # G
_COL_STOCK        = 7   # H — «Доступно»
_COL_TRANSIT      = 10  # K — «В транзите, всего»
_COL_PRICE_USD    = 11  # L
_COL_PRICE_RUB    = 12  # M

HEADER_ROW = 1
DATA_START_ROW = 2
_MAX_COL = 13


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


def _normalize(s) -> str:
    return (str(s).strip() if s is not None else "")


def _resolve_category(g1: str, g2: str, g3: str) -> str | None:
    return _CATEGORY_MAP.get((g1, g2, g3))


def _build_raw_path(g1: str, g2: str, g3: str) -> str:
    return " | ".join(x for x in (g1, g2, g3) if x)


class GreenPlaceLoader(BasePriceLoader):
    supplier_name = "Green Place"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "green_place" in name or "greenplace" in name or name.startswith("price_gp")

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            sheet_name = "Worksheet"
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Лист «{sheet_name}» не найден в файле {filepath}. "
                    f"Доступные листы: {wb.sheetnames}"
                )
            ws = wb[sheet_name]

            for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=DATA_START_ROW,
                    max_col=_MAX_COL,
                    values_only=True,
                ),
                start=DATA_START_ROW,
            ):
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue

                supplier_sku = _normalize(_cell(row, _COL_SUPPLIER_SKU))
                name         = _normalize(_cell(row, _COL_NAME))
                brand        = _normalize(_cell(row, _COL_BRAND)) or None
                mpn          = _normalize(_cell(row, _COL_MPN))   or None
                g1           = _normalize(_cell(row, _COL_GROUP_1))
                g2           = _normalize(_cell(row, _COL_GROUP_2))
                g3           = _normalize(_cell(row, _COL_GROUP_3))
                stock        = _parse_int(_cell(row, _COL_STOCK))
                transit      = _parse_int(_cell(row, _COL_TRANSIT))
                price_usd    = _parse_price(_cell(row, _COL_PRICE_USD))
                price_rub    = _parse_price(_cell(row, _COL_PRICE_RUB))

                if not supplier_sku:
                    continue

                if price_rub is not None:
                    price = price_rub
                    currency = "RUB"
                elif price_usd is not None:
                    price = price_usd
                    currency = "USD"
                else:
                    # Без цены позиция не пишется в supplier_prices,
                    # но и в unmapped тоже смысла нет — пропускаем.
                    continue

                our_category = _resolve_category(g1, g2, g3)

                yield PriceRow(
                    supplier_sku=supplier_sku,
                    mpn=mpn,
                    gtin=None,
                    brand=brand,
                    raw_category=_build_raw_path(g1, g2, g3),
                    our_category=our_category,
                    name=name,
                    price=price,
                    currency=currency,
                    stock=stock,
                    transit=transit,
                    row_number=row_idx,
                )
        finally:
            try:
                wb.close()
            except Exception:
                pass
