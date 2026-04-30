# Адаптер «Ресурс Медиа»: чтение прайс-листа «price_struct.xlsx».
#
# Особенности формата:
#   - Лист «Price».
#   - Строка 1 — служебная (дата формирования);
#     строка 2 — заголовки;
#     данные начинаются со строки 3.
#   - Иерархия категорий двухуровневая, реализована СТРОКАМИ-РАЗДЕЛИТЕЛЯМИ
#     (как у Treolan, не как у Merlion с тремя колонками):
#       * строка верхнего уровня:    заполнена ТОЛЬКО колонка A
#         («Аксессуары и периферия», «Жёсткие диски и оптические
#         носители», «Комплектующие и компоненты», ...);
#       * строка подкатегории:        заполнена ТОЛЬКО колонка B
#         («Адаптеры», «SSD диски», «Видеокарты», ...);
#       * строка данных:              заполнена колонка C (Артикул).
#     В строке данных колонка B содержит БРЕНД, а не подкатегорию —
#     поэтому отличаем по факту наличия Артикула в C.
#   - Заголовки (0-based):
#       A=0  «с/н» (служебный номер, у данных всегда пустой);
#       B=1  «Производитель» (в строках данных) или подкатегория-разделитель;
#       C=2  «Артикул» — внутренний SKU поставщика;
#       D=3  «Артикул производителя» — MPN;
#       E=4  «Номенклатура» — name;
#       F=5  «Объём, м3»;
#       G=6  «Вес, кг»;
#       H=7  «Цена, у.е.» — USD;
#       I=8  «Цена.руб»   — RUB;
#       J=9  «Доступно Москва» — основной остаток (число или
#             качественные маркеры «Мало»/«Средне»/«Много»/«Нет»);
#       K=10 «Факт Москва»     — качественный, дублирующий показатель;
#       L=11 «Ожидается Москва» — транзит (тоже качественный или пусто).
#   - Цена: приоритет RUB (I), fallback — USD (H). Совпадает с
#     Treolan/Merlion-логикой.
#   - GTIN в прайсе нет.
#
# Качественные маркеры остатка (Мало/Средне/Много/Нет) переводим в числа
# по аналогии с Treolan/Merlion — иначе фильтр `stock_qty > 0`
# в конфигураторе не увидел бы ни одной позиции.

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.services.price_loaders.base import BasePriceLoader
from app.services.price_loaders.models import PriceRow

logger = logging.getLogger(__name__)


# Пара (раздел, подраздел) → our_category. Разделы и подразделы — точно
# те, что встречаются в файле; всё остальное игнорируется.
_CATEGORY_MAP: dict[tuple[str, str], str] = {
    ("Жёсткие диски и оптические носители", "Внутренние жёсткие диски"): "storage",
    ("Комплектующие и компоненты", "SSD диски"):           "storage",
    ("Комплектующие и компоненты", "Видеокарты"):          "gpu",
    ("Комплектующие и компоненты", "Корпуса"):             "case",
    ("Комплектующие и компоненты", "Материнские платы"):   "motherboard",
    ("Комплектующие и компоненты", "Оперативная память"):  "ram",
    ("Комплектующие и компоненты", "Процессоры"):          "cpu",
    ("Комплектующие и компоненты", "Устройства охлаждения"): "cooler",
}


_COL_SECTION       = 0   # A — верхний раздел (только в строках-разделителях)
_COL_BRAND_OR_SUB  = 1   # B — подкатегория ИЛИ бренд (зависит от наличия Артикула)
_COL_SKU           = 2   # C — Артикул (supplier_sku)
_COL_MPN           = 3   # D — Артикул производителя
_COL_NAME          = 4   # E — Номенклатура
_COL_PRICE_USD     = 7   # H — Цена, у.е.
_COL_PRICE_RUB     = 8   # I — Цена.руб
_COL_STOCK         = 9   # J — Доступно Москва
_COL_TRANSIT       = 11  # L — Ожидается Москва

HEADER_ROW = 2
DATA_START_ROW = 3
_MAX_COL = 12


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


# Качественные маркеры остатка у Ресурс Медиа: «Мало», «Средне»,
# «Много», «Нет». Числа: ~5 / 20 / 100 / 0 — порядок величины из
# наблюдаемого распределения в реальном прайсе (~15k строк).
_RESURS_QUAL_STOCK: dict[str, int] = {
    "мало":   5,
    "средне": 20,
    "много":  100,
    "нет":    0,
}


def _parse_stock(value) -> int:
    if value is None:
        return 0
    s = str(value).strip().lower()
    if not s:
        return 0
    if s in _RESURS_QUAL_STOCK:
        return _RESURS_QUAL_STOCK[s]
    return _parse_int(value)


def _normalize(s) -> str:
    if s is None:
        return ""
    # Excel изредка отдаёт числовые SKU как float — нормализуем.
    if isinstance(s, float) and s.is_integer():
        return str(int(s))
    return str(s).strip()


def _is_data_row(row: tuple) -> bool:
    """Строка с данными определяется по непустому Артикулу (C)."""
    return bool(_normalize(_cell(row, _COL_SKU)))


class ResursMediaLoader(BasePriceLoader):
    supplier_name = "Ресурс Медиа"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return (
            "resurs" in name
            or "resurs_media" in name
            or "ресурс" in name
            or "медиа" in name
            or "price_struct" in name
        )

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            sheet_name = "Price"
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Лист «{sheet_name}» не найден в файле {filepath}. "
                    f"Доступные листы: {wb.sheetnames}"
                )
            ws = wb[sheet_name]

            current_section: str = ""
            current_subsection: str = ""

            for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=DATA_START_ROW,
                    max_col=_MAX_COL,
                    values_only=True,
                ),
                start=DATA_START_ROW,
            ):
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue

                a = _normalize(_cell(row, _COL_SECTION))
                b = _normalize(_cell(row, _COL_BRAND_OR_SUB))

                if not _is_data_row(row):
                    # Это строка-разделитель.
                    if a and not b:
                        current_section = a
                        # Новый раздел сбрасывает подкатегорию: до явной
                        # строки-подкатегории контекст некорректен.
                        current_subsection = ""
                    elif b and not a:
                        current_subsection = b
                    # Иначе — какой-то служебный мусор без c/d/e — игнорируем.
                    continue

                # Строка данных.
                supplier_sku = _normalize(_cell(row, _COL_SKU))
                mpn          = _normalize(_cell(row, _COL_MPN)) or None
                name         = _normalize(_cell(row, _COL_NAME))
                brand        = b or None
                price_usd    = _parse_price(_cell(row, _COL_PRICE_USD))
                price_rub    = _parse_price(_cell(row, _COL_PRICE_RUB))
                stock        = _parse_stock(_cell(row, _COL_STOCK))
                transit      = _parse_stock(_cell(row, _COL_TRANSIT))

                if not supplier_sku:
                    # Технически невозможно (мы попали сюда только из
                    # _is_data_row), но защищаемся.
                    continue

                # Цена: приоритет RUB, fallback USD.
                if price_rub is not None:
                    price = price_rub
                    currency = "RUB"
                elif price_usd is not None:
                    price = price_usd
                    currency = "USD"
                else:
                    continue

                our_category = _CATEGORY_MAP.get(
                    (current_section, current_subsection)
                )
                raw_category = " | ".join(
                    x for x in (current_section, current_subsection) if x
                )

                yield PriceRow(
                    supplier_sku=supplier_sku,
                    mpn=mpn,
                    gtin=None,
                    brand=brand,
                    raw_category=raw_category,
                    our_category=our_category,
                    name=name,
                    price=price,
                    currency=currency,
                    stock=stock,
                    transit=transit,
                    row_number=row_idx,
                )
        finally:
            try:
                wb.close()
            except Exception:
                pass
