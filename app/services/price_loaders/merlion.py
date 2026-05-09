# Адаптер Merlion: чтение прайса «Прайслист_Мерлион_Москва.xlsm».
#
# Особенности формата:
#   - Лист «Price List».
#   - Заголовки в строке 11; данные начинаются со строки 12.
#   - Иерархия категорий разнесена по трём колонкам: A «Группа 1»,
#     B «Группа 2», C «Группа 3». Полный путь строим как
#     «A | B | C» — он же идёт в raw_category.
#   - E «Номер»            — внутренний SKU Merlion (supplier_sku).
#   - G «Код производителя» — MPN → наш sku.
#   - H «Наименование»     — name.
#   - J «Цена»     — цена в USD.
#   - K «Цена(руб)» — цена в RUB.
#   - L «Доступно» — остаток, M «Ожидаемый приход» + N «На складе поставщика»
#     — транзит (для нас достаточно суммы).
#   - GTIN в прайсе Merlion Москва нет.
#
# Правило выбора цены/валюты: если есть RUB (K) — берём её, иначе
# берём USD (J). Это соответствует уже работающему OCS-подходу, где
# валюта из колонки J (currency) — один источник.

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Iterator

from openpyxl import load_workbook

from app.services.price_loaders.base import BasePriceLoader
from app.services.price_loaders.models import PriceRow

logger = logging.getLogger(__name__)


# Полный путь (Группа1 | Группа2 | Группа3) → our_category.
# Только то, что у нас есть в БД. Всё, что не в списке, пропускается.
_CATEGORY_MAP: dict[tuple[str, str, str], str] = {
    ("Комплектующие для компьютеров", "Материнские Платы", "Socket-1700"): "motherboard",
    ("Комплектующие для компьютеров", "Материнские Платы", "Socket-1851"): "motherboard",
    ("Комплектующие для компьютеров", "Материнские Платы", "Socket-AM4"):  "motherboard",
    ("Комплектующие для компьютеров", "Материнские Платы", "Socket-AM5"):  "motherboard",
    ("Комплектующие для компьютеров", "Память оперативная", "DDR3"):       "ram",
    ("Комплектующие для компьютеров", "Память оперативная", "DDR4"):       "ram",
    ("Комплектующие для компьютеров", "Память оперативная", "DDR5"):       "ram",
    ("Комплектующие для компьютеров", "Память оперативная", "SO-DIMM"):    "ram",
    ("Комплектующие для компьютеров", "Видеокарты", "PCI-E"):              "gpu",
    ("Оборудование для геймеров",     "Видеокарты", "Видеокарты"):         "gpu",
    ("Комплектующие для компьютеров", "Накопители SSD", "2.5\""):          "storage",
    ("Комплектующие для компьютеров", "Накопители SSD", "M.2"):            "storage",
    ("Комплектующие для компьютеров", "Жесткие Диски", "SATA"):            "storage",
    ("Комплектующие для компьютеров", "Корпуса", "ATX"):                   "case",
    ("Комплектующие для компьютеров", "Корпуса", "mATX"):                  "case",
    ("Комплектующие для компьютеров", "Корпуса", "Прочие"):                "case",
    ("Оборудование для геймеров",     "Корпуса", "Корпуса"):               "case",
    ("Комплектующие для компьютеров", "Блоки питания", "Блоки питания"):   "psu",
    ("Комплектующие для компьютеров", "Устройства охлаждения", "Все кулеры"):     "cooler",
    ("Комплектующие для компьютеров", "Устройства охлаждения", "Для INTEL"):      "cooler",
    ("Комплектующие для компьютеров", "Устройства охлаждения", "Универсальные"):  "cooler",
}


# Печатная техника Merlion (Этап 4 слияния, 2026-05-08).
# Если (g1, g2) попадает в _PRINTER_GROUPS — категоризуем по g3:
# {МФУ ...→ mfu, ...→ printer, прочее → ignore (явно)}.
# Запись printer/mfu в БД пока не подключена — orchestrator скипнет их
# с инкрементом счётчика pending_printers_mfu (Этап 6 даст таблицу
# `printers_mfu` и подключение в CATEGORY_TO_TABLE).
_PRINTER_GROUPS: set[tuple[str, str]] = {
    ("Периферия и аксессуары", "Принтеры"),
}

_G3_PRINTER_MFU_MAP: dict[str, str] = {
    "МФУ лазерные":      "mfu",
    "Лазерные":          "printer",
    "МФУ струйные":      "mfu",
    "Струйные":          "printer",
    # Явно отбрасываем подкатегории, которые нам неинтересны:
    # термо/матричные/мини-фото — не наш ассортимент.
    "Термопринтеры":     "ignore",
    "Мини-Фото-принтеры": "ignore",
    "Матричные":         "ignore",
    "":                  "ignore",
}


def _classify_merlion(g3: str) -> str:
    """Определяет печатную категорию Merlion по третьему уровню группы.

    Вызывается ТОЛЬКО когда (g1, g2) уже попало в _PRINTER_GROUPS;
    возвращает 'printer' / 'mfu' / 'ignore'. Неизвестный g3 → 'ignore'
    с INFO-логом для отладки прайса.
    """
    if g3 in _G3_PRINTER_MFU_MAP:
        return _G3_PRINTER_MFU_MAP[g3]
    logger.info("Merlion G3=%r: классифицирован как ignore (неизвестная подкатегория)", g3)
    return "ignore"


# Индексы колонок (0-based). Соответствуют буквам: A=0, B=1, ..., N=13.
_COL_GROUP_1   = 0   # A
_COL_GROUP_2   = 1   # B
_COL_GROUP_3   = 2   # C
_COL_BRAND     = 3   # D
_COL_NUMBER    = 4   # E  — Merlion SKU
_COL_MPN       = 6   # G  — «Код производителя»
_COL_NAME      = 7   # H
_COL_PRICE_USD = 9   # J
_COL_PRICE_RUB = 10  # K
_COL_STOCK     = 11  # L  — «Доступно»
_COL_TRANSIT_1 = 12  # M  — «Ожидаемый приход»
_COL_TRANSIT_2 = 13  # N  — «На складе поставщика»

HEADER_ROW = 11
DATA_START_ROW = 12


def _cell(row: tuple, idx: int):
    return row[idx] if len(row) > idx else None


def _parse_price(value) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def _parse_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


# Буквенные маркеры остатка у Merlion: в колонках L/M/N вместо чисел
# встречаются «+», «++», «+++», «++++» (градация «есть»→«очень много»)
# и «call» («позвоните, уточнить»). Пустая ячейка — товара нет.
# Без этого парсера все буквенные маркеры через _parse_int становились 0
# и конфигуратор не видел ни одной позиции Merlion (фильтр stock_qty > 0).
_MERLION_QUAL_STOCK: dict[str, int] = {
    "+":    5,
    "++":   15,
    "+++":  50,
    "++++": 100,
    # «call» консервативно трактуем как «точно нет» — менеджер должен
    # сам позвонить поставщику; в авто-подбор такие позиции не попадают.
    "call": 0,
}


def _parse_stock(value) -> int:
    """Остаток с учётом буквенных маркеров Merlion.

    Возвращает int >= 0. Если значение — число, возвращает его as-is.
    Если это один из буквенных маркеров — возвращает табличное значение.
    Иначе — 0 (пусто, неизвестный маркер).
    """
    if value is None:
        return 0
    s = str(value).strip().lower()
    if not s:
        return 0
    if s in _MERLION_QUAL_STOCK:
        return _MERLION_QUAL_STOCK[s]
    # Числовое значение (возможно, с запятой/пробелом).
    return _parse_int(value)


def _normalize(s) -> str:
    return (str(s).strip() if s is not None else "")


def _resolve_category(g1: str, g2: str, g3: str) -> str | None:
    """Сначала ищем точное соответствие в ПК-карте; если не нашли —
    проверяем, не печатная ли это группа (Этап 4 слияния).

    Возвращает: 'cpu'/'motherboard'/.../'storage' (ПК), либо
    'printer'/'mfu' (печать, обрабатывается orchestrator'ом как
    pending_printers_mfu до Этапа 6), либо None («не наша строка»).
    Подкатегории печати, явно отброшенные _classify_merlion как
    'ignore' (термо/матричные), для orchestrator'а отдаются как
    None — семантика «нам не интересно» в C-PC2 одна.
    """
    pc = _CATEGORY_MAP.get((g1, g2, g3))
    if pc is not None:
        return pc
    if (g1, g2) in _PRINTER_GROUPS:
        cat = _classify_merlion(g3)
        if cat in ("printer", "mfu"):
            return cat
    return None


def _build_raw_path(g1: str, g2: str, g3: str) -> str:
    return " | ".join(x for x in (g1, g2, g3) if x)


class MerlionLoader(BasePriceLoader):
    supplier_name = "Merlion"

    @classmethod
    def detect(cls, filename: str) -> bool:
        name = filename.lower()
        return "merlion" in name or "мерлион" in name

    def iter_rows(self, filepath: str) -> Iterator[PriceRow]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "Price List"
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист «{sheet_name}» не найден в файле {filepath}. "
                f"Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            g1 = _normalize(_cell(row, _COL_GROUP_1))
            g2 = _normalize(_cell(row, _COL_GROUP_2))
            g3 = _normalize(_cell(row, _COL_GROUP_3))
            brand        = _normalize(_cell(row, _COL_BRAND))   or None
            supplier_sku = _normalize(_cell(row, _COL_NUMBER))
            mpn          = _normalize(_cell(row, _COL_MPN))     or None
            name         = _normalize(_cell(row, _COL_NAME))
            price_usd    = _parse_price(_cell(row, _COL_PRICE_USD))
            price_rub    = _parse_price(_cell(row, _COL_PRICE_RUB))
            stock        = _parse_stock(_cell(row, _COL_STOCK))
            transit_1    = _parse_stock(_cell(row, _COL_TRANSIT_1))
            transit_2    = _parse_stock(_cell(row, _COL_TRANSIT_2))

            # Пустые строки-разделители категорий (если такие есть в самом
            # начале листа) — пропускаем.
            if not supplier_sku and not name:
                continue

            # У Merlion нет своего SKU — редкий, но возможный случай.
            # Тогда нет смысла вообще заводить строку: orchestrator не
            # сможет гарантировать идемпотентность загрузки.
            if not supplier_sku:
                logger.warning(
                    "Merlion строка %d: пустой «Номер» (supplier_sku) — строка пропущена.",
                    row_idx,
                )
                continue

            our_category = _resolve_category(g1, g2, g3)

            # Цена: приоритет RUB. Если нет — берём USD.
            if price_rub is not None:
                price = price_rub
                currency = "RUB"
            elif price_usd is not None:
                price = price_usd
                currency = "USD"
            else:
                # Без цены строка бесполезна.
                continue

            yield PriceRow(
                supplier_sku=supplier_sku,
                mpn=mpn,
                gtin=None,
                brand=brand,
                raw_category=_build_raw_path(g1, g2, g3),
                our_category=our_category,
                name=name,
                price=price,
                currency=currency,
                stock=stock,
                transit=transit_1 + transit_2,
                row_number=row_idx,
            )
