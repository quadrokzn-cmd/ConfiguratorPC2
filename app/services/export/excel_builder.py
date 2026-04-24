# Генератор Excel-выгрузки проекта по шаблону (этап 8.1).
#
# Берёт app/templates/export/project_template.xlsx как основу, сохраняет
# её структуру (заголовки, блок SUM/маржи, форматы валют) и заполняет
# реальными данными: шапкой проекта, курсом ЦБ, блоками конфигураций —
# по одной строке на системный блок («comp») и по строке на каждый
# компонент этой сборки.
#
# Продажные столбцы (F — цена руб. с НДС, M — конвертация, H — маржа)
# оставляем пустыми/нулевыми — их заполняет менеджер руками на своей
# стороне. Закупочная часть (N — $, 1 шт.; K — $ +конв; L — общ. $ +конв;
# I — руб закупки; J — общ. руб закупки; O — общ. $) — считаем формулами
# от фиксированного курса в $O$2 и конвертации M.

from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.routers.main_router import _CATEGORY_ORDER, _prepare_variants
from app.services import spec_service
from app.services.web_result_view import enrich_variants_with_specs


# Путь к шаблону — относительно корня проекта. Задаётся абсолютно,
# чтобы импорт работал из любой директории запуска (pytest, uvicorn).
_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[3]
    / "app" / "templates" / "export" / "project_template.xlsx"
)


# Соответствие «наша категория → таблица компонентов» — нужно для
# одного места: получение GTIN (бар-кода), которого нет в BuildResult.
_TABLE_BY_CAT = {
    "cpu":         "cpus",
    "motherboard": "motherboards",
    "ram":         "rams",
    "gpu":         "gpus",
    "storage":     "storages",
    "psu":         "psus",
    "case":        "cases",
    "cooler":      "coolers",
}


# ---------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------

_FONT = "Calibri"
_FONT_SIZE = 10

_HEADER_FILL = PatternFill("solid", start_color="FFE8F4F8")    # пастельный голубой — шапка проекта
_COL_HEADER_FILL = PatternFill("solid", start_color="FFD4E6F1")  # заголовки таблицы
_COMP_ROW_FILL = PatternFill("solid", start_color="FFE8F4F8")   # строка «Системный блок»

_THIN = Side(border_style="thin", color="FFBFBFBF")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_USD = '[$$-409]#,##0.00'
_FMT_RUB = '_-* #,##0.00"₽"_-;\\-* #,##0.00"₽"_-;_-* "-"??"₽"_-;_-@_-'
_FMT_PCT = '0.0%'

# Список колонок таблицы. L3 в шаблоне пустой — это не баг, это столбец
# «Общая стоимость $ +конв», считаемый формулой K*E; заголовок оставляем
# пустым, чтобы не расходиться с исходным шаблоном.
_COLS = list("ABCDEFGHIJKLMNO")


def _apply_font(cell, *, bold: bool = False) -> None:
    cell.font = Font(name=_FONT, size=_FONT_SIZE, bold=bold)


def _apply_align(cell, *, horizontal: str = "center", wrap: bool = True) -> None:
    cell.alignment = Alignment(
        horizontal=horizontal, vertical="center", wrap_text=wrap
    )


# ---------------------------------------------------------------------
# Подготовка данных
# ---------------------------------------------------------------------

def _load_project(db: Session, project_id: int) -> dict:
    row = db.execute(
        text(
            "SELECT p.id, p.name, p.created_at, "
            "       u.login AS author_login, u.name AS author_name "
            "FROM projects p JOIN users u ON u.id = p.user_id "
            "WHERE p.id = :pid"
        ),
        {"pid": project_id},
    ).first()
    if row is None:
        raise ValueError(f"Проект {project_id} не найден")
    return {
        "id":           int(row.id),
        "name":         row.name,
        "created_at":   row.created_at,
        "author_login": row.author_login,
        "author_name":  row.author_name,
    }


def _collect_blocks(
    db: Session,
    spec_items: list[dict],
) -> list[tuple[dict, dict | None, list[dict]]]:
    """На каждый spec_item возвращает (item, variant_dict, [компоненты]).

    Порядок компонентов — по _CATEGORY_ORDER; storage может быть несколько,
    все они идут подряд в месте категории 'storage'.
    """
    if not spec_items:
        return []

    query_ids = sorted({int(it["query_id"]) for it in spec_items})
    rows = db.execute(
        text("SELECT id, build_result_json FROM queries WHERE id = ANY(:ids)"),
        {"ids": query_ids},
    ).all()
    build_results = {int(r.id): r.build_result_json for r in rows}

    blocks: list[tuple[dict, dict | None, list[dict]]] = []
    # Enrich делаем пакетно: собираем все целевые варианты в один список.
    targets: list[dict] = []
    per_item_target: list[dict | None] = []

    for item in spec_items:
        br = build_results.get(int(item["query_id"]))
        variants = _prepare_variants(br) if br else []
        mfg = (item.get("variant_manufacturer") or "").lower()
        target = next(
            (v for v in variants
             if (v.get("manufacturer") or "").lower() == mfg),
            None,
        )
        per_item_target.append(target)
        if target is not None:
            targets.append(target)

    if targets:
        enrich_variants_with_specs(targets, db)

    for item, target in zip(spec_items, per_item_target):
        if target is None:
            blocks.append((item, None, []))
            continue
        comps_by_cat = target.get("components") or {}
        storages = target.get("storages_list") or []
        ordered: list[dict] = []
        for cat in _CATEGORY_ORDER:
            if cat == "storage":
                # У конфигурации может быть несколько накопителей.
                ordered.extend(storages)
            else:
                c = comps_by_cat.get(cat)
                if c:
                    ordered.append(c)
        # Если в categories_order не оказалось storage (редко, но
        # защитимся), добавим storages отдельно в конец.
        if not storages and comps_by_cat.get("storage"):
            ordered.append(comps_by_cat["storage"])
        blocks.append((item, target, ordered))

    return blocks


def _fetch_gtin_map(
    db: Session,
    blocks: list[tuple[dict, dict | None, list[dict]]],
) -> dict[tuple[str, int], str | None]:
    """Пакетный SELECT gtin из таблиц компонентов, по одному на категорию."""
    ids_by_cat: dict[str, set[int]] = defaultdict(set)
    for _item, _variant, comps in blocks:
        for c in comps:
            cat = c.get("category")
            cid = c.get("component_id")
            if cat in _TABLE_BY_CAT and cid is not None:
                ids_by_cat[cat].add(int(cid))

    out: dict[tuple[str, int], str | None] = {}
    for cat, ids in ids_by_cat.items():
        table = _TABLE_BY_CAT[cat]
        rows = db.execute(
            text(f"SELECT id, gtin FROM {table} WHERE id = ANY(:ids)"),
            {"ids": list(ids)},
        ).all()
        for r in rows:
            out[(cat, int(r.id))] = r.gtin
    return out


# ---------------------------------------------------------------------
# Заполнение ячеек
# ---------------------------------------------------------------------

def _row_style(ws, row: int, *, fill: PatternFill | None, bold: bool) -> None:
    """Общее оформление всей строки данных: шрифт/граница/заливка/высота."""
    for col in _COLS:
        cell = ws[f"{col}{row}"]
        _apply_font(cell, bold=bold)
        cell.border = _CELL_BORDER
        if fill is not None:
            cell.fill = fill
    ws.row_dimensions[row].height = 22

    # Выравнивания по колонкам
    _apply_align(ws[f"A{row}"], horizontal="center")
    _apply_align(ws[f"B{row}"], horizontal="center")
    _apply_align(ws[f"C{row}"], horizontal="center")
    _apply_align(ws[f"D{row}"], horizontal="left", wrap=True)
    for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
        _apply_align(ws[f"{col}{row}"], horizontal="right")


def _write_data_row(
    ws,
    row: int,
    *,
    pos_num: int | None,
    gtin: str | None,
    sku: str | None,
    name: str,
    qty: int,
    price_usd: float,
) -> None:
    """Пишет одну строку таблицы — comp или компонент, без заливки."""
    ws[f"A{row}"] = pos_num if pos_num is not None else None
    ws[f"B{row}"] = gtin or ""
    ws[f"C{row}"] = sku or ""
    ws[f"D{row}"] = name
    ws[f"E{row}"] = qty
    # Продажная часть (F, G, H, J) остаётся пустой/формульной — менеджер
    # сам заполняет цены продажи, а формулы (G=F*E, H=F/I-1, J=E*I)
    # продолжают работать, если он вобьёт числа.
    ws[f"G{row}"] = f"=F{row}*E{row}"
    ws[f"H{row}"] = f"=IFERROR(F{row}/I{row}-1,0)"
    ws[f"H{row}"].number_format = _FMT_PCT
    # Закупка: N — цена $ за единицу, M — конвертация (% накрутки на $).
    ws[f"M{row}"] = 0
    ws[f"M{row}"].number_format = _FMT_PCT
    ws[f"N{row}"] = round(float(price_usd), 2)
    ws[f"N{row}"].number_format = _FMT_USD
    ws[f"K{row}"] = f"=N{row}*(1+M{row})"
    ws[f"K{row}"].number_format = _FMT_USD
    ws[f"L{row}"] = f"=K{row}*E{row}"
    ws[f"L{row}"].number_format = _FMT_USD
    ws[f"O{row}"] = f"=N{row}*E{row}"
    ws[f"O{row}"].number_format = _FMT_USD
    # Закупочная в рублях: $ с конвертацией × курс ЦБ.
    ws[f"I{row}"] = f"=K{row}*$O$2"
    ws[f"I{row}"].number_format = _FMT_RUB
    ws[f"J{row}"] = f"=E{row}*I{row}"
    ws[f"J{row}"].number_format = _FMT_RUB
    # Цену продажи оставляем пустой — менеджер заполнит.
    ws[f"F{row}"].number_format = '#,##0'


def _clear_template_leftovers(ws) -> None:
    """Удаляет значения/формулы шаблонных строк 4+, оставляя границы/формат.

    Нам важно стереть и старые SUM-формулы (строка 8), и блок маржи
    (I9/I11/J11/I12/J12), чтобы они не оказались между нашими данными
    и пересозданными итогами.
    """
    for row in ws.iter_rows(min_row=4, max_row=max(ws.max_row, 4)):
        for cell in row:
            cell.value = None


# ---------------------------------------------------------------------
# Публичная функция
# ---------------------------------------------------------------------

def build_project_xlsx(
    project_id: int,
    db: Session,
    rate: Decimal,
    rate_date: date,
) -> bytes:
    """Собирает xlsx-файл проекта. Возвращает его содержимое в bytes.

    rate/rate_date приходят снаружи (получены exchange_rate.get_usd_rate)
    — это упрощает тестирование и отделяет сетевую операцию от чистой
    генерации файла.
    """
    project = _load_project(db, project_id)
    spec_items = spec_service.list_spec_items(db, project_id=project_id)
    blocks = _collect_blocks(db, spec_items)
    gtin_map = _fetch_gtin_map(db, blocks)

    wb = load_workbook(str(_TEMPLATE_PATH))
    ws = wb.active

    _clear_template_leftovers(ws)

    # --- Шапка проекта (строки 1-2, левее курса) ---
    # Разъединяем на всякий случай, затем мерджим заново — некоторые
    # шаблоны могут иметь предыдущие merge.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row in (1, 2) and rng.min_col == 1:
            ws.unmerge_cells(str(rng))

    ws["A1"] = f"Проект: {project['name']}"
    ws["A2"] = f"Создан: {project['created_at'].strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells("A1:M1")
    ws.merge_cells("A2:M2")

    c = ws["A1"]
    c.font = Font(name="Segoe UI", size=13, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = _HEADER_FILL
    ws.row_dimensions[1].height = 26

    c = ws["A2"]
    c.font = Font(name="Segoe UI", size=10, bold=False, italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = _HEADER_FILL
    ws.row_dimensions[2].height = 20

    # --- Курс в N1-O2 (N1/N2 текст остаётся из шаблона) ---
    ws["O1"] = rate_date.strftime("%d.%m.%Y")
    ws["O2"] = float(rate)
    ws["O2"].number_format = '0.0000'
    for coord in ("N1", "N2", "O1", "O2"):
        _apply_font(ws[coord], bold=coord in ("N1", "N2"))
        _apply_align(ws[coord], horizontal="center", wrap=False)

    # --- Заголовки таблицы (row 3) ---
    for col in _COLS:
        cell = ws[f"{col}3"]
        _apply_font(cell, bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        cell.fill = _COL_HEADER_FILL
        cell.border = _CELL_BORDER
    ws.row_dimensions[3].height = 40

    # --- Данные ---
    current_row = 4
    pos_num = 0
    for item, variant, comps in blocks:
        pos_num += 1
        comp_name = item.get("display_name") or item.get("auto_name") or "Конфигурация"
        comp_qty = int(item.get("quantity") or 1)
        comp_unit_usd = float(item.get("unit_usd") or 0.0)

        _write_data_row(
            ws,
            current_row,
            pos_num=pos_num,
            gtin=None,
            sku=None,
            name=comp_name,
            qty=comp_qty,
            price_usd=comp_unit_usd,
        )
        _row_style(ws, current_row, fill=_COMP_ROW_FILL, bold=True)
        current_row += 1

        for comp in comps:
            cat = comp.get("category")
            cid = comp.get("component_id")
            gtin = gtin_map.get((cat, int(cid))) if cat and cid is not None else None
            brand = (comp.get("manufacturer") or "").strip()
            model = (comp.get("model") or "").strip()
            short = (comp.get("specs_short") or "").strip()
            title = f"{brand} {model}".strip() if brand else model
            if short:
                title = f"{title} · {short}" if title else short
            if not title:
                title = "—"

            _write_data_row(
                ws,
                current_row,
                pos_num=None,
                gtin=gtin,
                sku=comp.get("sku"),
                name=title,
                qty=int(comp.get("quantity") or 1),
                price_usd=float(comp.get("price_usd") or 0.0),
            )
            _row_style(ws, current_row, fill=None, bold=False)
            current_row += 1

    last_data_row = current_row - 1

    # --- SUM и маржа (если данных не было — блок пустой) ---
    if last_data_row >= 4:
        sum_row = last_data_row + 2
        pct_row = sum_row + 2
        abs_row = pct_row + 1

        ws[f"F{sum_row}"] = "Итого:"
        _apply_font(ws[f"F{sum_row}"], bold=True)
        _apply_align(ws[f"F{sum_row}"], horizontal="right")

        ws[f"G{sum_row}"] = f"=SUM(G4:G{last_data_row})"
        ws[f"G{sum_row}"].number_format = _FMT_RUB
        _apply_font(ws[f"G{sum_row}"], bold=True)
        _apply_align(ws[f"G{sum_row}"], horizontal="right")

        ws[f"I{sum_row}"] = "Закупка ₽:"
        _apply_font(ws[f"I{sum_row}"], bold=True)
        _apply_align(ws[f"I{sum_row}"], horizontal="right")

        ws[f"J{sum_row}"] = f"=SUM(J4:J{last_data_row})"
        ws[f"J{sum_row}"].number_format = _FMT_RUB
        _apply_font(ws[f"J{sum_row}"], bold=True)
        _apply_align(ws[f"J{sum_row}"], horizontal="right")

        # Блок маржи
        ws[f"I{pct_row}"] = "Маржа %"
        _apply_font(ws[f"I{pct_row}"], bold=True)
        _apply_align(ws[f"I{pct_row}"], horizontal="right")
        ws[f"J{pct_row}"] = f"=IFERROR(G{sum_row}/J{sum_row}-1,0)"
        ws[f"J{pct_row}"].number_format = _FMT_PCT
        _apply_font(ws[f"J{pct_row}"], bold=True)
        _apply_align(ws[f"J{pct_row}"], horizontal="right")

        ws[f"I{abs_row}"] = "Маржа ₽"
        _apply_font(ws[f"I{abs_row}"], bold=True)
        _apply_align(ws[f"I{abs_row}"], horizontal="right")
        ws[f"J{abs_row}"] = f"=G{sum_row}-J{sum_row}"
        ws[f"J{abs_row}"].number_format = _FMT_RUB
        _apply_font(ws[f"J{abs_row}"], bold=True)
        _apply_align(ws[f"J{abs_row}"], horizontal="right")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
